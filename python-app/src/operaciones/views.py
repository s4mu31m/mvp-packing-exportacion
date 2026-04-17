"""
Vistas web del flujo operativo de packing.
Cada vista corresponde a una etapa del flujo.

Flujo de recepcion (preferido del MVP):
  RecepcionView — tres acciones: iniciar_lote / agregar_bin / cerrar_lote
  El lote activo se persiste en sesion como 'recepcion_lote_code'.

"""
import csv
import datetime
import json
import logging
import os
import tempfile
import threading
from io import BytesIO
from pathlib import Path
from urllib.parse import urlencode

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin

from usuarios.permissions import get_roles, is_admin


class RolRequiredMixin(UserPassesTestMixin):
    """Protege vistas por rol de negocio. Declarar roles_requeridos en subclase."""
    roles_requeridos: list[str] = []

    def test_func(self):
        if is_admin(self.request):
            return True
        return any(r in get_roles(self.request) for r in self.roles_requeridos)
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse, reverse_lazy
from django.views.generic import TemplateView, View
from django.utils import timezone

from operaciones.forms import (
    IniciarLoteForm,
    CerrarLoteForm,
    PesajeParcialCierreForm,
    PesajeParcialIngresoPackingForm,
    BinForm,
    EditBinVariableForm,
    CamaraMantencionForm,
    DesverdizadoForm,
    IngresoPackingForm,
    RegistroPackingForm,
    ControlProcesoPackingForm,
    CalidadPalletForm,
    CalidadPalletMuestraForm,
    CamaraFrioForm,
    MedicionTemperaturaForm,
    PlanillaDesverdizadoCalibreForm,
    PlanillaDesverdizadoSemillasForm,
    PlanillaCalidadPackingForm,
    PlanillaCalidadCamaraForm,
)
from operaciones.application.use_cases import (
    iniciar_lote_recepcion,
    agregar_bin_a_lote_abierto,
    cerrar_lote_recepcion,
    registrar_camara_mantencion,
    registrar_desverdizado,
    registrar_ingreso_packing,
    registrar_registro_packing,
    registrar_control_proceso_packing,
    registrar_calidad_pallet,
    cerrar_pallet,
    registrar_camara_frio,
    registrar_medicion_temperatura,
    guardar_muestra_calidad_pallet,
    registrar_planilla_desv_calibre,
    registrar_planilla_desv_semillas,
    registrar_planilla_calidad_packing,
    registrar_planilla_calidad_camara,
    eliminar_bin_de_lote,
    editar_bin_de_lote,
)
from operaciones.models import (
    Lote,
    LotePlantaEstado,
    Pallet,
    RegistroEtapa,
    RegistroPacking,
)

def _temporada(request) -> str:
    """Devuelve la temporada activa: del POST, sesion o año actual."""
    return (
        request.POST.get("temporada")
        or request.session.get("temporada_activa")
        or str(datetime.date.today().year)
    )


def _serialize_form_value(value):
    """
    Normaliza cleaned_data para payloads de vistas web.

    Los TimeField de Django salen como HH:mm:ss, pero los validadores de la
    aplicacion esperan HH:mm.
    """
    if isinstance(value, datetime.datetime):
        return value.isoformat()
    if isinstance(value, datetime.time):
        return value.strftime("%H:%M")
    if isinstance(value, datetime.date):
        return value.isoformat()
    return value


def _etapa_lote(lote: Lote) -> str:
    """
    Determina la etapa actual de un lote revisando qué registros existen.
    Usado para consulta jefatura y dashboard.
    """
    try:
        if hasattr(lote, "pallet_lotes") and lote.pallet_lotes.exists():
            pl = lote.pallet_lotes.first()
            if pl and hasattr(pl.pallet, "camara_frio"):
                return "Camara Frio"
            return "Pallet Cerrado"
    except Exception:
        pass
    try:
        if lote.registros_packing.exists() or lote.control_proceso_packing.exists():
            return "Packing / Proceso"
    except Exception:
        pass
    try:
        if hasattr(lote, "ingreso_packing"):
            return "Ingreso Packing"
    except Exception:
        pass
    try:
        if hasattr(lote, "desverdizado"):
            return "Desverdizado"
    except Exception:
        pass
    try:
        if hasattr(lote, "camara_mantencion"):
            return "Mantencion"
    except Exception:
        pass
    if lote.estado == LotePlantaEstado.CERRADO:
        if lote.requiere_desverdizado:
            return "Pendiente Desv."
        return "Pesado / Pendiente"
    if lote.estado == LotePlantaEstado.ABIERTO:
        return "Recepcion"
    return lote.get_estado_display()


def _kilos_recientes_lote(lote, *, include_bin_fallback: bool = True) -> tuple:
    """
    Retorna (kilos_bruto, kilos_neto) usando el pesaje mas reciente disponible.
    Prioridad: ingreso_packing > desverdizado salida > conformacion > suma bins.
    Usa reverse one-to-one accessors de Django ORM (lote.desverdizado, lote.ingreso_packing).
    """
    kb = float(lote.kilos_bruto_conformacion) if lote.kilos_bruto_conformacion is not None else None
    kn = float(lote.kilos_neto_conformacion) if lote.kilos_neto_conformacion is not None else None
    try:
        desv = lote.desverdizado
        if desv.kilos_bruto_salida is not None:
            kb = float(desv.kilos_bruto_salida)
        if desv.kilos_neto_salida is not None:
            kn = float(desv.kilos_neto_salida)
    except Exception:
        pass
    try:
        ip = lote.ingreso_packing
        if ip.kilos_bruto_ingreso_packing is not None:
            kb = float(ip.kilos_bruto_ingreso_packing)
        if ip.kilos_neto_ingreso_packing is not None:
            kn = float(ip.kilos_neto_ingreso_packing)
    except Exception:
        pass
    if include_bin_fallback and (kb is None or kn is None):
        try:
            bin_lotes = list(lote.bin_lotes.select_related("bin").all())
            if kb is None:
                total = sum(
                    float(bl.bin.kilos_bruto_ingreso)
                    for bl in bin_lotes
                    if bl.bin.kilos_bruto_ingreso is not None
                )
                if total > 0:
                    kb = total
            if kn is None:
                total = sum(
                    float(bl.bin.kilos_neto_ingreso)
                    for bl in bin_lotes
                    if bl.bin.kilos_neto_ingreso is not None
                )
                if total > 0:
                    kn = total
        except Exception:
            pass
    return kb, kn


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = "operaciones/dashboard.html"
    login_url = reverse_lazy("usuarios:login")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["page_title"] = "Dashboard — Turno Actual"
        temporada = (
            self.request.session.get("temporada_activa")
            or str(datetime.date.today().year)
        )
        ctx["temporada"] = temporada

        from django.conf import settings
        backend = getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip()

        if backend == "dataverse":
            ctx["kpis"], ctx["lotes_activos"] = self._context_dataverse()
        else:
            ctx["kpis"], ctx["lotes_activos"] = self._context_sqlite(temporada)

        return ctx

    def _context_sqlite(self, temporada: str):
        """KPIs y lista de lotes desde SQLite ORM."""
        try:
            from operaciones.models import BinLote
            from django.db.models import Sum

            hoy = datetime.date.today()
            bins_hoy = BinLote.objects.filter(created_at__date=hoy).count()

            lotes_qs = Lote.objects.filter(temporada=temporada, is_active=True)

            lotes_desverdizado_qs = lotes_qs.filter(
                desverdizado__isnull=False,
                ingreso_packing__isnull=True,
            )
            lotes_packing_qs = lotes_qs.filter(
                ingreso_packing__isnull=False,
            )
            lotes_recepcion_qs = lotes_qs.filter(
                estado=LotePlantaEstado.ABIERTO,
            )

            def _kg(qs):
                return float(qs.aggregate(total=Sum("kilos_neto_conformacion"))["total"] or 0)

            lotes_activos_qs = lotes_qs.exclude(
                estado=LotePlantaEstado.FINALIZADO
            ).order_by("-created_at")[:20]

            lotes_activos = [
                {
                    "codigo": lote.lote_code,
                    "bins": lote.cantidad_bins,
                    "etapa": _etapa_lote(lote),
                    "estado": lote.estado,
                }
                for lote in lotes_activos_qs
            ]
            kpis = {
                "lotes_desverdizado": lotes_desverdizado_qs.count(),
                "lotes_packing":      lotes_packing_qs.count(),
                "total_lotes":        lotes_qs.count(),
                "bins_hoy":           bins_hoy,
                "kg_recepcion":       _kg(lotes_recepcion_qs),
                "kg_desverdizado":    _kg(lotes_desverdizado_qs),
                "kg_packing":         _kg(lotes_packing_qs),
            }
            return kpis, lotes_activos
        except Exception:
            return {}, []

    def _context_dataverse(self):
        """
        KPIs y lista de lotes desde Dataverse.

        Usa crf21_etapa_actual como fuente principal de etapa (disponible desde 2026-03-31).
        Para lotes sin etapa_actual, resolve_etapa_lote retorna "Recepcion" como fallback.
        bins_hoy filtra por createdon = hoy en Dataverse (OData date-range filter).
        """
        try:
            from infrastructure.repository_factory import get_repositories
            from infrastructure.dataverse.repositories import resolve_etapa_lote

            repos = get_repositories()
            lotes = repos.lotes.list_recent(limit=50)

            _PACKING_ETAPAS = {
                "Ingreso Packing", "Packing / Proceso", "Paletizado",
                "Calidad Pallet", "Camara Frio", "Temperatura Salida",
            }

            lotes_activos = []
            lotes_desverdizado = 0
            lotes_packing = 0
            lotes_recepcion = 0
            kg_recepcion = 0.0
            kg_desverdizado = 0.0
            kg_packing = 0.0

            for l in lotes:
                etapa = resolve_etapa_lote(l)
                kn = float(l.kilos_neto_conformacion or 0)
                if etapa in ("Recepcion", "Pesaje"):
                    lotes_recepcion += 1
                    kg_recepcion += kn
                elif etapa == "Desverdizado":
                    lotes_desverdizado += 1
                    kg_desverdizado += kn
                elif etapa in _PACKING_ETAPAS:
                    lotes_packing += 1
                    kg_packing += kn
                lotes_activos.append({
                    "codigo":         l.lote_code,
                    "bins":           l.cantidad_bins,
                    "etapa":          etapa,
                    "estado":         l.estado,
                    # Fallback: fecha_conformacion para registros sin crf21_ultimo_cambio_estado_at
                    "ultimo_cambio":  l.ultimo_cambio_estado_at or l.fecha_conformacion,
                })

            kpis = {
                "lotes_desverdizado": lotes_desverdizado,
                "lotes_packing":      lotes_packing,
                "total_lotes":        len(lotes),
                "bins_hoy":           repos.lotes.count_bins_today(),
                "kg_recepcion":       kg_recepcion,
                "kg_desverdizado":    kg_desverdizado,
                "kg_packing":         kg_packing,
            }
            return kpis, lotes_activos
        except Exception as exc:
            import logging
            logging.getLogger(__name__).warning("Dashboard Dataverse error: %s", exc)
            return {}, []


# ---------------------------------------------------------------------------
# Recepcion de bins — flujo de lote abierto
# ---------------------------------------------------------------------------

class RecepcionView(LoginRequiredMixin, RolRequiredMixin, TemplateView):
    roles_requeridos = ["Recepcion"]
    template_name = "operaciones/recepcion.html"
    login_url = reverse_lazy("usuarios:login")

    CAMPOS_BASE = ["codigo_productor", "tipo_cultivo", "variedad_fruta", "color", "fecha_cosecha"]

    def _lote_activo(self, request):
        """
        Retorna el lote activo desde sesion.

        En SQLite retorna la instancia ORM (Lote).
        En Dataverse retorna un LoteRecord (dataclass) obtenido via repositorio.
        En ambos casos el objeto expone .lote_code, .cantidad_bins, .estado.
        """
        lote_code = request.session.get("lote_activo_code")
        if not lote_code:
            return None
        temporada = _temporada(request)

        from django.conf import settings
        if getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip() == "dataverse":
            return self._lote_activo_dataverse(request, temporada, lote_code)

        # -- Backend SQLite: usa ORM Django --
        try:
            lote = Lote.objects.get(temporada=temporada, lote_code=lote_code)
            if lote.estado != LotePlantaEstado.ABIERTO:
                request.session.pop("lote_activo_code", None)
                request.session.pop("lote_activo_campos_base", None)
                return None
            return lote
        except Lote.DoesNotExist:
            request.session.pop("lote_activo_code", None)
            request.session.pop("lote_activo_campos_base", None)
            return None

    def _lote_activo_dataverse(self, request, temporada, lote_code):
        """
        Resuelve el lote activo consultando Dataverse via repositorio.
        Usa etapa_actual para validar que el lote siga en 'Recepcion'.

        Politica de sesion:
          - Si el lote NO se encuentra (None): NO se limpia la sesion.
            Puede ser una inconsistencia temporal de Dataverse (el lote fue
            recien creado y aun no es queryable via OData). La sesion se
            preserva para que el proximo request intente de nuevo.
          - Si el lote se encuentra pero etapa_actual != 'Recepcion': se limpia
            la sesion, porque tenemos evidencia positiva de que el lote ya no
            corresponde a la etapa de recepcion.
        """
        try:
            from infrastructure.repository_factory import get_repositories
            from infrastructure.dataverse.repositories import resolve_etapa_lote
            repos = get_repositories()
            lote = repos.lotes.find_by_code(temporada, lote_code)
            if not lote:
                # No destruir la sesion — puede ser eventual consistency de Dataverse.
                # El lote fue recien creado y la cache fue precalentada en create(),
                # pero en caso de fallo de cache este path actua como resguardo conservador.
                return None
            # Validar que el lote siga en etapa Recepcion
            etapa = resolve_etapa_lote(lote)
            if etapa and etapa != "Recepcion":
                request.session.pop("lote_activo_code", None)
                request.session.pop("lote_activo_campos_base", None)
                return None
            return lote
        except Exception as exc:
            import logging
            logging.getLogger(__name__).warning(
                "RecepcionView._lote_activo_dataverse error para %s: %s", lote_code, exc
            )
            return None

    def _bins_de_lote(self, lote):
        """
        Retorna los bins del lote.
        En SQLite usa ORM. En Dataverse usa repos.bins.list_by_lote.
        """
        from django.conf import settings
        if getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip() == "dataverse":
            try:
                from infrastructure.repository_factory import get_repositories
                repos = get_repositories()
                return repos.bins.list_by_lote(lote.id)
            except Exception as exc:
                import logging
                logging.getLogger(__name__).warning(
                    "RecepcionView._bins_de_lote dataverse error: %s", exc
                )
                return []
        return [bl.bin for bl in lote.bin_lotes.select_related("bin").order_by("created_at")]

    def get(self, request, *args, **kwargs):
        lote = self._lote_activo(request)
        campos_base = request.session.get("lote_activo_campos_base", {})

        bins = []
        total_peso_neto = 0
        form_bin = BinForm(initial=campos_base) if campos_base else BinForm()

        if lote:
            bins = self._bins_de_lote(lote)
            total_peso_neto = sum((b.kilos_neto_ingreso or 0) for b in bins)
        lote_campos_unicos = _unique_lote_field_values(bins) if bins else {}
        lote_campos_unicos_display = [
            {
                "label": label,
                "values": ", ".join(lote_campos_unicos.get(key, [])),
            }
            for label, key in _MULTI_VALUE_FIELD_LABELS
            if lote_campos_unicos.get(key)
        ]

        # Pesajes parciales de cierre (acumulados en sesion)
        pesajes_cierre = request.session.get("pesajes_cierre", [])
        bins_pesados = sum(p["cantidad_bins"] for p in pesajes_cierre)
        bins_pendientes = (lote.cantidad_bins - bins_pesados) if lote else 0

        from decimal import Decimal as D
        neto_acumulado = sum(D(str(p["kilos_netos"])) for p in pesajes_cierre)
        bruto_acumulado = sum(D(str(p["kilos_brutos"])) for p in pesajes_cierre)

        ctx = {
            "page_title": "Recepcion de Bins",
            "lote": lote,
            "bins_del_lote": bins,
            "total_peso_neto": total_peso_neto,
            "campos_base": campos_base,
            "form_iniciar": IniciarLoteForm(),
            "form_bin": form_bin,
            "form_cerrar": CerrarLoteForm(),
            "form_pesaje_parcial": PesajeParcialCierreForm(),
            "pesajes_cierre": pesajes_cierre,
            "bins_pesados": bins_pesados,
            "bins_pendientes": bins_pendientes,
            "neto_acumulado": neto_acumulado,
            "bruto_acumulado": bruto_acumulado,
            "campos_base_keys": self.CAMPOS_BASE,
            "lote_campos_unicos": lote_campos_unicos,
            "lote_campos_unicos_display": lote_campos_unicos_display,
        }
        return render(request, self.template_name, ctx)

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action", "")
        if action == "iniciar":
            return self._handle_iniciar(request)
        if action == "agregar_bin":
            return self._handle_agregar_bin(request)
        if action == "agregar_pesaje_cierre":
            return self._handle_agregar_pesaje_cierre(request)
        if action == "eliminar_pesaje_cierre":
            return self._handle_eliminar_pesaje_cierre(request)
        if action == "cerrar":
            return self._handle_cerrar(request)
        if action == "eliminar_bin":
            return self._handle_eliminar_bin(request)
        if action == "editar_bin":
            return self._handle_editar_bin(request)
        messages.error(request, "Accion desconocida.")
        return redirect("operaciones:recepcion")

    def _handle_iniciar(self, request):
        # Guard: si ya existe un lote activo en Dataverse, no crear otro.
        # Evita generar lotes vacíos en cadena cuando el usuario pulsa "Iniciar"
        # repetidamente mientras el backend aún está procesando el primero.
        existing_lote = self._lote_activo(request)
        if existing_lote:
            messages.warning(
                request,
                f"Ya existe un lote activo ({existing_lote.lote_code}). "
                "Complete o cierre el lote actual antes de iniciar uno nuevo.",
            )
            return redirect("operaciones:recepcion")

        form = IniciarLoteForm(request.POST)
        if not form.is_valid():
            messages.error(request, "Formulario invalido para iniciar lote.")
            return redirect("operaciones:recepcion")
        cd = form.cleaned_data
        payload = {
            "temporada": _temporada(request),
            "operator_code": request.session.get("crf21_codigooperador", ""),
            "source_system": "web",
        }
        result = iniciar_lote_recepcion(payload)
        if result.ok:
            request.session["lote_activo_code"] = result.data["lote_code"]
            request.session.pop("lote_activo_campos_base", None)
            request.session.pop("pesajes_cierre", None)
            messages.success(
                request,
                f"Lote {result.data['lote_code']} iniciado. Agregue los bins.",
            )
        else:
            for err in result.errors:
                messages.error(request, err)
        return redirect("operaciones:recepcion")

    def _handle_agregar_bin(self, request):
        lote_code = request.session.get("lote_activo_code")
        if not lote_code:
            messages.error(request, "No hay lote abierto. Inicie un lote primero.")
            return redirect("operaciones:recepcion")

        form = BinForm(request.POST)
        if not form.is_valid():
            messages.error(request, "Formulario invalido.")
            return redirect("operaciones:recepcion")

        cd = form.cleaned_data
        campos_base = request.session.get("lote_activo_campos_base", {})

        # Validar consistencia con campos base si ya existe al menos un bin
        if campos_base:
            incompatibles = []
            mapeo = {
                "codigo_productor": "Codigo productor",
                "tipo_cultivo": "Tipo cultivo",
                "variedad_fruta": "Variedad",
                "color": "Color",
                "fecha_cosecha": "Fecha cosecha",
            }
            for campo, etiqueta in mapeo.items():
                val_nuevo = str(cd.get(campo) or "").strip()
                val_base = str(campos_base.get(campo, "")).strip()
                if val_base and val_nuevo and val_nuevo != val_base:
                    incompatibles.append(f"{etiqueta}: esperado '{val_base}', recibido '{val_nuevo}'")
            if incompatibles:
                messages.error(
                    request,
                    "Bin incompatible con el lote — " + " | ".join(incompatibles),
                )
                return redirect("operaciones:recepcion")

        payload = {
            "temporada": _temporada(request),
            "lote_code": lote_code,
            "operator_code": request.session.get("crf21_codigooperador", ""),
            "source_system": "web",
            "fecha_cosecha": _serialize_form_value(cd["fecha_cosecha"]) if cd.get("fecha_cosecha") else None,
            "variedad_fruta": cd.get("variedad_fruta") or "",
            "codigo_productor": cd.get("codigo_productor") or "",
            "tipo_cultivo": cd.get("tipo_cultivo") or "",
            "numero_cuartel": cd.get("numero_cuartel") or "",
            "nombre_cuartel": cd.get("nombre_cuartel") or "",
            "codigo_sag_csg": cd.get("codigo_sag_csg") or "",
            "codigo_sag_csp": cd.get("codigo_sag_csp") or "",
            "codigo_sdp": cd.get("codigo_sdp") or "",
            "lote_productor": cd.get("lote_productor") or "",
            "color": cd.get("color") or "",
            "hora_recepcion": _serialize_form_value(cd["hora_recepcion"]) if cd.get("hora_recepcion") else "",
            "kilos_bruto_ingreso": float(cd["kilos_bruto_ingreso"]) if cd.get("kilos_bruto_ingreso") else None,
            "kilos_neto_ingreso": float(cd["kilos_neto_ingreso"]) if cd.get("kilos_neto_ingreso") else None,
            # cantidad_bins_grupo y tara_bin ya no se ingresan por bin individual
            "a_o_r": cd.get("a_o_r") or None,
            "observaciones": cd.get("observaciones") or "",
        }

        result = agregar_bin_a_lote_abierto(payload)
        if result.ok:
            # Actualizar campos_base con los valores del bin — solo rellena campos vacios
            current = dict(campos_base)
            for campo in ("codigo_productor", "tipo_cultivo", "variedad_fruta", "color"):
                if not current.get(campo) and cd.get(campo):
                    current[campo] = cd[campo]
            if not current.get("fecha_cosecha") and cd.get("fecha_cosecha"):
                current["fecha_cosecha"] = _serialize_form_value(cd["fecha_cosecha"])
            request.session["lote_activo_campos_base"] = current
            _consulta_cache_invalidate()
            messages.success(
                request,
                f"Bin {result.data['bin_code']} agregado al lote {lote_code}.",
            )
        else:
            for err in result.errors:
                messages.error(request, err)
        return redirect("operaciones:recepcion")

    def _handle_agregar_pesaje_cierre(self, request):
        lote = self._lote_activo(request)
        if not lote:
            messages.error(request, "No hay lote abierto.")
            return redirect("operaciones:recepcion")

        form = PesajeParcialCierreForm(request.POST)
        if not form.is_valid():
            for errs in form.errors.values():
                for e in errs:
                    messages.error(request, e)
            return redirect("operaciones:recepcion")

        cd = form.cleaned_data
        cantidad = cd["cantidad_bins"]
        kilos_brutos = float(cd["kilos_brutos_grupo"])
        tara = float(cd["tara"])
        kilos_netos = float(cd["kilos_netos_grupo"])

        pesajes_cierre = list(request.session.get("pesajes_cierre", []))
        bins_pesados = sum(p["cantidad_bins"] for p in pesajes_cierre)
        bins_pendientes = lote.cantidad_bins - bins_pesados

        if cantidad > bins_pendientes:
            messages.error(
                request,
                f"No puede registrar {cantidad} bin(s): solo quedan {bins_pendientes} pendiente(s) de pesar.",
            )
            return redirect("operaciones:recepcion")

        pesajes_cierre.append({
            "cantidad_bins": cantidad,
            "kilos_brutos": kilos_brutos,
            "tara": tara,
            "kilos_netos": round(kilos_netos, 2),
        })
        request.session["pesajes_cierre"] = pesajes_cierre
        messages.success(
            request,
            f"Pesaje registrado: {cantidad} bin(s) — {kilos_netos:.2f} kg netos.",
        )
        return redirect("operaciones:recepcion")

    def _handle_eliminar_pesaje_cierre(self, request):
        try:
            idx = int(request.POST.get("pesaje_idx", -1))
        except (ValueError, TypeError):
            idx = -1

        pesajes_cierre = list(request.session.get("pesajes_cierre", []))
        if 0 <= idx < len(pesajes_cierre):
            pesajes_cierre.pop(idx)
            request.session["pesajes_cierre"] = pesajes_cierre
            messages.success(request, "Pesaje eliminado.")
        else:
            messages.error(request, "Pesaje no encontrado.")
        return redirect("operaciones:recepcion")

    def _handle_cerrar(self, request):
        lote_code = request.session.get("lote_activo_code")
        if not lote_code:
            messages.error(request, "No hay lote abierto para cerrar.")
            return redirect("operaciones:recepcion")

        lote = self._lote_activo(request)
        if not lote:
            messages.error(request, "No hay lote abierto.")
            return redirect("operaciones:recepcion")

        # Validar que todos los bins del lote hayan sido pesados
        pesajes_cierre = request.session.get("pesajes_cierre", [])
        bins_pesados = sum(p["cantidad_bins"] for p in pesajes_cierre)
        if bins_pesados != lote.cantidad_bins:
            pendientes = lote.cantidad_bins - bins_pesados
            messages.error(
                request,
                f"Quedan {pendientes} bin(s) pendientes de pesar. "
                "Complete el pesaje antes de cerrar el lote.",
            )
            return redirect("operaciones:recepcion")

        form = CerrarLoteForm(request.POST)
        if not form.is_valid():
            messages.error(request, "Formulario invalido para cerrar lote.")
            return redirect("operaciones:recepcion")

        cd = form.cleaned_data

        # Calcular totales desde pesajes parciales
        from decimal import Decimal as D
        kilos_bruto_total = sum(D(str(p["kilos_brutos"])) for p in pesajes_cierre)
        kilos_neto_total = sum(D(str(p["kilos_netos"])) for p in pesajes_cierre)

        payload = {
            "temporada": _temporada(request),
            "lote_code": lote_code,
            "operator_code": request.session.get("crf21_codigooperador", ""),
            "source_system": "web",
            "requiere_desverdizado": cd.get("requiere_desverdizado") or False,
            "kilos_bruto_conformacion": float(kilos_bruto_total),
            "kilos_neto_conformacion": float(kilos_neto_total),
        }
        result = cerrar_lote_recepcion(payload)
        if result.ok:
            request.session.pop("lote_activo_code", None)
            request.session.pop("lote_activo_campos_base", None)
            request.session.pop("pesajes_cierre", None)
            _consulta_cache_invalidate()
            messages.success(request, result.message)
        else:
            for err in result.errors:
                messages.error(request, err)
        return redirect("operaciones:recepcion")

    def _handle_eliminar_bin(self, request):
        lote_code = request.session.get("lote_activo_code")
        if not lote_code:
            messages.error(request, "No hay lote abierto.")
            return redirect("operaciones:recepcion")

        bin_code = request.POST.get("bin_code", "").strip()
        if not bin_code:
            messages.error(request, "Bin no identificado.")
            return redirect("operaciones:recepcion")

        result = eliminar_bin_de_lote({
            "temporada": _temporada(request),
            "lote_code": lote_code,
            "bin_code": bin_code,
        })
        if result.ok:
            if result.data.get("cantidad_bins_restantes", 1) == 0:
                request.session.pop("lote_activo_campos_base", None)
            messages.success(request, f"Bin {bin_code} eliminado del lote.")
        else:
            for err in result.errors:
                messages.error(request, err)
        return redirect("operaciones:recepcion")

    def _handle_editar_bin(self, request):
        lote_code = request.session.get("lote_activo_code")
        if not lote_code:
            messages.error(request, "No hay lote abierto.")
            return redirect("operaciones:recepcion")

        bin_code = request.POST.get("bin_code", "").strip()
        if not bin_code:
            messages.error(request, "Bin no identificado.")
            return redirect("operaciones:recepcion")

        form = EditBinVariableForm(request.POST)
        if not form.is_valid():
            for field_errors in form.errors.values():
                for err in field_errors:
                    messages.error(request, err)
            return redirect("operaciones:recepcion")

        cd = form.cleaned_data
        result = editar_bin_de_lote({
            "temporada": _temporada(request),
            "lote_code": lote_code,
            "bin_code": bin_code,
            "campos": {
                "numero_cuartel":      cd.get("numero_cuartel") or "",
                "hora_recepcion":      _serialize_form_value(cd["hora_recepcion"]) if cd.get("hora_recepcion") else "",
                "kilos_bruto_ingreso": float(cd["kilos_bruto_ingreso"]) if cd.get("kilos_bruto_ingreso") else None,
                "kilos_neto_ingreso":  float(cd["kilos_neto_ingreso"]) if cd.get("kilos_neto_ingreso") else None,
                "a_o_r":               cd.get("a_o_r") or None,
                "observaciones":       cd.get("observaciones") or "",
            },
        })
        if result.ok:
            messages.success(request, f"Bin {bin_code} actualizado correctamente.")
        else:
            for err in result.errors:
                messages.error(request, err)
        return redirect("operaciones:recepcion")


# ---------------------------------------------------------------------------
# Desverdizado (camara mantencion + desverdizado)
# ---------------------------------------------------------------------------

def _lotes_pendientes_desverdizado(temporada: str):
    """
    Retorna lotes cerrados que requieren desverdizado y aun no tienen
    registro de desverdizado. Se muestran en el selector de la vista.
    """
    from django.conf import settings
    if getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip() == "dataverse":
        try:
            from infrastructure.repository_factory import get_repositories
            repos = get_repositories()
            lotes = repos.lotes.list_recent(limit=200)
            return [
                l for l in lotes
                if l.is_active and l.requiere_desverdizado
                and l.etapa_actual in ("Pesaje", "Mantencion")
            ]
        except Exception:
            return []
    from operaciones.models import Desverdizado
    lotes_con_desv = set(
        Desverdizado.objects.values_list("lote_id", flat=True)
    )
    qs = (
        Lote.objects
        .filter(temporada=temporada, is_active=True, estado=LotePlantaEstado.CERRADO)
        .exclude(id__in=lotes_con_desv)
        .order_by("-created_at")
    )
    return list(qs)


class DesverdizadoView(LoginRequiredMixin, RolRequiredMixin, TemplateView):
    roles_requeridos = ["Desverdizado"]
    template_name = "operaciones/desverdizado.html"
    login_url = reverse_lazy("usuarios:login")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["page_title"] = "Desverdizado"
        ctx["form_mantencion"] = CamaraMantencionForm()
        ctx["form_desverdizado"] = DesverdizadoForm()
        temporada = (
            self.request.session.get("temporada_activa")
            or str(datetime.date.today().year)
        )
        lotes = _lotes_pendientes_desverdizado(temporada)
        ctx["lotes_pendientes"] = lotes
        from django.conf import settings
        if getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip() == "dataverse":
            from infrastructure.repository_factory import get_repositories
            ctx["lotes_data_json"] = _lotes_json_from_records(lotes, get_repositories())
        else:
            ctx["lotes_data_json"] = _lotes_data_json(temporada, lotes)
        return ctx

    def _lote_context(self, temporada: str, lote_code: str) -> dict:
        """Datos del lote para mostrar en el panel de contexto."""
        try:
            lote = Lote.objects.get(temporada=temporada, lote_code=lote_code)
            primer_bin = None
            bin_lotes = list(lote.bin_lotes.select_related("bin").order_by("created_at"))
            if bin_lotes:
                primer_bin = bin_lotes[0].bin
            _kb, _kn = _kilos_recientes_lote(lote)
            return {
                "lote_code": lote.lote_code,
                "cantidad_bins": lote.cantidad_bins,
                "kilos_bruto": _kb,
                "kilos_neto": _kn,
                "requiere_desverdizado": lote.requiere_desverdizado,
                "disponibilidad_camara": lote.disponibilidad_camara_desverdizado,
                "productor": primer_bin.codigo_productor if primer_bin else "",
                "variedad": primer_bin.variedad_fruta if primer_bin else "",
                "color": primer_bin.color if primer_bin else "",
                "fecha_cosecha": primer_bin.fecha_cosecha if primer_bin else None,
            }
        except Lote.DoesNotExist:
            return {}

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action", "mantencion")
        temporada = _temporada(request)
        lote_code = request.POST.get("lote_code", "").strip()

        if action == "mantencion":
            form = CamaraMantencionForm(request.POST)
            if form.is_valid():
                cd = form.cleaned_data
                payload = {
                    "temporada": temporada,
                    "lote_code": lote_code,
                    "operator_code": request.session.get("crf21_codigooperador", ""),
                    "source_system": "web",
                    "extra": {
                        "camara_numero": cd.get("camara_numero"),
                        "fecha_ingreso": _serialize_form_value(cd["fecha_ingreso"]) if cd.get("fecha_ingreso") else None,
                        "hora_ingreso": _serialize_form_value(cd["hora_ingreso"]) if cd.get("hora_ingreso") else None,
                        "temperatura_camara": float(cd["temperatura_camara"]) if cd.get("temperatura_camara") else None,
                        "humedad_relativa": float(cd["humedad_relativa"]) if cd.get("humedad_relativa") else None,
                        "observaciones": cd.get("observaciones"),
                    },
                }
                result = registrar_camara_mantencion(payload)
                _handle_result(request, result)
                if result.ok:
                    _consulta_cache_invalidate()
            else:
                messages.error(request, "Formulario de camara mantencion invalido.")

        elif action == "desverdizado":
            form = DesverdizadoForm(request.POST)
            if form.is_valid():
                cd = form.cleaned_data
                payload = {
                    "temporada": temporada,
                    "lote_code": lote_code,
                    "operator_code": request.session.get("crf21_codigooperador", ""),
                    "source_system": "web",
                    "disponibilidad_camara_desverdizado": cd.get("disponibilidad_camara_desverdizado") or None,
                    "extra": {
                        "numero_camara": cd.get("numero_camara"),
                        "fecha_ingreso": _serialize_form_value(cd["fecha_ingreso"]) if cd.get("fecha_ingreso") else None,
                        "hora_ingreso": _serialize_form_value(cd["hora_ingreso"]) if cd.get("hora_ingreso") else None,
                        "color_salida": cd.get("color") or "",
                        "horas_desverdizado": cd.get("horas_desverdizado"),
                    },
                }
                result = registrar_desverdizado(payload)
                _handle_result(request, result)
                if result.ok:
                    _consulta_cache_invalidate()
            else:
                messages.error(request, "Formulario de desverdizado invalido.")

        return redirect("operaciones:desverdizado")


# ---------------------------------------------------------------------------
# Ingreso a packing
# ---------------------------------------------------------------------------

def _lotes_pendientes_ingreso_packing(temporada: str):
    """
    Retorna lotes cerrados que aun no tienen registro de ingreso a packing.
    """
    from django.conf import settings
    if getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip() == "dataverse":
        try:
            from infrastructure.repository_factory import get_repositories
            repos = get_repositories()
            lotes = repos.lotes.list_recent(limit=200)
            _ANTES_INGRESO = ("Pesaje", "Mantencion", "Desverdizado")
            return [l for l in lotes if l.is_active and l.etapa_actual in _ANTES_INGRESO]
        except Exception:
            return []
    from operaciones.models import IngresoAPacking
    lotes_con_ingreso = set(
        IngresoAPacking.objects.values_list("lote_id", flat=True)
    )
    qs = (
        Lote.objects
        .filter(temporada=temporada, is_active=True, estado=LotePlantaEstado.CERRADO)
        .exclude(id__in=lotes_con_ingreso)
        .order_by("-created_at")
    )
    return list(qs)


def _lote_info(temporada: str, lote_code: str) -> dict:
    """
    Datos base de un lote para autocompletar formularios.
    Todos los valores son JSON-serializables (str, int, bool, None).
    """
    from django.conf import settings
    if getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip() == "dataverse":
        try:
            from infrastructure.repository_factory import get_repositories
            repos = get_repositories()
            lote = repos.lotes.find_by_code(temporada, lote_code)
            if not lote:
                return {}
            kilos_bruto = float(lote.kilos_bruto_conformacion) if lote.kilos_bruto_conformacion else None
            kilos_neto = float(lote.kilos_neto_conformacion) if lote.kilos_neto_conformacion else None
            via_desv = lote.etapa_actual in ("Desverdizado", "Mantencion") if lote.etapa_actual else False
            try:
                desv = repos.desverdizados.find_by_lote(lote.id)
                if desv:
                    if desv.kilos_bruto_salida is not None:
                        kilos_bruto = float(desv.kilos_bruto_salida)
                    if desv.kilos_neto_salida is not None:
                        kilos_neto = float(desv.kilos_neto_salida)
                    via_desv = True
            except Exception:
                pass
            try:
                ip = repos.ingresos_packing.find_by_lote(lote.id)
                if ip:
                    if ip.kilos_bruto_ingreso_packing is not None:
                        kilos_bruto = float(ip.kilos_bruto_ingreso_packing)
                    if ip.kilos_neto_ingreso_packing is not None:
                        kilos_neto = float(ip.kilos_neto_ingreso_packing)
            except Exception:
                pass
            # B4: obtener campos heredables desde TODOS los bins (fallback por campo)
            bins_dv: list = []
            try:
                bins_map = repos.bins.all_bins_by_lotes([lote.id])
                bins_dv = bins_map.get(lote.id) or []
            except Exception:
                pass
            merged_dv = _merge_bin_fields(bins_dv)
            return {
                "lote_code": lote.lote_code,
                "estado": lote.etapa_actual or lote.estado,
                "cantidad_bins": lote.cantidad_bins,
                "kilos_bruto": kilos_bruto,
                "kilos_neto": kilos_neto,
                "via_desverdizado": via_desv,
                "requiere_desverdizado": lote.requiere_desverdizado,
                "productor": merged_dv["codigo_productor"] or getattr(lote, "codigo_productor", "") or "",
                "variedad":     merged_dv["variedad_fruta"] or "",
                "color":        merged_dv["color"] or "",
                "fecha_cosecha": str(merged_dv["fecha_cosecha"]) if merged_dv["fecha_cosecha"] else "",
                "tipo_cultivo": merged_dv["tipo_cultivo"] or "",
            }
        except Exception:
            return {}
    try:
        lote = Lote.objects.get(temporada=temporada, lote_code=lote_code)
        bin_lotes = list(lote.bin_lotes.select_related("bin").order_by("created_at"))
        primer_bin = bin_lotes[0].bin if bin_lotes else None

        # Kilos heredables desde desverdizado si existe
        kilos_bruto = float(lote.kilos_bruto_conformacion) if lote.kilos_bruto_conformacion is not None else None
        kilos_neto  = float(lote.kilos_neto_conformacion)  if lote.kilos_neto_conformacion  is not None else None
        via_desv = False
        try:
            desv = lote.desverdizado
            if desv.kilos_bruto_salida is not None:
                kilos_bruto = float(desv.kilos_bruto_salida)
            if desv.kilos_neto_salida is not None:
                kilos_neto = float(desv.kilos_neto_salida)
            via_desv = True
        except Exception:
            pass
        try:
            lote.camara_mantencion
            via_desv = True
        except Exception:
            pass
        try:
            ip = lote.ingreso_packing
            if ip.kilos_bruto_ingreso_packing is not None:
                kilos_bruto = float(ip.kilos_bruto_ingreso_packing)
            if ip.kilos_neto_ingreso_packing is not None:
                kilos_neto = float(ip.kilos_neto_ingreso_packing)
        except Exception:
            pass

        # Fallback: suma de kilos de bins si no hay conformacion ni desverdizado
        if kilos_neto is None:
            total_bin_neto = sum(
                float(bl.bin.kilos_neto_ingreso)
                for bl in bin_lotes
                if bl.bin.kilos_neto_ingreso is not None
            )
            if total_bin_neto > 0:
                kilos_neto = total_bin_neto
        if kilos_bruto is None:
            total_bin_bruto = sum(
                float(bl.bin.kilos_bruto_ingreso)
                for bl in bin_lotes
                if bl.bin.kilos_bruto_ingreso is not None
            )
            if total_bin_bruto > 0:
                kilos_bruto = total_bin_bruto

        fecha_cosecha_str = ""
        if primer_bin and primer_bin.fecha_cosecha:
            fecha_cosecha_str = str(primer_bin.fecha_cosecha)

        # color: usar el primer bin que lo tenga
        color = next(
            (bl.bin.color for bl in bin_lotes if bl.bin.color),
            "",
        )

        return {
            "lote_code":           lote.lote_code,
            "estado":              lote.estado,
            "cantidad_bins":       lote.cantidad_bins,
            "kilos_bruto":         kilos_bruto,
            "kilos_neto":          kilos_neto,
            "via_desverdizado":    via_desv,
            "requiere_desverdizado": lote.requiere_desverdizado,
            "productor":    primer_bin.codigo_productor if primer_bin else "",
            "variedad":     primer_bin.variedad_fruta   if primer_bin else "",
            "color":        color,
            "fecha_cosecha": fecha_cosecha_str,
            "tipo_cultivo": primer_bin.tipo_cultivo     if primer_bin else "",
        }
    except Lote.DoesNotExist:
        return {}


def _lotes_data_json(temporada: str, lotes: list) -> str:
    """
    Serializa la lista de lotes como JSON para consumo del JS del template.
    Retorna una cadena JSON: {lote_code: {productor, variedad, ...}, ...}
    """
    data = {}
    for lote in lotes:
        info = _lote_info(temporada, lote.lote_code)
        if info:
            data[lote.lote_code] = info
    return json.dumps(data, ensure_ascii=False)


# ---------------------------------------------------------------------------
# Campos heredables desde bins hacia el lote (para reportería y enriquecimiento)
# ---------------------------------------------------------------------------
_HEREDABLE_FIELDS = (
    "codigo_productor",
    "variedad_fruta",
    "tipo_cultivo",
    "color",
    "fecha_cosecha",
    "codigo_sag_csg",
    "codigo_sag_csp",
    "codigo_sdp",
    "numero_cuartel",
    "nombre_cuartel",
    "lote_productor",
)


def _merge_bin_fields(bins: list) -> dict:
    """
    Para cada campo heredable, retorna el primer valor NO vacío de la lista de bins
    (en el orden en que vienen — estable desde la query: createdon asc en Dataverse,
    id asc en SQLite).

    Garantía: si al menos un bin del lote tiene el dato, este helper lo devuelve.
    No devuelve vacío si el dato existe en cualquier bin del lote.
    """
    merged: dict = {f: None for f in _HEREDABLE_FIELDS}
    for b in bins:
        for field in _HEREDABLE_FIELDS:
            if merged[field]:
                continue
            val = getattr(b, field, None)
            if val:
                merged[field] = val
        if all(merged[f] for f in _HEREDABLE_FIELDS):
            break
    return merged


_MULTI_VALUE_FIELDS = (
    "codigo_sag_csg",
    "codigo_sag_csp",
    "codigo_sdp",
    "nombre_cuartel",
    "lote_productor",
)
_MULTI_VALUE_FIELD_LABELS = (
    ("CSG", "codigo_sag_csg"),
    ("CSP", "codigo_sag_csp"),
    ("SDP", "codigo_sdp"),
    ("Nombre Cuartel", "nombre_cuartel"),
    ("Lote campo", "lote_productor"),
)


def _unique_lote_field_values(bins: list, fields=_MULTI_VALUE_FIELDS) -> dict:
    """Devuelve, por campo, la lista ordenada de valores únicos no vacíos de los bins del lote."""
    result = {f: [] for f in fields}
    for b in bins:
        for f in fields:
            v = getattr(b, f, None)
            if v and v not in result[f]:
                result[f].append(v)
    return result


def _lotes_json_from_records(lotes, repos=None) -> str:
    """
    Serializa LoteRecord objects como JSON para autocomplete del template.
    Usado en modo Dataverse donde los lotes ya fueron cargados via repos.
    Si se pasa repos, obtiene todos los bins, desverdizado e ingreso packing
    de cada lote en batch (≤3 queries totales, sin N+1).
    Usa _merge_bin_fields para garantizar fallback entre bins por campo.
    """
    import logging
    _log = logging.getLogger(__name__)

    # Batch pre-load: 3 queries totales independientemente del número de lotes
    todos_bins_por_lote: dict = {}
    desvs_map: dict = {}
    ips_map: dict = {}
    if repos is not None and lotes:
        lote_ids = [l.id for l in lotes if l.id]
        _log.debug("_lotes_json_from_records: batch fetch para %d lotes", len(lote_ids))
        try:
            todos_bins_por_lote = repos.bins.all_bins_by_lotes(lote_ids)
        except Exception as exc:
            _log.warning("_lotes_json_from_records: error all_bins_by_lotes: %s", exc, exc_info=True)
        try:
            desvs_map = repos.desverdizados.list_by_lotes(lote_ids)
        except Exception as exc:
            _log.warning("_lotes_json_from_records: error desverdizados.list_by_lotes: %s", exc, exc_info=True)
        try:
            ips_map = repos.ingresos_packing.list_by_lotes(lote_ids)
        except Exception as exc:
            _log.warning("_lotes_json_from_records: error ingresos_packing.list_by_lotes: %s", exc, exc_info=True)

    data = {}
    for lote in lotes:
        if lote.lote_code in data:
            continue
        via_desv = (
            lote.etapa_actual in ("Desverdizado", "Mantencion")
            if lote.etapa_actual else False
        )
        bins = todos_bins_por_lote.get(lote.id) or []
        merged = _merge_bin_fields(bins)

        # Resolver kilos mas recientes: conformacion → desverdizado salida → ingreso packing
        kilos_bruto = float(lote.kilos_bruto_conformacion) if lote.kilos_bruto_conformacion else None
        kilos_neto = float(lote.kilos_neto_conformacion) if lote.kilos_neto_conformacion else None
        desv = desvs_map.get(lote.id)
        if desv:
            if desv.kilos_bruto_salida is not None:
                kilos_bruto = float(desv.kilos_bruto_salida)
            if desv.kilos_neto_salida is not None:
                kilos_neto = float(desv.kilos_neto_salida)
            via_desv = True
        ip = ips_map.get(lote.id)
        if ip:
            if ip.kilos_bruto_ingreso_packing is not None:
                kilos_bruto = float(ip.kilos_bruto_ingreso_packing)
            if ip.kilos_neto_ingreso_packing is not None:
                kilos_neto = float(ip.kilos_neto_ingreso_packing)

        data[lote.lote_code] = {
            "lote_code": lote.lote_code,
            "estado": lote.etapa_actual or lote.estado,
            "cantidad_bins": lote.cantidad_bins,
            "kilos_bruto": kilos_bruto,
            "kilos_neto": kilos_neto,
            "via_desverdizado": via_desv,
            "requiere_desverdizado": lote.requiere_desverdizado,
            "disponibilidad_camara": lote.disponibilidad_camara_desverdizado,
            "productor": merged["codigo_productor"] or getattr(lote, "codigo_productor", "") or "",
            "variedad": merged["variedad_fruta"] or "",
            "color": merged["color"] or "",
            "fecha_cosecha": str(merged["fecha_cosecha"]) if merged["fecha_cosecha"] else "",
            "tipo_cultivo": merged["tipo_cultivo"] or "",
        }
    return json.dumps(data, ensure_ascii=False)


def _build_pallet_context_map_from_records(pallets, repos=None) -> dict:
    """
    Enriquecer pallets Dataverse con el lote asociado y su primer bin.

    La relacion pallet->lote no se expande en list_recent(), por lo que aquí
    resolvemos pallet_lote y luego reutilizamos first_bin_by_lotes() para traer
    el contexto heredado visible en UI.
    """
    data = {
        p.pallet_code: {
            "pallet_code": p.pallet_code,
            "lote_code": "",
            "tipo_caja": p.tipo_caja or "",
            "peso_total": float(p.peso_total_kg) if p.peso_total_kg else None,
            "productor": "",
            "codigo_sag_csg": "",
            "codigo_sag_csp": "",
            "tipo_cultivo": "",
            "variedad": "",
            "color": "",
            "fecha_cosecha": "",
        }
        for p in pallets
    }
    if repos is None or not pallets:
        return data

    ordered_codes: list[str] = []
    deduped_pallets = []
    for pallet in pallets:
        if pallet.pallet_code in data and pallet.pallet_code in ordered_codes:
            continue
        ordered_codes.append(pallet.pallet_code)
        deduped_pallets.append(pallet)

    data = {p.pallet_code: data[p.pallet_code] for p in deduped_pallets}
    pallets = deduped_pallets

    pallet_to_lote: dict = {}
    lote_ids: list = []
    for pallet in pallets:
        try:
            pl = repos.pallet_lotes.find_by_pallet(pallet.id)
        except Exception:
            pl = None
        if pl and pl.lote_id:
            pallet_to_lote[pallet.id] = pl.lote_id
            if pl.lote_id not in lote_ids:
                lote_ids.append(pl.lote_id)

    lotes_by_id = {}
    for lote_id in lote_ids:
        try:
            lote = repos.lotes.find_by_id(lote_id)
        except Exception:
            lote = None
        if lote:
            lotes_by_id[lote_id] = lote

    try:
        todos_bins_por_lote = repos.bins.all_bins_by_lotes(lote_ids) if lote_ids else {}
    except Exception:
        todos_bins_por_lote = {}

    ingreso_packing_by_lote = {}
    for lote_id in lote_ids:
        try:
            ingreso = repos.ingresos_packing.find_by_lote(lote_id)
            if ingreso:
                ingreso_packing_by_lote[lote_id] = ingreso
        except Exception:
            pass

    for pallet in pallets:
        lote_id = pallet_to_lote.get(pallet.id)
        lote = lotes_by_id.get(lote_id)
        bins = todos_bins_por_lote.get(lote_id) or [] if lote_id else []
        merged = _merge_bin_fields(bins)
        peso_total = data[pallet.pallet_code]["peso_total"]
        ingreso = ingreso_packing_by_lote.get(lote_id)
        if ingreso and ingreso.kilos_neto_ingreso_packing:
            peso_total = float(ingreso.kilos_neto_ingreso_packing)
        data[pallet.pallet_code].update({
            "lote_code": lote.lote_code if lote else "",
            "peso_total": peso_total,
            "productor": merged["codigo_productor"] or (getattr(lote, "codigo_productor", "") if lote else "") or "",
            "codigo_sag_csg": merged["codigo_sag_csg"] or "",
            "codigo_sag_csp": merged["codigo_sag_csp"] or "",
            "tipo_cultivo": merged["tipo_cultivo"] or "",
            "variedad": merged["variedad_fruta"] or "",
            "color": merged["color"] or "",
            "fecha_cosecha": str(merged["fecha_cosecha"]) if merged["fecha_cosecha"] else "",
        })
    return data


def _pallets_json_from_records(pallets, repos=None) -> str:
    """Serializa PalletRecord objects como JSON para autocomplete del template."""
    return json.dumps(_build_pallet_context_map_from_records(pallets, repos), ensure_ascii=False)


def _campos_base_lote(lote: Lote) -> dict:
    """
    Deriva los campos base del lote (productor, tipo_cultivo, variedad, color,
    fecha_cosecha) desde el primer bin asociado. No depende de sesion.
    """
    bin_lotes = list(lote.bin_lotes.select_related("bin").order_by("created_at")[:1])
    primer_bin = bin_lotes[0].bin if bin_lotes else None
    return {
        "productor":     primer_bin.codigo_productor if primer_bin else "",
        "tipo_cultivo":  primer_bin.tipo_cultivo     if primer_bin else "",
        "variedad":      primer_bin.variedad_fruta   if primer_bin else "",
        "color":         primer_bin.color            if primer_bin else "",
        "fecha_cosecha": str(primer_bin.fecha_cosecha) if primer_bin and primer_bin.fecha_cosecha else "",
    }


def _lotes_para_paletizar(temporada: str) -> list:
    """
    Retorna lotes con ingreso a packing que aun no han sido vinculados a un pallet.
    Estos son los lotes disponibles para ser convertidos en pallet.
    """
    from django.conf import settings
    if getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip() == "dataverse":
        try:
            from infrastructure.repository_factory import get_repositories
            repos = get_repositories()
            lotes = repos.lotes.list_recent(limit=200)
            return [l for l in lotes if l.is_active and l.etapa_actual in ("Packing / Proceso", "Ingreso Packing")]
        except Exception:
            return []
    from operaciones.models import IngresoAPacking, PalletLote
    paletizados_ids = set(PalletLote.objects.values_list("lote_id", flat=True))
    return list(
        Lote.objects
        .filter(temporada=temporada, is_active=True, ingreso_packing__isnull=False)
        .exclude(id__in=paletizados_ids)
        .order_by("-created_at")[:50]
    )


def _pallets_pendientes_calidad(temporada: str) -> list:
    """Pallets activos que aun no tienen registro de CalidadPallet."""
    from django.conf import settings
    if getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip() == "dataverse":
        try:
            from infrastructure.repository_factory import get_repositories
            # En Dataverse retornamos los pallets recientes sin filtrar por calidad;
            # la logica de negocio en el use case previene registros duplicados.
            return get_repositories().pallets.list_recent(limit=30)
        except Exception:
            return []
    from operaciones.models import CalidadPallet
    pallets_con_calidad = set(CalidadPallet.objects.values_list("pallet_id", flat=True))
    return list(
        Pallet.objects
        .filter(temporada=temporada, is_active=True)
        .exclude(id__in=pallets_con_calidad)
        .order_by("-created_at")[:30]
    )


def _pallets_pendientes_camara_frio(temporada: str) -> list:
    """Pallets activos que aun no tienen registro de CamaraFrio."""
    from django.conf import settings
    if getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip() == "dataverse":
        try:
            from infrastructure.repository_factory import get_repositories
            return get_repositories().pallets.list_recent(limit=30)
        except Exception:
            return []
    from operaciones.models import CamaraFrio
    pallets_con_camara = set(CamaraFrio.objects.values_list("pallet_id", flat=True))
    return list(
        Pallet.objects
        .filter(temporada=temporada, is_active=True)
        .exclude(id__in=pallets_con_camara)
        .order_by("-created_at")[:30]
    )


def _pallet_info(temporada: str, pallet_code: str) -> dict:
    """Datos base de un pallet para autocompletar formularios."""
    from django.conf import settings
    if getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip() == "dataverse":
        try:
            from infrastructure.repository_factory import get_repositories
            repos = get_repositories()
            pallet = repos.pallets.find_by_code(temporada, pallet_code)
            if not pallet:
                return {}
            return _build_pallet_context_map_from_records([pallet], repos).get(pallet_code, {})
        except Exception:
            return {}
    try:
        pallet = Pallet.objects.get(temporada=temporada, pallet_code=pallet_code)
        lote_code = ""
        campos = {}
        peso_total = None
        pl = pallet.pallet_lotes.select_related("lote").first()
        if pl:
            lote_code = pl.lote.lote_code
            campos = _campos_base_lote(pl.lote)
            ingreso = pl.lote.ingreso_packing
            if ingreso and ingreso.kilos_neto_ingreso_packing:
                peso_total = float(ingreso.kilos_neto_ingreso_packing)
        if peso_total is None:
            peso_total = float(pallet.peso_total_kg) if pallet.peso_total_kg else None
        return {
            "pallet_code": pallet.pallet_code,
            "lote_code":   lote_code,
            "tipo_caja":   pallet.tipo_caja,
            "peso_total":  peso_total,
            **campos,
        }
    except Pallet.DoesNotExist:
        return {}


def _pallets_data_json(temporada: str, pallets: list) -> str:
    data = {}
    for p in pallets:
        info = _pallet_info(temporada, p.pallet_code)
        if info:
            data[p.pallet_code] = info
    return json.dumps(data, ensure_ascii=False)


class IngresoPackingView(LoginRequiredMixin, RolRequiredMixin, TemplateView):
    roles_requeridos = ["Ingreso Packing"]
    template_name = "operaciones/ingreso_packing.html"
    login_url = reverse_lazy("usuarios:login")

    # ------------------------------------------------------------------
    # Context helper
    # ------------------------------------------------------------------

    def _build_context_base(self, request):
        temporada = (
            request.session.get("temporada_activa")
            or str(datetime.date.today().year)
        )
        lotes = _lotes_pendientes_ingreso_packing(temporada)

        from django.conf import settings
        if getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip() == "dataverse":
            from infrastructure.repository_factory import get_repositories
            lotes_data_json = _lotes_json_from_records(lotes, get_repositories())
        else:
            lotes_data_json = _lotes_data_json(temporada, lotes)

        lote_activo_code = request.session.get("lote_activo_ingreso_packing")
        pesajes = request.session.get("pesajes_ingreso_packing", [])

        lote_info = {}
        bins_pesados = 0
        bins_pendientes = 0
        from decimal import Decimal as D
        bruto_acumulado = D("0")
        neto_acumulado = D("0")

        if lote_activo_code:
            lote_info = _lote_info(temporada, lote_activo_code) or {}
            bins_pesados = sum(p["cantidad_bins"] for p in pesajes)
            cantidad_bins_lote = lote_info.get("cantidad_bins", 0) or 0
            bins_pendientes = cantidad_bins_lote - bins_pesados
            bruto_acumulado = sum(D(str(p["kilos_brutos"])) for p in pesajes)
            neto_acumulado = sum(D(str(p["kilos_netos"])) for p in pesajes)

        return {
            "page_title": "Ingreso a Packing",
            "lotes_pendientes": lotes,
            "lotes_data_json": lotes_data_json,
            "lote_activo_code": lote_activo_code,
            "lote_info": lote_info,
            "pesajes_ingreso_packing": pesajes,
            "bins_pesados": bins_pesados,
            "bins_pendientes": bins_pendientes,
            "bruto_acumulado": bruto_acumulado,
            "neto_acumulado": neto_acumulado,
            "form_pesaje_parcial": PesajeParcialIngresoPackingForm(),
            "form_datos": IngresoPackingForm(),
            "codigo_operador": request.session.get("crf21_codigooperador", ""),
        }

    # ------------------------------------------------------------------
    # GET
    # ------------------------------------------------------------------

    def get(self, request, *args, **kwargs):
        ctx = self._build_context_base(request)
        return render(request, self.template_name, ctx)

    # ------------------------------------------------------------------
    # POST dispatcher
    # ------------------------------------------------------------------

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action", "")
        dispatch = {
            "seleccionar_lote":       self._handle_seleccionar_lote,
            "agregar_pesaje_packing": self._handle_agregar_pesaje_packing,
            "eliminar_pesaje_packing": self._handle_eliminar_pesaje_packing,
            "registrar_ingreso":      self._handle_registrar_ingreso,
        }
        handler = dispatch.get(action)
        if not handler:
            messages.error(request, "Accion desconocida.")
            return redirect("operaciones:ingreso_packing")
        return handler(request)

    # ------------------------------------------------------------------
    # Action handlers
    # ------------------------------------------------------------------

    def _handle_seleccionar_lote(self, request):
        lote_code = request.POST.get("lote_code", "").strip()
        if not lote_code:
            request.session.pop("lote_activo_ingreso_packing", None)
            request.session.pop("pesajes_ingreso_packing", None)
            return redirect("operaciones:ingreso_packing")
        temporada = _temporada(request)
        info = _lote_info(temporada, lote_code)
        if not info:
            messages.error(request, f"Lote {lote_code} no encontrado.")
            return redirect("operaciones:ingreso_packing")
        request.session["lote_activo_ingreso_packing"] = lote_code
        request.session["pesajes_ingreso_packing"] = []
        messages.success(request, f"Lote {lote_code} seleccionado. Registre los pesajes parciales.")
        return redirect("operaciones:ingreso_packing")

    def _handle_agregar_pesaje_packing(self, request):
        lote_code = request.session.get("lote_activo_ingreso_packing")
        if not lote_code:
            messages.error(request, "No hay lote activo. Seleccione un lote primero.")
            return redirect("operaciones:ingreso_packing")

        form = PesajeParcialIngresoPackingForm(request.POST)
        if not form.is_valid():
            for errs in form.errors.values():
                for e in errs:
                    messages.error(request, e)
            return redirect("operaciones:ingreso_packing")

        cd = form.cleaned_data
        cantidad = cd["cantidad_bins"]
        kilos_brutos = float(cd["kilos_brutos_grupo"])
        tara = float(cd["tara"])
        kilos_netos = float(cd["kilos_netos_grupo"])

        temporada = _temporada(request)
        lote_info = _lote_info(temporada, lote_code) or {}
        cantidad_bins_lote = lote_info.get("cantidad_bins", 0) or 0

        if cantidad_bins_lote == 0:
            messages.error(request, "No se pudo determinar la cantidad de bins del lote. Intente nuevamente.")
            return redirect("operaciones:ingreso_packing")

        pesajes = list(request.session.get("pesajes_ingreso_packing", []))
        bins_pesados = sum(p["cantidad_bins"] for p in pesajes)
        bins_pendientes = cantidad_bins_lote - bins_pesados

        if cantidad > bins_pendientes:
            messages.error(
                request,
                f"No puede registrar {cantidad} bin(s): solo quedan {bins_pendientes} pendiente(s) de pesar.",
            )
            return redirect("operaciones:ingreso_packing")

        pesajes.append({
            "cantidad_bins": cantidad,
            "kilos_brutos": kilos_brutos,
            "tara": tara,
            "kilos_netos": round(kilos_netos, 2),
        })
        request.session["pesajes_ingreso_packing"] = pesajes
        messages.success(
            request,
            f"Pesaje registrado: {cantidad} bin(s) — {kilos_netos:.2f} kg netos.",
        )
        return redirect("operaciones:ingreso_packing")

    def _handle_eliminar_pesaje_packing(self, request):
        try:
            idx = int(request.POST.get("pesaje_idx", -1))
        except (ValueError, TypeError):
            idx = -1
        pesajes = list(request.session.get("pesajes_ingreso_packing", []))
        if 0 <= idx < len(pesajes):
            pesajes.pop(idx)
            request.session["pesajes_ingreso_packing"] = pesajes
            messages.success(request, "Pesaje eliminado.")
        else:
            messages.error(request, "Pesaje no encontrado.")
        return redirect("operaciones:ingreso_packing")

    def _handle_registrar_ingreso(self, request):
        lote_code = request.session.get("lote_activo_ingreso_packing")
        if not lote_code:
            messages.error(request, "No hay lote activo.")
            return redirect("operaciones:ingreso_packing")

        temporada = _temporada(request)
        lote_info = _lote_info(temporada, lote_code) or {}
        cantidad_bins_lote = lote_info.get("cantidad_bins", 0) or 0

        pesajes = request.session.get("pesajes_ingreso_packing", [])
        bins_pesados = sum(p["cantidad_bins"] for p in pesajes)

        if bins_pesados != cantidad_bins_lote:
            pendientes = cantidad_bins_lote - bins_pesados
            messages.error(
                request,
                f"Quedan {pendientes} bin(s) pendientes de pesar. "
                "Complete el pesaje antes de registrar el ingreso.",
            )
            return redirect("operaciones:ingreso_packing")

        form = IngresoPackingForm(request.POST)
        if not form.is_valid():
            messages.error(request, "Formulario de datos invalido.")
            return redirect("operaciones:ingreso_packing")

        cd = form.cleaned_data
        from decimal import Decimal as D
        kilos_bruto_total = float(sum(D(str(p["kilos_brutos"])) for p in pesajes))
        kilos_neto_total = float(sum(D(str(p["kilos_netos"])) for p in pesajes))

        payload = {
            "temporada": temporada,
            "lote_code": lote_code,
            "operator_code": request.session.get("crf21_codigooperador", ""),
            "source_system": "web",
            "via_desverdizado": cd.get("via_desverdizado") or lote_info.get("via_desverdizado", False),
            "extra": {
                "fecha_ingreso": _serialize_form_value(cd["fecha_ingreso"]) if cd.get("fecha_ingreso") else None,
                "hora_ingreso": _serialize_form_value(cd["hora_ingreso"]) if cd.get("hora_ingreso") else None,
                "kilos_bruto_ingreso_packing": kilos_bruto_total,
                "kilos_neto_ingreso_packing": kilos_neto_total,
                "observaciones": cd.get("observaciones"),
            },
        }
        result = registrar_ingreso_packing(payload)
        if result.ok:
            request.session.pop("lote_activo_ingreso_packing", None)
            request.session.pop("pesajes_ingreso_packing", None)
            _consulta_cache_invalidate()
            messages.success(request, result.message)
            return redirect("operaciones:ingreso_packing")
        for err in result.errors:
            messages.error(request, err)
        return redirect("operaciones:ingreso_packing")


# ---------------------------------------------------------------------------
# Proceso packing (registro de produccion)
# ---------------------------------------------------------------------------

class ProcesoView(LoginRequiredMixin, RolRequiredMixin, TemplateView):
    roles_requeridos = ["Proceso"]
    template_name = "operaciones/proceso.html"
    login_url = reverse_lazy("usuarios:login")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["page_title"] = "Proceso Packing"
        ctx["form"] = RegistroPackingForm()
        temporada = (
            self.request.session.get("temporada_activa")
            or str(datetime.date.today().year)
        )
        from django.conf import settings
        backend = getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip()
        if backend == "dataverse":
            try:
                from infrastructure.repository_factory import get_repositories
                repos = get_repositories()
                todos = repos.lotes.list_recent(limit=200)
                _CON_INGRESO = ("Ingreso Packing", "Packing / Proceso", "Paletizado",
                                "Calidad Pallet", "Camara Frio", "Temperatura Salida")
                lotes = [l for l in todos if l.is_active and l.etapa_actual in _CON_INGRESO][:30]
            except Exception:
                lotes = []
            ctx["lotes_pendientes"] = lotes
            ctx["lotes_data_json"] = _lotes_json_from_records(lotes, repos)
        else:
            # Lotes que ya tienen ingreso packing (aptos para proceso)
            from operaciones.models import IngresoAPacking
            lotes_con_ingreso = set(
                IngresoAPacking.objects.values_list("lote_id", flat=True)
            )
            lotes = list(
                Lote.objects
                .filter(temporada=temporada, is_active=True, id__in=lotes_con_ingreso)
                .order_by("-created_at")[:30]
            )
            ctx["lotes_pendientes"] = lotes
            ctx["lotes_data_json"] = _lotes_data_json(temporada, lotes)
        return ctx

    def post(self, request, *args, **kwargs):
        form = RegistroPackingForm(request.POST)
        if not form.is_valid():
            messages.error(request, "Formulario invalido.")
            ctx = self.get_context_data()
            ctx["form"] = form
            return render(request, self.template_name, ctx)

        cd = form.cleaned_data
        lote_code = request.POST.get("lote_code", "").strip()
        payload = {
            "temporada": _temporada(request),
            "lote_code": lote_code,
            "operator_code": request.session.get("crf21_codigooperador", ""),
            "source_system": "web",
            "extra": {
                "fecha": _serialize_form_value(cd["fecha"]) if cd.get("fecha") else None,
                "hora_inicio": _serialize_form_value(cd["hora_inicio"]) if cd.get("hora_inicio") else None,
                "linea_proceso": cd.get("linea_proceso"),
                "categoria_calidad": cd.get("categoria_calidad"),
                "calibre": cd.get("calibre"),
                "tipo_envase": cd.get("tipo_envase"),
                "cantidad_cajas_producidas": cd.get("cantidad_cajas_producidas"),
                "merma_seleccion_pct": float(cd["merma_seleccion_pct"]) if cd.get("merma_seleccion_pct") else None,
            },
        }
        result = registrar_registro_packing(payload)
        _handle_result(request, result)
        return redirect("operaciones:proceso")


# ---------------------------------------------------------------------------
# Control proceso packing (parámetros máquina volcado/tina)
# ---------------------------------------------------------------------------

class ControlProcesoView(LoginRequiredMixin, RolRequiredMixin, TemplateView):
    roles_requeridos = ["Control"]
    template_name = "operaciones/control_proceso.html"
    login_url = reverse_lazy("usuarios:login")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["page_title"] = "Control Proceso Packing"
        form = ControlProcesoPackingForm()
        form.fields['rol'].initial = self.request.session.get("crf21_codigooperador", "")
        ctx["form"] = form
        temporada = (
            self.request.session.get("temporada_activa")
            or str(datetime.date.today().year)
        )
        from django.conf import settings
        backend = getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip()
        if backend == "dataverse":
            try:
                from infrastructure.repository_factory import get_repositories
                repos = get_repositories()
                todos = repos.lotes.list_recent(limit=200)
                _CON_INGRESO = ("Ingreso Packing", "Packing / Proceso", "Paletizado",
                                "Calidad Pallet", "Camara Frio", "Temperatura Salida")
                lotes = [l for l in todos if l.is_active and l.etapa_actual in _CON_INGRESO][:30]
            except Exception:
                lotes = []
            ctx["lotes_pendientes"] = lotes
            ctx["lotes_data_json"] = _lotes_json_from_records(lotes, repos)
        else:
            from operaciones.models import IngresoAPacking
            lotes_con_ingreso = set(IngresoAPacking.objects.values_list("lote_id", flat=True))
            lotes = list(
                Lote.objects
                .filter(temporada=temporada, is_active=True, id__in=lotes_con_ingreso)
                .order_by("-created_at")[:30]
            )
            ctx["lotes_pendientes"] = lotes
            ctx["lotes_data_json"] = _lotes_data_json(temporada, lotes)
        return ctx

    def post(self, request, *args, **kwargs):
        form = ControlProcesoPackingForm(request.POST)
        if not form.is_valid():
            messages.error(request, "Formulario invalido.")
            ctx = self.get_context_data()
            ctx["form"] = form
            return render(request, self.template_name, ctx)

        cd = form.cleaned_data
        lote_code = request.POST.get("lote_code", "").strip()

        def _dec(key):
            v = cd.get(key)
            return float(v) if v is not None else None

        payload = {
            "temporada": _temporada(request),
            "lote_code": lote_code,
            "operator_code": request.session.get("crf21_codigooperador", ""),
            "source_system": "web",
            "extra": {
                "fecha": _serialize_form_value(cd["fecha"]) if cd.get("fecha") else None,
                "hora": _serialize_form_value(cd["hora"]) if cd.get("hora") else None,
                "n_bins_procesados": cd.get("n_bins_procesados"),
                "velocidad_volcador": _dec("velocidad_volcador"),
                "obs_volcador": cd.get("obs_volcador"),
                "temp_agua_tina": _dec("temp_agua_tina"),
                "cloro_libre_ppm": _dec("cloro_libre_ppm"),
                "ph_agua": _dec("ph_agua"),
                "tiempo_inmersion_seg": cd.get("tiempo_inmersion_seg"),
                "recambio_agua": cd.get("recambio_agua"),
                "temp_aire_secado": _dec("temp_aire_secado"),
                "velocidad_ventiladores": _dec("velocidad_ventiladores"),
                "fruta_sale_seca": cd.get("fruta_sale_seca"),
                "tipo_cera": cd.get("tipo_cera"),
                "dosis_cera_ml_min": _dec("dosis_cera_ml_min"),
                "temp_cera": _dec("temp_cera"),
                "cobertura_uniforme": cd.get("cobertura_uniforme"),
                "n_operarios_seleccion": cd.get("n_operarios_seleccion"),
                "fruta_dano_condicion_kg": _dec("fruta_dano_condicion_kg"),
                "fruta_dano_calidad_kg": _dec("fruta_dano_calidad_kg"),
                "fruta_pudricion_kg": _dec("fruta_pudricion_kg"),
                "merma_total_seleccion_kg": _dec("merma_total_seleccion_kg"),
                "equipo_calibrador": cd.get("equipo_calibrador"),
                "calibre_predominante": cd.get("calibre_predominante"),
                "pct_calibre_export": _dec("pct_calibre_export"),
                "pct_calibres_menores": _dec("pct_calibres_menores"),
                "tipo_caja": cd.get("tipo_caja"),
                "peso_promedio_caja_kg": _dec("peso_promedio_caja_kg"),
                "n_cajas_producidas": cd.get("n_cajas_producidas"),
                "rendimiento_lote_pct": _dec("rendimiento_lote_pct"),
                "observaciones_generales": cd.get("observaciones_generales"),
                "rol": request.session.get("crf21_codigooperador", ""),
            },
        }
        result = registrar_control_proceso_packing(payload)
        _handle_result(request, result)
        return redirect("operaciones:control_proceso")


# ---------------------------------------------------------------------------
# Control de Calidad — Índice
# ---------------------------------------------------------------------------

class ControlCalidadIndexView(LoginRequiredMixin, RolRequiredMixin, TemplateView):
    roles_requeridos = ["Control"]
    template_name = "operaciones/control_calidad_index.html"
    login_url = reverse_lazy("usuarios:login")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["page_title"] = "Control de Calidad"
        return ctx


# ---------------------------------------------------------------------------
# Control de Calidad — Desverdizado (Planilla 1: Calibres, Planilla 2: Semillas)
# ---------------------------------------------------------------------------

class ControlCalidadDesverdizadoView(LoginRequiredMixin, RolRequiredMixin, TemplateView):
    roles_requeridos = ["Control"]
    template_name = "operaciones/control_calidad_desverdizado.html"
    login_url = reverse_lazy("usuarios:login")

    def _get_lotes(self, request):
        temporada = (
            request.session.get("temporada_activa")
            or str(datetime.date.today().year)
        )
        from django.conf import settings
        backend = getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip()
        if backend == "dataverse":
            try:
                from infrastructure.repository_factory import get_repositories
                repos = get_repositories()
                todos = repos.lotes.list_recent(limit=200)
                lotes = [l for l in todos if l.is_active][:50]
            except Exception:
                lotes = []
            return lotes, backend, temporada
        else:
            lotes = list(
                Lote.objects
                .filter(temporada=temporada, is_active=True)
                .order_by("-created_at")[:50]
            )
            return lotes, backend, temporada

    def _build_lotes_data_json(self, lotes, backend, temporada) -> str:
        """Construye JSON con datos del lote+bin para autocompletar los formularios."""
        data = {}
        if backend == "dataverse":
            try:
                from infrastructure.repository_factory import get_repositories
                repos = get_repositories()
                lote_ids = [l.id for l in lotes if l.id]
                todos_bins_por_lote = repos.bins.all_bins_by_lotes(lote_ids) if lote_ids else {}
            except Exception:
                todos_bins_por_lote = {}
            for lote in lotes:
                bins = todos_bins_por_lote.get(lote.id) or []
                m = _merge_bin_fields(bins)
                data[lote.lote_code] = {
                    "productor": m["codigo_productor"] or getattr(lote, "codigo_productor", "") or "",
                    "variedad":  m["variedad_fruta"] or "",
                    "color":     m["color"] or "",
                    "fecha_cosecha": str(m["fecha_cosecha"]) if m["fecha_cosecha"] else "",
                    "trazabilidad": lote.lote_code,
                    "cuartel": m["nombre_cuartel"] or m["numero_cuartel"] or "",
                    "sector":  "",
                }
        else:
            from operaciones.models import BinLote
            # Batch: prefetch bins para todos los lotes de una vez
            lote_ids = [l.pk for l in lotes]
            bins_qs = (
                BinLote.objects
                .filter(lote_id__in=lote_ids)
                .select_related("bin", "lote")
                .order_by("lote_id", "created_at")
            )
            primer_bin_map = {}
            for bl in bins_qs:
                if bl.lote_id not in primer_bin_map:
                    primer_bin_map[bl.lote_id] = bl.bin
            for lote in lotes:
                b = primer_bin_map.get(lote.pk)
                data[lote.lote_code] = {
                    "productor": b.productor if b else "",
                    "variedad":  b.variedad_fruta if b else "",
                    "color":     b.color if b else "",
                    "fecha_cosecha": str(b.fecha_cosecha) if b and b.fecha_cosecha else "",
                    "trazabilidad": lote.lote_code,
                    "cuartel": (b.nombre_cuartel or b.numero_cuartel) if b else "",
                    "sector":  b.sector if b else "",
                }
        return json.dumps(data, ensure_ascii=False)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["page_title"] = "Control Calidad Desverdizado"
        lotes, backend, temporada = self._get_lotes(self.request)
        ctx["lotes_pendientes"] = lotes
        ctx["lotes_data_json"] = self._build_lotes_data_json(lotes, backend, temporada)
        ctx["supervisor_default"] = self.request.session.get("crf21_codigooperador", "")
        ctx["form_calibres"] = PlanillaDesverdizadoCalibreForm()
        ctx["form_semillas"] = PlanillaDesverdizadoSemillasForm()
        ctx["calibres_nombres"] = ["1xx", "1x", "1", "2", "3", "4", "5", "Precalibre"]
        ctx["grupos_calibre"] = [1, 2, 3]
        ctx["grupos_semillas"] = [1, 2, 3, 4, 5]
        ctx["frutas_range"] = list(range(1, 11))
        ctx["defectos_nombres"] = [
            ("oleocelosis", "Oleocelosis"),
            ("heridas_abiertas", "Heridas abiertas"),
            ("rugoso", "Rugoso"),
            ("deforme", "Deforme"),
            ("golpe_sol", "Golpe sol"),
            ("verdes", "Verdes"),
            ("pre_calibre", "Pre-calibre"),
            ("palo_largo", "Palo largo"),
        ]
        return ctx

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action", "calibres")
        lote_code = request.POST.get("lote_code", "").strip()

        if action == "calibres":
            form = PlanillaDesverdizadoCalibreForm(request.POST)
            if not form.is_valid():
                messages.error(request, "Formulario de calibres inválido.")
                ctx = self.get_context_data()
                ctx["form_calibres"] = form
                ctx["tab_activo"] = "calibres"
                return render(request, self.template_name, ctx)
            cd = form.cleaned_data
            # Parse calibres grupos from dynamic POST fields (3 groups)
            calibres_grupos = []
            for g in range(1, 4):
                calibres = {}
                for cal in ["1xx", "1x", "1", "2", "3", "4", "5", "Precalibre"]:
                    key = f"grupo_{g}_cal_{cal}"
                    val = request.POST.get(key, "")
                    try:
                        calibres[cal] = int(val) if val else None
                    except ValueError:
                        calibres[cal] = None
                calibres_grupos.append({
                    "color": request.POST.get(f"grupo_{g}_color", ""),
                    "calibres": calibres,
                    "observacion": request.POST.get(f"grupo_{g}_obs", ""),
                })
            payload = {
                "lote_code": lote_code,
                "operator_code": request.session.get("crf21_codigooperador", ""),
                "source_system": "web",
                "extra": {
                    **{k: _serialize_form_value(v)
                       for k, v in cd.items()},
                    "calibres_grupos": calibres_grupos,
                },
            }
            result = registrar_planilla_desv_calibre(payload)
            _handle_result(request, result)
            return redirect("operaciones:control_desverdizado")

        elif action == "semillas":
            form = PlanillaDesverdizadoSemillasForm(request.POST)
            if not form.is_valid():
                messages.error(request, "Formulario de semillas inválido.")
                ctx = self.get_context_data()
                ctx["form_semillas"] = form
                ctx["tab_activo"] = "semillas"
                return render(request, self.template_name, ctx)
            cd = form.cleaned_data
            # Parse 50 frutas: g{G}_f{N}_semillas (G=1-5, N=1-10)
            frutas_data = []
            n = 1
            for g in range(1, 6):
                for f in range(1, 11):
                    val = request.POST.get(f"g{g}_f{f}_semillas", "")
                    try:
                        semillas = int(val) if val else 0
                    except ValueError:
                        semillas = 0
                    frutas_data.append({"n_fruto": n, "n_semillas": semillas})
                    n += 1
            payload = {
                "lote_code": lote_code,
                "operator_code": request.session.get("crf21_codigooperador", ""),
                "source_system": "web",
                "extra": {
                    **{k: _serialize_form_value(v)
                       for k, v in cd.items()},
                    "frutas_data": frutas_data,
                },
            }
            result = registrar_planilla_desv_semillas(payload)
            _handle_result(request, result)
            return redirect("operaciones:control_desverdizado")

        messages.error(request, "Acción desconocida.")
        return redirect("operaciones:control_desverdizado")


# ---------------------------------------------------------------------------
# Control de Calidad — Packing Cítricos
# ---------------------------------------------------------------------------

class ControlCalidadPackingView(LoginRequiredMixin, RolRequiredMixin, TemplateView):
    roles_requeridos = ["Control"]
    template_name = "operaciones/control_calidad_packing.html"
    login_url = reverse_lazy("usuarios:login")

    def _get_pallets(self, request):
        from django.conf import settings
        backend = getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip()
        if backend == "dataverse":
            try:
                from infrastructure.repository_factory import get_repositories
                repos = get_repositories()
                pallets = repos.pallets.list_recent(limit=50)
            except Exception:
                pallets = []
            return pallets
        else:
            return list(Pallet.objects.order_by("-created_at")[:50])

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["page_title"] = "Control Calidad Packing Cítricos"
        ctx["pallets"] = self._get_pallets(self.request)
        ctx["form"] = PlanillaCalidadPackingForm()
        return ctx

    def post(self, request, *args, **kwargs):
        form = PlanillaCalidadPackingForm(request.POST)
        if not form.is_valid():
            messages.error(request, "Formulario inválido.")
            ctx = self.get_context_data()
            ctx["form"] = form
            return render(request, self.template_name, ctx)
        cd = form.cleaned_data
        pallet_id = request.POST.get("pallet_id", "").strip() or None
        payload = {
            "pallet_id": pallet_id,
            "operator_code": request.session.get("crf21_codigooperador", ""),
            "source_system": "web",
            "extra": {
                k: _serialize_form_value(v)
                for k, v in cd.items()
            },
        }
        result = registrar_planilla_calidad_packing(payload)
        _handle_result(request, result)
        return redirect("operaciones:control_packing")


# ---------------------------------------------------------------------------
# Control de Calidad — Cámaras
# ---------------------------------------------------------------------------

class ControlCalidadCamarasView(LoginRequiredMixin, RolRequiredMixin, TemplateView):
    roles_requeridos = ["Control"]
    template_name = "operaciones/control_calidad_camaras.html"
    login_url = reverse_lazy("usuarios:login")

    def _get_pallets(self, request):
        from django.conf import settings
        backend = getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip()
        if backend == "dataverse":
            try:
                from infrastructure.repository_factory import get_repositories
                repos = get_repositories()
                pallets = repos.pallets.list_recent(limit=50)
            except Exception:
                pallets = []
            return pallets
        else:
            return list(Pallet.objects.order_by("-created_at")[:50])

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["page_title"] = "Control Calidad Cámaras"
        ctx["pallets"] = self._get_pallets(self.request)
        ctx["form"] = PlanillaCalidadCamaraForm()
        return ctx

    def post(self, request, *args, **kwargs):
        form = PlanillaCalidadCamaraForm(request.POST)
        if not form.is_valid():
            messages.error(request, "Formulario inválido.")
            ctx = self.get_context_data()
            ctx["form"] = form
            return render(request, self.template_name, ctx)
        cd = form.cleaned_data
        pallet_id = request.POST.get("pallet_id", "").strip() or None
        # Parse dynamic hourly measurement rows: med_{i}_{campo}
        mediciones = []
        i = 0
        while True:
            hora = request.POST.get(f"med_{i}_hora", "")
            if not hora:
                break
            fila = {"hora": hora}
            for campo in ["ambiente", "pulpa_ext_entrada", "pulpa_ext_medio",
                          "pulpa_ext_salida", "pulpa_int_entrada", "pulpa_int_media",
                          "pulpa_int_salida"]:
                fila[campo] = request.POST.get(f"med_{i}_{campo}", "") or None
            mediciones.append(fila)
            i += 1
        payload = {
            "pallet_id": pallet_id,
            "operator_code": request.session.get("crf21_codigooperador", ""),
            "source_system": "web",
            "extra": {
                **{k: _serialize_form_value(v)
                   for k, v in cd.items()},
                "mediciones": mediciones,
            },
        }
        result = registrar_planilla_calidad_camara(payload)
        _handle_result(request, result)
        return redirect("operaciones:control_camaras")


# ---------------------------------------------------------------------------
# Paletizado (calidad + cerrar pallet)
# ---------------------------------------------------------------------------

class PaletizadoView(LoginRequiredMixin, RolRequiredMixin, TemplateView):
    roles_requeridos = ["Paletizado"]
    template_name = "operaciones/paletizado.html"
    login_url = reverse_lazy("usuarios:login")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["page_title"] = "Paletizado"
        temporada = (
            self.request.session.get("temporada_activa")
            or str(datetime.date.today().year)
        )
        # Lotes listos para convertirse en pallet (accion primaria)
        lotes = _lotes_para_paletizar(temporada)
        ctx["lotes_para_paletizar"] = lotes
        from django.conf import settings
        backend = getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip()
        if backend == "dataverse":
            from infrastructure.repository_factory import get_repositories
            repos = get_repositories()
            ctx["lotes_data_json"] = _lotes_json_from_records(lotes, repos)
        else:
            ctx["lotes_data_json"] = _lotes_data_json(temporada, lotes)
        # Pallets activos para control de calidad y cierre
        pallets = _pallets_pendientes_calidad(temporada)
        ctx["pallets_pendientes"] = pallets
        if backend == "dataverse":
            ctx["pallets_data_json"] = _pallets_json_from_records(pallets)
        else:
            ctx["pallets_data_json"] = _pallets_data_json(temporada, pallets)
        ctx["form_muestra"] = CalidadPalletMuestraForm()
        return ctx

    def post(self, request, *args, **kwargs):
        temporada = _temporada(request)
        lote_code = request.POST.get("lote_code", "").strip()
        pallet_code = request.POST.get("pallet_code", "").strip()
        payload = {
            "temporada": temporada,
            "lote_codes": [lote_code] if lote_code else [],
            "pallet_code": pallet_code,
            "operator_code": request.session.get("crf21_codigooperador", ""),
            "source_system": "web",
        }
        result = cerrar_pallet(payload)
        _handle_result(request, result)
        return redirect("operaciones:paletizado")


# ---------------------------------------------------------------------------
# Camaras (frio + medicion temperatura)
# ---------------------------------------------------------------------------

class CamarasView(LoginRequiredMixin, RolRequiredMixin, TemplateView):
    roles_requeridos = ["Camaras"]
    template_name = "operaciones/camaras.html"
    login_url = reverse_lazy("usuarios:login")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["page_title"] = "Camaras de Frio"
        ctx["form_camara"] = CamaraFrioForm()
        ctx["form_medicion"] = MedicionTemperaturaForm()
        temporada = (
            self.request.session.get("temporada_activa")
            or str(datetime.date.today().year)
        )
        pallets = _pallets_pendientes_camara_frio(temporada)
        ctx["pallets_pendientes"] = pallets
        from django.conf import settings
        if getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip() == "dataverse":
            from infrastructure.repository_factory import get_repositories
            ctx["pallets_data_json"] = _pallets_json_from_records(pallets, get_repositories())
        else:
            ctx["pallets_data_json"] = _pallets_data_json(temporada, pallets)
        return ctx

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action", "camara")
        temporada = _temporada(request)
        pallet_code = request.POST.get("pallet_code", "").strip()

        if action == "camara":
            form = CamaraFrioForm(request.POST)
            if form.is_valid():
                cd = form.cleaned_data
                payload = {
                    "temporada": temporada,
                    "pallet_code": pallet_code,
                    "operator_code": request.session.get("crf21_codigooperador", ""),
                    "source_system": "web",
                    "extra": {
                        "camara_numero": cd.get("camara_numero"),
                        "temperatura_camara": float(cd["temperatura_camara"]) if cd.get("temperatura_camara") else None,
                        "humedad_relativa": float(cd["humedad_relativa"]) if cd.get("humedad_relativa") else None,
                        "fecha_ingreso": _serialize_form_value(cd["fecha_ingreso"]) if cd.get("fecha_ingreso") else None,
                        "hora_ingreso": _serialize_form_value(cd["hora_ingreso"]) if cd.get("hora_ingreso") else None,
                        "destino_despacho": cd.get("destino_despacho"),
                    },
                }
                result = registrar_camara_frio(payload)
                _handle_result(request, result)
            else:
                messages.error(request, "Formulario de camara frio invalido.")

        elif action == "medicion":
            form = MedicionTemperaturaForm(request.POST)
            if form.is_valid():
                cd = form.cleaned_data
                payload = {
                    "temporada": temporada,
                    "pallet_code": pallet_code,
                    "operator_code": request.session.get("crf21_codigooperador", ""),
                    "source_system": "web",
                    "extra": {
                        "fecha": _serialize_form_value(cd["fecha"]) if cd.get("fecha") else None,
                        "hora": _serialize_form_value(cd["hora"]) if cd.get("hora") else None,
                        "temperatura_pallet": float(cd["temperatura_pallet"]) if cd.get("temperatura_pallet") else None,
                        "punto_medicion": cd.get("punto_medicion"),
                        "dentro_rango": cd.get("dentro_rango"),
                        "observaciones": cd.get("observaciones"),
                    },
                }
                result = registrar_medicion_temperatura(payload)
                _handle_result(request, result)
            else:
                messages.error(request, "Formulario de medicion invalido.")

        return redirect("operaciones:camaras")


# ---------------------------------------------------------------------------
# Consulta jefatura — helpers internos
# ---------------------------------------------------------------------------

CONSULTA_TAB_LOTES = "lotes"
CONSULTA_TAB_PALLETS = "pallets"
CONSULTA_TAB_DEFAULT = CONSULTA_TAB_LOTES
CONSULTA_ESTADO_EN_CAMARA_FRIO = "en_camara_frio"

CONSULTA_COLUMNAS_LOTES = [
    ("Temporada",          "temporada"),
    ("Lote (code)",        "lote_code"),
    ("Estado",             "estado_display"),
    ("Etapa actual",       "etapa"),
    ("Fecha conformacion", "fecha_conformacion"),
    ("Ult. Cambio Etapa",  "ultimo_cambio_etapa"),
    ("Productor (SAC)",    "productor"),
    ("CSG",                "codigo_sag_csg"),
    ("CSP",                "codigo_sag_csp"),
    ("SDP",                "codigo_sdp"),
    ("N Cuartel",          "numero_cuartel"),
    ("Nombre Cuartel",     "nombre_cuartel"),
    ("Tipo cultivo",       "tipo_cultivo"),
    ("Variedad",           "variedad"),
    ("Color",              "color"),
    ("Fecha cosecha",      "fecha_cosecha"),
    ("Bins",               "cantidad_bins"),
    ("Kg bruto",           "kilos_bruto"),
    ("Kg neto",            "kilos_neto"),
    ("Cajas producidas",   "cajas_producidas"),
]

CONSULTA_COLUMNAS_PALLETS = [
    ("Temporada",        "temporada"),
    ("Pallet (code)",    "pallet_code"),
    ("Lote relacionado", "lote_code"),
    ("Tipo caja",        "tipo_caja"),
    ("Cajas x pallet",   "cajas_por_pallet"),
    ("Peso total (kg)",  "peso_total"),
    ("Fecha pallet",     "fecha"),
    ("Productor",        "productor"),
    ("CSG",              "codigo_sag_csg"),
    ("CSP",              "codigo_sag_csp"),
    ("Variedad",         "variedad"),
    ("Color",            "color"),
    ("Fecha cosecha",    "fecha_cosecha"),
]

CONSULTA_CACHE_VERSION = 1
_CONSULTA_CACHE_REFRESH_LOCK = threading.Lock()
_CONSULTA_CACHE_REFRESH_IN_PROGRESS: set[str] = set()
_CONSULTA_CACHE_TYPE_KEY = "__consulta_cache_type__"
_CONSULTA_CACHE_TYPE_DATE = "date"
_CONSULTA_CACHE_TYPE_DATETIME = "datetime"


def _consulta_now_utc() -> datetime.datetime:
    return datetime.datetime.now(datetime.timezone.utc)


def _consulta_cache_file_path() -> Path:
    from django.conf import settings
    raw_path = getattr(
        settings,
        "CONSULTA_DATAVERSE_CACHE_FILE",
        str(Path(settings.BASE_DIR) / ".cache" / "consulta_dataverse.json"),
    )
    return Path(raw_path)


def _consulta_cache_ttl_seconds() -> int:
    from django.conf import settings
    ttl = getattr(settings, "CONSULTA_DATAVERSE_CACHE_TTL_SECONDS", 3600)
    try:
        ttl_int = int(ttl)
    except (TypeError, ValueError):
        ttl_int = 3600
    return max(60, ttl_int)


def _consulta_cache_empty_payload() -> dict:
    return {
        "version": CONSULTA_CACHE_VERSION,
        "backend": "dataverse",
        "updated_at": None,
        "temporadas": {},
    }


def _consulta_cache_json_default(value):
    if isinstance(value, datetime.datetime):
        return {
            _CONSULTA_CACHE_TYPE_KEY: _CONSULTA_CACHE_TYPE_DATETIME,
            "value": value.isoformat(),
        }
    if isinstance(value, datetime.date):
        return {
            _CONSULTA_CACHE_TYPE_KEY: _CONSULTA_CACHE_TYPE_DATE,
            "value": value.isoformat(),
        }
    raise TypeError(f"Tipo no serializable para cache: {type(value)!r}")


def _consulta_cache_json_hook(payload: dict):
    marker = payload.get(_CONSULTA_CACHE_TYPE_KEY)
    if not marker:
        return payload
    raw = payload.get("value")
    try:
        if marker == _CONSULTA_CACHE_TYPE_DATETIME:
            return datetime.datetime.fromisoformat(raw)
        if marker == _CONSULTA_CACHE_TYPE_DATE:
            return datetime.date.fromisoformat(raw)
    except Exception:
        return raw
    return payload


def _consulta_cache_read() -> dict:
    cache_path = _consulta_cache_file_path()
    if not cache_path.exists():
        return _consulta_cache_empty_payload()
    try:
        with cache_path.open("r", encoding="utf-8") as handler:
            data = json.load(handler, object_hook=_consulta_cache_json_hook)
        if not isinstance(data, dict):
            return _consulta_cache_empty_payload()
        data.setdefault("version", CONSULTA_CACHE_VERSION)
        data.setdefault("backend", "dataverse")
        data.setdefault("temporadas", {})
        return data
    except Exception as exc:
        logging.getLogger(__name__).warning("No se pudo leer cache consulta: %s", exc)
        return _consulta_cache_empty_payload()


def _consulta_cache_write(payload: dict) -> bool:
    cache_path = _consulta_cache_file_path()
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = None
    try:
        fd, tmp_path = tempfile.mkstemp(
            prefix=f"{cache_path.name}.",
            suffix=".tmp",
            dir=str(cache_path.parent),
        )
        with os.fdopen(fd, "w", encoding="utf-8") as handler:
            json.dump(
                payload,
                handler,
                ensure_ascii=False,
                default=_consulta_cache_json_default,
            )
        os.replace(tmp_path, cache_path)
        return True
    except Exception as exc:
        logging.getLogger(__name__).warning("No se pudo escribir cache consulta: %s", exc)
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass
        return False


def _consulta_cache_updated_at(payload: dict) -> datetime.datetime | None:
    raw = payload.get("updated_at")
    if isinstance(raw, datetime.datetime):
        if raw.tzinfo is None:
            return raw.replace(tzinfo=datetime.timezone.utc)
        return raw.astimezone(datetime.timezone.utc)
    if isinstance(raw, str):
        try:
            parsed = datetime.datetime.fromisoformat(raw)
            if parsed.tzinfo is None:
                return parsed.replace(tzinfo=datetime.timezone.utc)
            return parsed.astimezone(datetime.timezone.utc)
        except Exception:
            return None
    return None


def _consulta_cache_temporada(payload: dict, temporada: str) -> dict | None:
    temporadas = payload.get("temporadas") or {}
    data = temporadas.get(str(temporada))
    if not isinstance(data, dict):
        return None
    lotes = data.get("lotes")
    pallets = data.get("pallets")
    if not isinstance(lotes, list) or not isinstance(pallets, list):
        return None
    return data


def _consulta_cache_is_fresh(payload: dict) -> bool:
    updated_at = _consulta_cache_updated_at(payload)
    if not updated_at:
        return False
    age = (_consulta_now_utc() - updated_at).total_seconds()
    return age <= _consulta_cache_ttl_seconds()


def _consulta_cache_invalidate() -> None:
    """
    Invalida la cache JSON de consulta eliminando el archivo.
    Debe llamarse despues de cualquier mutacion que cambie datos visibles en reporteria
    (agregar bin, cerrar lote, desverdizado, ingreso packing, pallet).
    La proxima carga de _consulta_dataset reconstruira la cache desde Dataverse.
    """
    try:
        path = _consulta_cache_file_path()
        if path.exists():
            path.unlink()
    except Exception:
        pass


def _consulta_dataverse_dataset_live(temporada: str) -> dict:
    lotes = _lotes_enriquecidos_dataverse("", "")
    pallets = _pallets_enriquecidos_dataverse(temporada, "", "")
    lotes_rows = [{k: v for k, v in item.items() if k != "lote"} for item in lotes]
    pallets_rows = [{k: v for k, v in item.items() if k != "pallet"} for item in pallets]
    return {"lotes": lotes_rows, "pallets": pallets_rows}


def _consulta_dataverse_cache_refresh_now(temporada: str) -> tuple[dict, datetime.datetime, bool]:
    data = _consulta_dataverse_dataset_live(temporada)
    payload = _consulta_cache_read()
    updated_at = _consulta_now_utc()
    payload["version"] = CONSULTA_CACHE_VERSION
    payload["backend"] = "dataverse"
    payload["updated_at"] = updated_at
    payload.setdefault("temporadas", {})
    payload["temporadas"][str(temporada)] = data
    wrote = _consulta_cache_write(payload)
    return data, updated_at, wrote


def _consulta_dataverse_cache_refresh_background(temporada: str) -> bool:
    key = str(temporada)
    with _CONSULTA_CACHE_REFRESH_LOCK:
        if key in _CONSULTA_CACHE_REFRESH_IN_PROGRESS:
            return False
        _CONSULTA_CACHE_REFRESH_IN_PROGRESS.add(key)

    def _worker():
        try:
            _consulta_dataverse_cache_refresh_now(temporada)
        except Exception as exc:
            logging.getLogger(__name__).warning(
                "Fallo refresco en segundo plano de consulta (temporada=%s): %s",
                temporada,
                exc,
            )
        finally:
            with _CONSULTA_CACHE_REFRESH_LOCK:
                _CONSULTA_CACHE_REFRESH_IN_PROGRESS.discard(key)

    threading.Thread(target=_worker, daemon=True).start()
    return True


def _filtrar_lotes_consulta(rows: list, filtro_productor: str, filtro_estado: str) -> list:
    resultado = []
    filtro_productor_lc = (filtro_productor or "").lower()
    for raw in rows or []:
        item = dict(raw)
        estado = (item.get("estado") or "").strip()
        etapa = (item.get("etapa") or "").strip()
        productor = (item.get("productor") or "").strip()
        if filtro_productor_lc and filtro_productor_lc not in productor.lower():
            continue
        if not _coincide_filtro_estado_lote(filtro_estado, estado=estado, etapa=etapa):
            continue
        resultado.append(item)
    return resultado


def _filtrar_pallets_consulta(rows: list, filtro_productor: str, filtro_estado: str) -> list:
    resultado = []
    filtro_productor_lc = (filtro_productor or "").lower()
    for raw in rows or []:
        item = dict(raw)
        productor = (item.get("productor") or "").strip()
        if filtro_productor_lc and filtro_productor_lc not in productor.lower():
            continue
        if not _coincide_filtro_estado_pallet(
            filtro_estado,
            en_camara_frio=bool(item.get("en_camara_frio")),
            lote_estado=(item.get("estado") or "").strip(),
            lote_etapa=(item.get("etapa") or "").strip(),
        ):
            continue
        resultado.append(item)
    return resultado


def _consulta_dataset(
    temporada: str,
    filtro_productor: str,
    filtro_estado: str,
    *,
    fecha_desde: str = "",
    fecha_hasta: str = "",
    force_refresh: bool = False,
) -> tuple[list, list, dict]:
    from django.conf import settings
    backend = getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip()
    if backend != "dataverse":
        lotes = _filtrar_por_fecha(
            _lotes_enriquecidos_qs(temporada, filtro_productor, filtro_estado),
            fecha_desde, fecha_hasta,
        )
        pallets = _filtrar_por_fecha(
            _pallets_enriquecidos_qs(temporada, filtro_productor, filtro_estado),
            fecha_desde, fecha_hasta,
        )
        return (
            lotes,
            pallets,
            {
                "is_dataverse_backend": False,
                "cache_source": "none",
                "cache_last_updated": None,
                "cache_is_stale": False,
                "background_refresh_started": False,
                "refreshed_now": False,
                "warning_message": "",
            },
        )

    cache_payload = _consulta_cache_read()
    cache_rows = _consulta_cache_temporada(cache_payload, temporada)
    cache_updated_at = _consulta_cache_updated_at(cache_payload)
    cache_is_fresh = _consulta_cache_is_fresh(cache_payload)

    meta = {
        "is_dataverse_backend": True,
        "cache_source": "cache",
        "cache_last_updated": cache_updated_at,
        "cache_is_stale": bool(cache_rows) and (not cache_is_fresh),
        "background_refresh_started": False,
        "refreshed_now": False,
        "warning_message": "",
    }

    if force_refresh:
        try:
            live_data, updated_at, _ = _consulta_dataverse_cache_refresh_now(temporada)
            meta["cache_source"] = "live"
            meta["cache_last_updated"] = updated_at
            meta["cache_is_stale"] = False
            meta["refreshed_now"] = True
            lotes = _filtrar_por_fecha(_filtrar_lotes_consulta(live_data.get("lotes", []), filtro_productor, filtro_estado), fecha_desde, fecha_hasta)
            pallets = _filtrar_por_fecha(_filtrar_pallets_consulta(live_data.get("pallets", []), filtro_productor, filtro_estado), fecha_desde, fecha_hasta)
            return lotes, pallets, meta
        except Exception as exc:
            logging.getLogger(__name__).warning("Refresco manual consulta fallo: %s", exc)
            if cache_rows:
                meta["warning_message"] = (
                    "No fue posible sincronizar Dataverse. Se muestra cache disponible."
                )
                lotes = _filtrar_por_fecha(_filtrar_lotes_consulta(cache_rows.get("lotes", []), filtro_productor, filtro_estado), fecha_desde, fecha_hasta)
                pallets = _filtrar_por_fecha(_filtrar_pallets_consulta(cache_rows.get("pallets", []), filtro_productor, filtro_estado), fecha_desde, fecha_hasta)
                return lotes, pallets, meta
            meta["warning_message"] = (
                "No fue posible sincronizar Dataverse y no existe cache disponible."
            )
            return [], [], meta

    if cache_rows and cache_is_fresh:
        lotes = _filtrar_por_fecha(_filtrar_lotes_consulta(cache_rows.get("lotes", []), filtro_productor, filtro_estado), fecha_desde, fecha_hasta)
        pallets = _filtrar_por_fecha(_filtrar_pallets_consulta(cache_rows.get("pallets", []), filtro_productor, filtro_estado), fecha_desde, fecha_hasta)
        return lotes, pallets, meta

    if cache_rows and not cache_is_fresh:
        meta["background_refresh_started"] = _consulta_dataverse_cache_refresh_background(temporada)
        lotes = _filtrar_por_fecha(_filtrar_lotes_consulta(cache_rows.get("lotes", []), filtro_productor, filtro_estado), fecha_desde, fecha_hasta)
        pallets = _filtrar_por_fecha(_filtrar_pallets_consulta(cache_rows.get("pallets", []), filtro_productor, filtro_estado), fecha_desde, fecha_hasta)
        return lotes, pallets, meta

    try:
        live_data, updated_at, _ = _consulta_dataverse_cache_refresh_now(temporada)
        meta["cache_source"] = "live"
        meta["cache_last_updated"] = updated_at
        meta["cache_is_stale"] = False
        lotes = _filtrar_por_fecha(_filtrar_lotes_consulta(live_data.get("lotes", []), filtro_productor, filtro_estado), fecha_desde, fecha_hasta)
        pallets = _filtrar_por_fecha(_filtrar_pallets_consulta(live_data.get("pallets", []), filtro_productor, filtro_estado), fecha_desde, fecha_hasta)
        return lotes, pallets, meta
    except Exception as exc:
        logging.getLogger(__name__).warning("Carga inicial Dataverse para consulta fallo: %s", exc)
        meta["warning_message"] = (
            "No fue posible consultar Dataverse y no hay cache disponible."
        )
        return [], [], meta


def _normalizar_tab_consulta(tab_raw: str) -> str:
    tab = (tab_raw or "").strip().lower()
    if tab in {CONSULTA_TAB_LOTES, CONSULTA_TAB_PALLETS}:
        return tab
    return CONSULTA_TAB_DEFAULT


def _consulta_estado_choices() -> list[tuple[str, str]]:
    choices = list(LotePlantaEstado.choices)
    choices.append((CONSULTA_ESTADO_EN_CAMARA_FRIO, "En Camara de Frio"))
    return choices


def _consulta_kpis(lotes: list, pallets: list) -> dict:
    """
    Calcula KPIs de kilos por etapa para la vista Control de Gestion.
    Opera sobre la lista ya filtrada, por lo que respeta los filtros activos.
    """
    kg_recepcionados = kg_mantencion = kg_desverdizado = kg_packing = 0.0
    kg_camara_frio = 0.0
    lotes_recepcionados = lotes_mantencion = lotes_desverdizado = lotes_packing = 0
    pallets_camara = cajas_camara = 0

    for item in lotes:
        kn    = float(item.get("kilos_neto") or 0)
        etapa = item.get("etapa", "")
        if etapa == "Recepcion":
            kg_recepcionados += kn
            lotes_recepcionados += 1
        elif etapa == "Mantencion":
            kg_mantencion += kn
            lotes_mantencion += 1
        elif etapa == "Desverdizado":
            kg_desverdizado += kn
            lotes_desverdizado += 1
        elif etapa in ("Ingreso Packing", "Packing / Proceso"):
            kg_packing += kn
            lotes_packing += 1

    for item in pallets:
        if item.get("estado") == CONSULTA_ESTADO_EN_CAMARA_FRIO:
            kg_camara_frio += float(item.get("peso_total") or 0)
            cajas_camara   += int(item.get("cajas_por_pallet") or 0)
            pallets_camara += 1

    return {
        "kg_recepcionados":    round(kg_recepcionados, 1),
        "lotes_recepcionados": lotes_recepcionados,
        "kg_mantencion":       round(kg_mantencion, 1),
        "lotes_mantencion":    lotes_mantencion,
        "kg_desverdizado":     round(kg_desverdizado, 1),
        "lotes_desverdizado":  lotes_desverdizado,
        "kg_packing":          round(kg_packing, 1),
        "lotes_packing":       lotes_packing,
        "kg_camara_frio":      round(kg_camara_frio, 1),
        "cajas_camara_frio":   cajas_camara,
        "pallets_camara_frio": pallets_camara,
    }


def _estado_display_consulta(estado: str, *, fallback: str = "") -> str:
    if estado == CONSULTA_ESTADO_EN_CAMARA_FRIO:
        return "En Camara de Frio"
    if not estado:
        return fallback or "—"
    return dict(LotePlantaEstado.choices).get(estado, fallback or estado)


def _coincide_filtro_estado_lote(filtro_estado: str, *, estado: str, etapa: str) -> bool:
    if not filtro_estado:
        return True
    if filtro_estado == CONSULTA_ESTADO_EN_CAMARA_FRIO:
        return etapa == "Camara Frio"
    if filtro_estado == LotePlantaEstado.ABIERTO:
        return etapa == "Recepcion" or estado == LotePlantaEstado.ABIERTO
    if filtro_estado == LotePlantaEstado.CERRADO:
        return etapa != "Recepcion" or estado == LotePlantaEstado.CERRADO
    return estado == filtro_estado


def _coincide_filtro_estado_pallet(
    filtro_estado: str,
    *,
    en_camara_frio: bool,
    lote_estado: str,
    lote_etapa: str,
) -> bool:
    if not filtro_estado:
        return True
    if filtro_estado == CONSULTA_ESTADO_EN_CAMARA_FRIO:
        return en_camara_frio
    if en_camara_frio and filtro_estado == LotePlantaEstado.CERRADO:
        return True
    if en_camara_frio:
        return False
    return _coincide_filtro_estado_lote(
        filtro_estado,
        estado=lote_estado or "",
        etapa=lote_etapa or "",
    )


def _query_consulta(
    tab: str,
    filtro_productor: str,
    filtro_estado: str,
    *,
    fecha_desde: str = "",
    fecha_hasta: str = "",
    refresh: bool = False,
) -> str:
    params = {"tab": _normalizar_tab_consulta(tab)}
    if filtro_productor:
        params["productor"] = filtro_productor
    if filtro_estado:
        params["estado"] = filtro_estado
    if fecha_desde:
        params["fecha_desde"] = fecha_desde
    if fecha_hasta:
        params["fecha_hasta"] = fecha_hasta
    if refresh:
        params["refresh"] = "1"
    return urlencode(params)


def _url_consulta(
    tab: str,
    filtro_productor: str = "",
    filtro_estado: str = "",
    *,
    fecha_desde: str = "",
    fecha_hasta: str = "",
    refresh: bool = False,
) -> str:
    base = reverse("operaciones:consulta")
    qs = _query_consulta(tab, filtro_productor, filtro_estado, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, refresh=refresh)
    return f"{base}?{qs}" if qs else base


def _filtros_consulta_request(request, default_tab: str = CONSULTA_TAB_DEFAULT) -> tuple[str, str, str]:
    tab = _normalizar_tab_consulta(request.GET.get("tab", default_tab))
    filtro_productor = request.GET.get("productor", "").strip()
    filtro_estado = request.GET.get("estado", "").strip()
    return tab, filtro_productor, filtro_estado


def _fechas_consulta_request(request) -> tuple[str, str]:
    """Extrae fecha_desde y fecha_hasta (YYYY-MM-DD) del request GET. Devuelve ('', '') si no están."""
    def _clean_date(raw: str) -> str:
        raw = (raw or "").strip()
        if not raw:
            return ""
        try:
            datetime.date.fromisoformat(raw)
            return raw
        except ValueError:
            return ""
    return _clean_date(request.GET.get("fecha_desde", "")), _clean_date(request.GET.get("fecha_hasta", ""))


def _filtrar_por_fecha(rows: list, fecha_desde: str, fecha_hasta: str) -> list:
    """Filtra una lista de dicts con clave 'fecha' (date|datetime|None) por rango de fechas."""
    if not fecha_desde and not fecha_hasta:
        return rows
    desde = datetime.date.fromisoformat(fecha_desde) if fecha_desde else None
    hasta = datetime.date.fromisoformat(fecha_hasta) if fecha_hasta else None
    resultado = []
    for item in rows:
        f = item.get("fecha")
        if f is None:
            continue
        if isinstance(f, datetime.datetime):
            f = f.date()
        if desde and f < desde:
            continue
        if hasta and f > hasta:
            continue
        resultado.append(item)
    return resultado


def _refresh_consulta_request(request) -> bool:
    raw = (request.GET.get("refresh") or "").strip().lower()
    return raw in {"1", "true", "si", "yes"}


def _cache_last_updated_display(updated_at: datetime.datetime | None) -> str:
    if not updated_at:
        return ""
    try:
        local_dt = timezone.localtime(updated_at)
    except Exception:
        local_dt = updated_at
    return local_dt.strftime("%d/%m/%Y %H:%M")


def _fmt_ultimo_cambio(value) -> str:
    """
    Formatea el ultimo cambio de estado para la UI.
    - datetime → "d/m/Y H:i" (con hora real)
    - date → "d/m/Y" (solo fecha, fallback cuando no hay timestamp exacto)
    - None → ""
    Retorna siempre str para que el template pueda usar |default:"-" sin errores.
    """
    import datetime as _dt
    if value is None:
        return ""
    if isinstance(value, _dt.datetime):
        return value.strftime("%d/%m/%Y %H:%M")
    if isinstance(value, _dt.date):
        return value.strftime("%d/%m/%Y")
    return str(value)


def _float_or_none(value):
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _export_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, bool):
        return "Si" if value else "No"
    return value


_CAMPOS_BASE_LOTE_VACIO = {
    "productor": "",
    "tipo_cultivo": "",
    "variedad": "",
    "color": "",
    "fecha_cosecha": "",
    "codigo_sag_csg": "",
    "codigo_sag_csp": "",
    "codigo_sdp": "",
    "numero_cuartel": "",
    "nombre_cuartel": "",
}


def _campos_base_lote_prefetch(lote: Lote | None) -> dict:
    if not lote:
        return dict(_CAMPOS_BASE_LOTE_VACIO)
    primer_bin = None
    try:
        # list() usa el prefetch cache en lugar de emitir una query extra por lote
        bin_rels = list(lote.bin_lotes.all())
        primer_bin = bin_rels[0].bin if bin_rels else None
    except Exception:
        primer_bin = None
    if not primer_bin:
        return dict(_CAMPOS_BASE_LOTE_VACIO)
    return {
        "productor": primer_bin.codigo_productor or "",
        "tipo_cultivo": primer_bin.tipo_cultivo or "",
        "variedad": primer_bin.variedad_fruta or "",
        "color": primer_bin.color or "",
        "fecha_cosecha": str(primer_bin.fecha_cosecha) if primer_bin.fecha_cosecha else "",
        "codigo_sag_csg": getattr(primer_bin, "codigo_sag_csg", None) or "",
        "codigo_sag_csp": getattr(primer_bin, "codigo_sag_csp", None) or "",
        "codigo_sdp": getattr(primer_bin, "codigo_sdp", None) or "",
        "numero_cuartel": getattr(primer_bin, "numero_cuartel", None) or "",
        "nombre_cuartel": getattr(primer_bin, "nombre_cuartel", None) or "",
    }

def _es_jefatura(user) -> bool:
    """
    Compatibilidad: usa is_staff/is_superuser como fallback.
    Preferir is_jefatura(request) desde usuarios.permissions cuando se dispone del request.
    """
    return user.is_active and (user.is_staff or user.is_superuser)


def _lotes_enriquecidos_qs(temporada: str, filtro_productor: str, filtro_estado: str) -> list:
    """
    Devuelve lista de dicts con datos enriquecidos de lotes filtrados.
    Encapsulado para reutilizar en vista HTML y exportaciones.
    """
    from django.conf import settings
    backend = getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip()

    if backend == "dataverse":
        return _lotes_enriquecidos_dataverse(filtro_productor, filtro_estado)

    from django.db.models import Sum
    lotes_raw = list(
        Lote.objects
        .filter(temporada=temporada, is_active=True)
        .prefetch_related(
            "desverdizado",
            "ingreso_packing",
            "camara_mantencion",
            "bin_lotes__bin",
            "pallet_lotes__pallet__camara_frio",
            "registros_packing",
            "control_proceso_packing",
        )
        .order_by("-created_at")[:500]
    )
    lote_ids = [l.id for l in lotes_raw]
    cajas_por_lote: dict = {}
    if lote_ids:
        cajas_por_lote = dict(
            RegistroPacking.objects
            .filter(lote_id__in=lote_ids)
            .values("lote_id")
            .annotate(total=Sum("cantidad_cajas_producidas"))
            .values_list("lote_id", "total")
        )

    # B3: batch query para obtener la fecha del último evento de trazabilidad por lote
    from django.db.models import Max as _Max
    ultimo_evento_por_lote: dict = {}
    if lote_ids:
        ultimo_evento_por_lote = dict(
            RegistroEtapa.objects
            .filter(lote_id__in=lote_ids)
            .values("lote_id")
            .annotate(ultimo=_Max("occurred_at"))
            .values_list("lote_id", "ultimo")
        )

    resultado = []
    filtro_productor_lc = (filtro_productor or "").lower()
    for lote in lotes_raw:
        etapa = _etapa_lote(lote)
        if not _coincide_filtro_estado_lote(filtro_estado, estado=lote.estado, etapa=etapa):
            continue

        campos = _campos_base_lote_prefetch(lote)
        productor = campos["productor"]
        if filtro_productor_lc and filtro_productor_lc not in productor.lower():
            continue

        kilos_bruto, kilos_neto = _kilos_recientes_lote(lote, include_bin_fallback=False)

        # B3: fecha = último cambio de etapa real; fecha_conformacion separado
        _ultimo_dt = ultimo_evento_por_lote.get(lote.id)
        _ultimo_date = (
            _ultimo_dt.date() if isinstance(_ultimo_dt, datetime.datetime) else _ultimo_dt
        ) or lote.fecha_conformacion or (lote.created_at.date() if lote.created_at else None)

        resultado.append({
            "lote": lote,
            "lote_code": lote.lote_code,
            "estado": lote.estado,
            "estado_display": lote.get_estado_display(),
            "etapa": etapa,
            "cantidad_bins": lote.cantidad_bins,
            "kilos_bruto": kilos_bruto,
            "kilos_neto": kilos_neto,
            "cajas_producidas": cajas_por_lote.get(lote.id),
            "productor": productor,
            "codigo_sag_csg": campos["codigo_sag_csg"],
            "codigo_sag_csp": campos["codigo_sag_csp"],
            "codigo_sdp": campos["codigo_sdp"],
            "numero_cuartel": campos["numero_cuartel"],
            "nombre_cuartel": campos["nombre_cuartel"],
            "variedad": campos["variedad"],
            "tipo_cultivo": campos["tipo_cultivo"],
            "color": campos["color"],
            "fecha_cosecha": campos["fecha_cosecha"],
            "fecha": _ultimo_date,               # usado por _filtrar_por_fecha
            "fecha_conformacion": lote.fecha_conformacion,
            "ultimo_cambio_etapa": _ultimo_date,
        })
    return resultado


def _lotes_enriquecidos_dataverse(filtro_productor: str, filtro_estado: str) -> list:
    """
    Version Dataverse de _lotes_enriquecidos_qs.
    Usa etapa_actual persistida en crf21_etapa_actual como fuente principal.
    """
    try:
        from infrastructure.repository_factory import get_repositories
        from infrastructure.dataverse.repositories import resolve_etapa_lote
        repos = get_repositories()
        lotes = repos.lotes.list_recent(limit=500)
        lote_ids = [l.id for l in lotes if l.id]
        todos_bins_por_lote = repos.bins.all_bins_by_lotes(lote_ids) if lote_ids else {}

        # Batch fetch: 2 llamadas en lugar de hasta 1,000 llamadas individuales
        desv_por_lote: dict = repos.desverdizados.list_by_lotes(lote_ids) if lote_ids else {}
        ip_por_lote: dict = repos.ingresos_packing.list_by_lotes(lote_ids) if lote_ids else {}

        # B5: cajas producidas por lote (guardado: sum_cajas_by_lotes es Fase 2)
        cajas_por_lote_dv: dict = {}
        try:
            if (lote_ids
                    and hasattr(repos, "registros_packing")
                    and hasattr(repos.registros_packing, "sum_cajas_by_lotes")):
                cajas_por_lote_dv = repos.registros_packing.sum_cajas_by_lotes(lote_ids)
        except Exception:
            pass

        resultado = []
        filtro_productor_lc = (filtro_productor or "").lower()
        for lote in lotes:
            etapa = resolve_etapa_lote(lote)
            estado = (lote.estado or "").strip() or etapa
            if not _coincide_filtro_estado_lote(filtro_estado, estado=estado, etapa=etapa):
                continue
            bins = todos_bins_por_lote.get(lote.id) or []
            merged = _merge_bin_fields(bins)
            productor = merged["codigo_productor"] or lote.codigo_productor or ""
            if filtro_productor_lc and filtro_productor_lc not in productor.lower():
                continue

            _kb = _float_or_none(lote.kilos_bruto_conformacion)
            _kn = _float_or_none(lote.kilos_neto_conformacion)
            _d = desv_por_lote.get(lote.id)
            if _d:
                if _d.kilos_bruto_salida is not None:
                    _kb = float(_d.kilos_bruto_salida)
                if _d.kilos_neto_salida is not None:
                    _kn = float(_d.kilos_neto_salida)
            _ip = ip_por_lote.get(lote.id)
            if _ip:
                if _ip.kilos_bruto_ingreso_packing is not None:
                    _kb = float(_ip.kilos_bruto_ingreso_packing)
                if _ip.kilos_neto_ingreso_packing is not None:
                    _kn = float(_ip.kilos_neto_ingreso_packing)

            # B3: fecha_conformacion separada de ultimo_cambio_etapa
            _ultimo = lote.ultimo_cambio_estado_at or lote.fecha_conformacion

            resultado.append({
                "lote": lote,
                "lote_code": lote.lote_code,
                "estado": estado,
                "estado_display": _estado_display_consulta(estado, fallback=etapa),
                "etapa": etapa,
                "cantidad_bins": lote.cantidad_bins,
                "kilos_bruto": _kb,
                "kilos_neto": _kn,
                "cajas_producidas": cajas_por_lote_dv.get(lote.id),
                "productor": productor,
                "codigo_sag_csg": merged["codigo_sag_csg"] or "",
                "codigo_sag_csp": merged["codigo_sag_csp"] or "",
                "codigo_sdp": merged["codigo_sdp"] or "",
                "numero_cuartel": merged["numero_cuartel"] or "",
                "nombre_cuartel": merged["nombre_cuartel"] or "",
                "variedad": merged["variedad_fruta"] or "",
                "tipo_cultivo": merged["tipo_cultivo"] or "",
                "color": merged["color"] or "",
                "fecha_cosecha": str(merged["fecha_cosecha"]) if merged["fecha_cosecha"] else "",
                "fecha": _ultimo,               # usado por _filtrar_por_fecha
                "fecha_conformacion": lote.fecha_conformacion,
                "ultimo_cambio_etapa": _ultimo,
            })
        return resultado
    except Exception as exc:
        import logging
        logging.getLogger(__name__).warning("_lotes_enriquecidos_dataverse error: %s", exc)
        return []


def _pallets_enriquecidos_qs(temporada: str, filtro_productor: str, filtro_estado: str) -> list:
    from django.conf import settings
    backend = getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip()
    if backend == "dataverse":
        return _pallets_enriquecidos_dataverse(temporada, filtro_productor, filtro_estado)

    pallets = list(
        Pallet.objects
        .filter(temporada=temporada, is_active=True)
        .select_related("camara_frio")
        .prefetch_related("pallet_lotes__lote__bin_lotes__bin")
        .order_by("-created_at")[:500]
    )

    resultado = []
    filtro_productor_lc = (filtro_productor or "").lower()
    for pallet in pallets:
        relacion = pallet.pallet_lotes.first()
        lote = relacion.lote if relacion else None
        campos = _campos_base_lote_prefetch(lote)

        productor = campos["productor"]
        if filtro_productor_lc and filtro_productor_lc not in productor.lower():
            continue

        en_camara_frio = False
        try:
            en_camara_frio = bool(pallet.camara_frio)
        except Exception:
            en_camara_frio = False

        lote_etapa = _etapa_lote(lote) if lote else ""
        lote_estado = lote.estado if lote else ""
        if not _coincide_filtro_estado_pallet(
            filtro_estado,
            en_camara_frio=en_camara_frio,
            lote_estado=lote_estado,
            lote_etapa=lote_etapa,
        ):
            continue

        estado = CONSULTA_ESTADO_EN_CAMARA_FRIO if en_camara_frio else (lote_estado or "")
        estado_display = (
            "En Camara de Frio"
            if en_camara_frio
            else (lote.get_estado_display() if lote else "Sin lote relacionado")
        )
        etapa = "Camara Frio" if en_camara_frio else (lote_etapa or "Sin lote relacionado")
        resultado.append({
            "pallet": pallet,
            "pallet_code": pallet.pallet_code,
            "lote_code": lote.lote_code if lote else "",
            "tipo_caja": pallet.tipo_caja or "",
            "cajas_por_pallet": getattr(pallet, "cajas_por_pallet", None),
            "peso_total": _float_or_none(pallet.peso_total_kg),
            "productor": productor,
            "codigo_sag_csg": campos["codigo_sag_csg"],
            "codigo_sag_csp": campos["codigo_sag_csp"],
            "variedad": campos["variedad"],
            "color": campos["color"],
            "fecha_cosecha": campos["fecha_cosecha"],
            "estado": estado,
            "estado_display": estado_display,
            "etapa": etapa,
            "en_camara_frio": en_camara_frio,
            "fecha": pallet.fecha or (pallet.created_at.date() if pallet.created_at else None),
        })
    return resultado


def _pallets_enriquecidos_dataverse(temporada: str, filtro_productor: str, filtro_estado: str) -> list:
    try:
        from infrastructure.repository_factory import get_repositories
        from infrastructure.dataverse.repositories import resolve_etapa_lote

        repos = get_repositories()
        pallets = repos.pallets.list_recent(limit=500)
        contexto = _build_pallet_context_map_from_records(pallets, repos)

        resultado = []
        filtro_productor_lc = (filtro_productor or "").lower()
        lote_cache: dict[str, object] = {}

        for pallet in pallets:
            base = contexto.get(pallet.pallet_code, {})
            productor = (base.get("productor") or "").strip()
            if filtro_productor_lc and filtro_productor_lc not in productor.lower():
                continue

            lote_code = (base.get("lote_code") or "").strip()
            lote = None
            if lote_code:
                if lote_code not in lote_cache:
                    lote_cache[lote_code] = repos.lotes.find_by_code(temporada, lote_code)
                lote = lote_cache.get(lote_code)

            lote_estado = getattr(lote, "estado", "") if lote else ""
            lote_etapa = resolve_etapa_lote(lote, repos) if lote else ""

            en_camara_frio = False
            try:
                en_camara_frio = bool(repos.camara_frios.find_by_pallet(pallet.id))
            except Exception:
                en_camara_frio = False

            if not _coincide_filtro_estado_pallet(
                filtro_estado,
                en_camara_frio=en_camara_frio,
                lote_estado=lote_estado,
                lote_etapa=lote_etapa,
            ):
                continue

            estado = CONSULTA_ESTADO_EN_CAMARA_FRIO if en_camara_frio else (lote_estado or lote_etapa)
            estado_display = (
                "En Camara de Frio"
                if en_camara_frio
                else _estado_display_consulta(estado, fallback=lote_etapa or "Sin lote relacionado")
            )
            etapa = "Camara Frio" if en_camara_frio else (lote_etapa or "Sin lote relacionado")

            resultado.append({
                "pallet": pallet,
                "pallet_code": pallet.pallet_code,
                "lote_code": lote_code,
                "tipo_caja": pallet.tipo_caja or "",
                "cajas_por_pallet": getattr(pallet, "cajas_por_pallet", None),
                "peso_total": _float_or_none(pallet.peso_total_kg),
                "productor": productor,
                "codigo_sag_csg": base.get("codigo_sag_csg", ""),
                "codigo_sag_csp": base.get("codigo_sag_csp", ""),
                "variedad": base.get("variedad", ""),
                "color": base.get("color", ""),
                "fecha_cosecha": base.get("fecha_cosecha", ""),
                "estado": estado,
                "estado_display": estado_display,
                "etapa": etapa,
                "en_camara_frio": en_camara_frio,
                "fecha": getattr(pallet, "fecha", None),
            })
        return resultado
    except Exception as exc:
        import logging
        logging.getLogger(__name__).warning("_pallets_enriquecidos_dataverse error: %s", exc)
        return []


def _detalle_lote_context(temporada: str, lote_code: str) -> dict:
    from django.conf import settings
    backend = getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip()
    if backend == "dataverse":
        try:
            from infrastructure.repository_factory import get_repositories
            from infrastructure.dataverse.repositories import resolve_etapa_lote

            repos = get_repositories()
            lote = repos.lotes.find_by_code(temporada, lote_code)
            if not lote:
                return {}

            info = _lote_info(temporada, lote_code)
            bins = repos.bins.list_by_lote(lote.id) if lote.id else []
            bins_data = []
            for b in bins:
                bins_data.append({
                    "bin_code": getattr(b, "bin_code", ""),
                    "productor": getattr(b, "codigo_productor", ""),
                    "variedad": getattr(b, "variedad_fruta", ""),
                    "color": getattr(b, "color", ""),
                    "fecha_cosecha": getattr(b, "fecha_cosecha", None),
                    "kilos_bruto": _float_or_none(getattr(b, "kilos_bruto_ingreso", None)),
                    "kilos_neto": _float_or_none(getattr(b, "kilos_neto_ingreso", None)),
                    "tipo_cultivo": getattr(b, "tipo_cultivo", ""),
                })

            etapa = resolve_etapa_lote(lote, repos)
            estado = info.get("estado") or getattr(lote, "estado", "") or etapa

            # B9: packing data (guardado: requiere list_by_lote en repos, Fase 2)
            packing_data_dv = {}
            try:
                if (hasattr(repos, "registros_packing")
                        and hasattr(repos.registros_packing, "list_by_lote")):
                    registros_dv = repos.registros_packing.list_by_lote(lote.id) or []
                    if registros_dv:
                        packing_data_dv = {
                            "cajas_producidas": sum(getattr(r, "cantidad_cajas_producidas", 0) or 0 for r in registros_dv),
                            "merma_kg":         round(sum(float(getattr(r, "merma_seleccion_kg", 0) or 0) for r in registros_dv), 2),
                            "kg_comercial":     round(sum(float(getattr(r, "kilos_fruta_comercial", 0) or 0) for r in registros_dv), 2),
                            "registros": [],
                        }
            except Exception:
                pass

            return {
                "lote_code": lote.lote_code,
                "estado": estado,
                "estado_display": _estado_display_consulta(estado, fallback=etapa),
                "etapa": etapa,
                "cantidad_bins": lote.cantidad_bins,
                "kilos_bruto": info.get("kilos_bruto"),
                "kilos_neto": info.get("kilos_neto"),
                "via_desverdizado": bool(info.get("via_desverdizado")),
                "requiere_desverdizado": bool(info.get("requiere_desverdizado")),
                "productor": info.get("productor", ""),
                "variedad": info.get("variedad", ""),
                "tipo_cultivo": info.get("tipo_cultivo", ""),
                "color": info.get("color", ""),
                "fecha_cosecha": info.get("fecha_cosecha", ""),
                "fecha_conformacion": getattr(lote, "fecha_conformacion", None),
                "ultimo_cambio_etapa": _fmt_ultimo_cambio(
                    lote.ultimo_cambio_estado_at or lote.fecha_conformacion
                ),
                "bins": bins_data,
                "packing": packing_data_dv,
            }
        except Exception:
            return {}

    try:
        lote = Lote.objects.get(temporada=temporada, lote_code=lote_code, is_active=True)
    except Lote.DoesNotExist:
        return {}

    info = _lote_info(temporada, lote_code)
    etapa = _etapa_lote(lote)
    bin_lotes = list(lote.bin_lotes.select_related("bin").order_by("created_at"))
    bins_data = []
    for rel in bin_lotes:
        b = rel.bin
        bins_data.append({
            "bin_code": b.bin_code,
            "productor": b.codigo_productor or "",
            "variedad": b.variedad_fruta or "",
            "color": b.color or "",
            "fecha_cosecha": b.fecha_cosecha,
            "kilos_bruto": _float_or_none(b.kilos_bruto_ingreso),
            "kilos_neto": _float_or_none(b.kilos_neto_ingreso),
            "tipo_cultivo": b.tipo_cultivo or "",
        })

    # Desverdizado (opcional)
    desv_data = {}
    try:
        desv = lote.desverdizado
        desv_data = {
            "numero_camara": desv.numero_camara or "",
            "fecha_ingreso": desv.fecha_ingreso,
            "hora_ingreso": desv.hora_ingreso or "",
            "horas_desverdizado": desv.horas_desverdizado,
            "color_salida": desv.color_salida or "",
            "kilos_bruto_salida": _float_or_none(desv.kilos_bruto_salida),
            "kilos_neto_salida": _float_or_none(desv.kilos_neto_salida),
        }
    except Exception:
        pass

    # Ingreso a packing (obligatorio una vez en packing)
    ip_data = {}
    try:
        ip = lote.ingreso_packing
        ip_data = {
            "fecha_ingreso_packing": ip.fecha_ingreso,
            "hora_ingreso_packing": ip.hora_ingreso or "",
            "kilos_bruto_ingreso_packing": _float_or_none(ip.kilos_bruto_ingreso_packing),
            "kilos_neto_ingreso_packing": _float_or_none(ip.kilos_neto_ingreso_packing),
        }
    except Exception:
        pass

    # Pallet relacionado (primer pallet activo del lote)
    pallet_code = None
    try:
        pallet_lote = lote.pallet_lotes.select_related("pallet").order_by("-created_at").first()
        if pallet_lote:
            pallet_code = pallet_lote.pallet.pallet_code
    except Exception:
        pass

    # B9: resultados de packing (cajas, merma, kg comercial)
    packing_data = {}
    try:
        registros = list(lote.registros_packing.all())
        if registros:
            packing_data = {
                "cajas_producidas":  sum(r.cantidad_cajas_producidas or 0 for r in registros),
                "merma_kg":          round(sum(float(r.merma_seleccion_kg or 0) for r in registros), 2),
                "kg_comercial":      round(sum(float(r.kilos_fruta_comercial or 0) for r in registros), 2),
                "registros": [
                    {
                        "fecha":        r.fecha,
                        "linea":        r.linea_proceso or "",
                        "categoria":    r.categoria_calidad or "",
                        "calibre":      r.calibre or "",
                        "cajas":        r.cantidad_cajas_producidas,
                        "peso_prom_kg": _float_or_none(r.peso_promedio_caja_kg),
                        "merma_kg":     _float_or_none(r.merma_seleccion_kg),
                        "kg_comercial": _float_or_none(r.kilos_fruta_comercial),
                    }
                    for r in registros
                ],
            }
    except Exception:
        pass

    return {
        "lote_code": lote.lote_code,
        "estado": lote.estado,
        "estado_display": lote.get_estado_display(),
        "etapa": etapa,
        "cantidad_bins": lote.cantidad_bins,
        "kilos_bruto": info.get("kilos_bruto"),
        "kilos_neto": info.get("kilos_neto"),
        "via_desverdizado": bool(info.get("via_desverdizado")),
        "requiere_desverdizado": bool(info.get("requiere_desverdizado")),
        "productor": info.get("productor", ""),
        "variedad": info.get("variedad", ""),
        "tipo_cultivo": info.get("tipo_cultivo", ""),
        "color": info.get("color", ""),
        "fecha_cosecha": info.get("fecha_cosecha", ""),
        "fecha_conformacion": lote.fecha_conformacion,
        "ultimo_cambio_etapa": _fmt_ultimo_cambio(
            getattr(lote, "ultimo_cambio_estado_at", None) or lote.fecha_conformacion
        ),
        "bins": bins_data,
        "desverdizado": desv_data,
        "ingreso_packing": ip_data,
        "packing": packing_data,
        "pallet_code": pallet_code,
    }


def _detalle_pallet_context(temporada: str, pallet_code: str) -> dict:
    from django.conf import settings
    backend = getattr(settings, "PERSISTENCE_BACKEND", "sqlite").lower().strip()
    if backend == "dataverse":
        try:
            from infrastructure.repository_factory import get_repositories
            from infrastructure.dataverse.repositories import resolve_etapa_lote

            repos = get_repositories()
            pallet = repos.pallets.find_by_code(temporada, pallet_code)
            if not pallet:
                return {}

            base = _pallet_info(temporada, pallet_code)
            lote_code = (base.get("lote_code") or "").strip()
            lote = repos.lotes.find_by_code(temporada, lote_code) if lote_code else None
            lote_etapa = resolve_etapa_lote(lote, repos) if lote else ""
            lote_estado = getattr(lote, "estado", "") if lote else ""

            en_camara_frio = False
            try:
                en_camara_frio = bool(repos.camara_frios.find_by_pallet(pallet.id))
            except Exception:
                en_camara_frio = False

            estado = CONSULTA_ESTADO_EN_CAMARA_FRIO if en_camara_frio else (lote_estado or lote_etapa)
            estado_display = (
                "En Camara de Frio"
                if en_camara_frio
                else _estado_display_consulta(estado, fallback=lote_etapa or "Sin lote relacionado")
            )
            return {
                "pallet_code": pallet.pallet_code,
                "tipo_caja": pallet.tipo_caja or "",
                "peso_total": _float_or_none(pallet.peso_total_kg),
                "lote_code": lote_code,
                "productor": base.get("productor", ""),
                "variedad": base.get("variedad", ""),
                "color": base.get("color", ""),
                "fecha_cosecha": base.get("fecha_cosecha", ""),
                "estado": estado,
                "estado_display": estado_display,
                "etapa": "Camara Frio" if en_camara_frio else (lote_etapa or "Sin lote relacionado"),
                "en_camara_frio": en_camara_frio,
                "fecha": getattr(pallet, "fecha", None),
            }
        except Exception:
            return {}

    try:
        pallet = Pallet.objects.get(temporada=temporada, pallet_code=pallet_code, is_active=True)
    except Pallet.DoesNotExist:
        return {}

    relacion = pallet.pallet_lotes.select_related("lote").first()
    lote = relacion.lote if relacion else None
    campos = _campos_base_lote(lote) if lote else {
        "productor": "",
        "tipo_cultivo": "",
        "variedad": "",
        "color": "",
        "fecha_cosecha": "",
    }

    en_camara_frio = False
    try:
        en_camara_frio = bool(pallet.camara_frio)
    except Exception:
        en_camara_frio = False

    lote_etapa = _etapa_lote(lote) if lote else ""
    lote_estado = lote.estado if lote else ""
    estado = CONSULTA_ESTADO_EN_CAMARA_FRIO if en_camara_frio else lote_estado
    estado_display = (
        "En Camara de Frio"
        if en_camara_frio
        else (lote.get_estado_display() if lote else "Sin lote relacionado")
    )
    return {
        "pallet_code": pallet.pallet_code,
        "tipo_caja": pallet.tipo_caja or "",
        "peso_total": _float_or_none(pallet.peso_total_kg),
        "lote_code": lote.lote_code if lote else "",
        "productor": campos.get("productor", ""),
        "variedad": campos.get("variedad", ""),
        "color": campos.get("color", ""),
        "fecha_cosecha": campos.get("fecha_cosecha", ""),
        "estado": estado,
        "estado_display": estado_display,
        "etapa": "Camara Frio" if en_camara_frio else (lote_etapa or "Sin lote relacionado"),
        "en_camara_frio": en_camara_frio,
        "fecha": pallet.fecha or (pallet.created_at.date() if pallet.created_at else None),
    }


def _consulta_export_bundle(
    temporada: str,
    tab: str,
    filtro_productor: str,
    filtro_estado: str,
    *,
    fecha_desde: str = "",
    fecha_hasta: str = "",
    force_refresh: bool = False,
) -> tuple[list[tuple[str, str]], list[dict], str, dict]:
    tab_norm = _normalizar_tab_consulta(tab)
    lotes, pallets, meta = _consulta_dataset(
        temporada,
        filtro_productor,
        filtro_estado,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        force_refresh=force_refresh,
    )
    if tab_norm == CONSULTA_TAB_PALLETS:
        rows = [dict(item) for item in pallets]
        for item in rows:
            item["temporada"] = temporada
        return CONSULTA_COLUMNAS_PALLETS, rows, CONSULTA_TAB_PALLETS, meta

    rows = [dict(item) for item in lotes]
    for item in rows:
        item["temporada"] = temporada
    return CONSULTA_COLUMNAS_LOTES, rows, CONSULTA_TAB_LOTES, meta


class JefaturaRequiredMixin(UserPassesTestMixin):
    """Restringe el acceso a usuarios con rol Jefatura o Administrador."""
    def test_func(self):
        from usuarios.permissions import is_jefatura
        return is_jefatura(self.request)

    def handle_no_permission(self):
        messages.error(self.request, "Acceso denegado: se requiere rol Jefatura o Administrador.")
        return redirect("usuarios:portal")


# ---------------------------------------------------------------------------
# Consulta jefatura — vista HTML
# ---------------------------------------------------------------------------

class ConsultaJefaturaView(LoginRequiredMixin, JefaturaRequiredMixin, TemplateView):
    template_name = "operaciones/consulta.html"
    login_url = reverse_lazy("usuarios:login")
    raise_exception = True

    def test_func(self):
        """Solo jefatura y administradores (is_staff o is_superuser)."""
        user = self.request.user
        return user.is_active and (user.is_staff or user.is_superuser)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        temporada = self.request.session.get("temporada_activa") or str(datetime.date.today().year)
        tab_actual, filtro_productor, filtro_estado = _filtros_consulta_request(
            self.request,
            CONSULTA_TAB_DEFAULT,
        )
        fecha_desde, fecha_hasta = _fechas_consulta_request(self.request)
        force_refresh = _refresh_consulta_request(self.request)

        lotes, pallets, cache_meta = _consulta_dataset(
            temporada,
            filtro_productor,
            filtro_estado,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            force_refresh=force_refresh,
        )
        query_consulta = _query_consulta(tab_actual, filtro_productor, filtro_estado, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)

        ctx["page_title"] = "Control de Gestion"
        ctx["temporada"] = temporada
        ctx["filtro_productor"] = filtro_productor
        ctx["filtro_estado"] = filtro_estado
        ctx["fecha_desde"] = fecha_desde
        ctx["fecha_hasta"] = fecha_hasta
        ctx["tab_actual"] = tab_actual
        ctx["lotes"] = lotes
        ctx["pallets"] = pallets
        ctx["lotes_count"] = len(lotes)
        ctx["pallets_count"] = len(pallets)
        ctx["kpis"] = _consulta_kpis(lotes, pallets)
        ctx["estados_choices"] = _consulta_estado_choices()
        ctx["query_consulta"] = query_consulta
        ctx["tab_lotes_url"] = _url_consulta(CONSULTA_TAB_LOTES, filtro_productor, filtro_estado, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)
        ctx["tab_pallets_url"] = _url_consulta(CONSULTA_TAB_PALLETS, filtro_productor, filtro_estado, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta)
        ctx["refresh_url"] = _url_consulta(
            tab_actual,
            filtro_productor,
            filtro_estado,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            refresh=True,
        )
        ctx["export_csv_url"] = f"{reverse('operaciones:exportar_consulta')}?{query_consulta}"
        ctx["export_excel_url"] = f"{reverse('operaciones:exportar_consulta_excel')}?{query_consulta}"
        ctx["is_dataverse_backend"] = cache_meta.get("is_dataverse_backend", False)
        ctx["cache_is_stale"] = cache_meta.get("cache_is_stale", False)
        ctx["cache_source"] = cache_meta.get("cache_source", "none")
        ctx["cache_last_updated"] = cache_meta.get("cache_last_updated")
        ctx["cache_last_updated_display"] = _cache_last_updated_display(ctx["cache_last_updated"])
        ctx["background_refresh_started"] = cache_meta.get("background_refresh_started", False)
        ctx["refreshed_now"] = cache_meta.get("refreshed_now", False)

        warning_message = cache_meta.get("warning_message", "").strip()
        if warning_message:
            messages.warning(self.request, warning_message)
        elif force_refresh and cache_meta.get("refreshed_now"):
            messages.success(self.request, "Consulta sincronizada desde Dataverse.")
        return ctx


# ---------------------------------------------------------------------------
# Exportación CSV — mismos filtros que la vista
# ---------------------------------------------------------------------------

class ConsultaLoteDetalleView(LoginRequiredMixin, JefaturaRequiredMixin, TemplateView):
    template_name = "operaciones/consulta_lote_detalle.html"
    login_url = reverse_lazy("usuarios:login")
    raise_exception = True

    def get(self, request, *args, **kwargs):
        temporada = request.session.get("temporada_activa") or str(datetime.date.today().year)
        lote_code = (kwargs.get("lote_code") or "").strip()
        tab, filtro_productor, filtro_estado = _filtros_consulta_request(
            request,
            CONSULTA_TAB_LOTES,
        )
        detalle = _detalle_lote_context(temporada, lote_code)
        if not detalle:
            messages.error(request, f"Lote '{lote_code}' no encontrado.")
            return redirect(_url_consulta(tab, filtro_productor, filtro_estado))

        context = self.get_context_data(
            temporada=temporada,
            detalle=detalle,
            back_url=_url_consulta(tab, filtro_productor, filtro_estado),
            query_consulta=_query_consulta(tab, filtro_productor, filtro_estado),
        )
        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        detalle = kwargs["detalle"]
        ctx["page_title"] = f"Detalle Lote {detalle['lote_code']}"
        ctx["temporada"] = kwargs["temporada"]
        ctx["detalle"] = detalle
        ctx["back_url"] = kwargs["back_url"]
        ctx["query_consulta"] = kwargs["query_consulta"]
        return ctx


class ConsultaPalletDetalleView(LoginRequiredMixin, JefaturaRequiredMixin, TemplateView):
    template_name = "operaciones/consulta_pallet_detalle.html"
    login_url = reverse_lazy("usuarios:login")
    raise_exception = True

    def get(self, request, *args, **kwargs):
        temporada = request.session.get("temporada_activa") or str(datetime.date.today().year)
        pallet_code = (kwargs.get("pallet_code") or "").strip()
        tab, filtro_productor, filtro_estado = _filtros_consulta_request(
            request,
            CONSULTA_TAB_PALLETS,
        )
        detalle = _detalle_pallet_context(temporada, pallet_code)
        if not detalle:
            messages.error(request, f"Pallet '{pallet_code}' no encontrado.")
            return redirect(_url_consulta(tab, filtro_productor, filtro_estado))

        context = self.get_context_data(
            temporada=temporada,
            detalle=detalle,
            back_url=_url_consulta(tab, filtro_productor, filtro_estado),
            query_consulta=_query_consulta(tab, filtro_productor, filtro_estado),
            query_lote=_query_consulta(CONSULTA_TAB_LOTES, filtro_productor, filtro_estado),
        )
        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        detalle = kwargs["detalle"]
        ctx["page_title"] = f"Detalle Pallet {detalle['pallet_code']}"
        ctx["temporada"] = kwargs["temporada"]
        ctx["detalle"] = detalle
        ctx["back_url"] = kwargs["back_url"]
        ctx["query_consulta"] = kwargs["query_consulta"]
        ctx["query_lote"] = kwargs["query_lote"]
        return ctx


class ExportarConsultaCSVView(LoginRequiredMixin, JefaturaRequiredMixin, View):
    login_url = reverse_lazy("usuarios:login")

    def get(self, request):
        temporada = request.session.get("temporada_activa") or str(datetime.date.today().year)
        tab, filtro_productor, filtro_estado = _filtros_consulta_request(
            request,
            CONSULTA_TAB_DEFAULT,
        )
        fecha_desde, fecha_hasta = _fechas_consulta_request(request)
        force_refresh = _refresh_consulta_request(request)
        columnas, rows, tab_norm, _ = _consulta_export_bundle(
            temporada,
            tab,
            filtro_productor,
            filtro_estado,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            force_refresh=force_refresh,
        )

        fecha_hoy = datetime.date.today().strftime("%Y%m%d")
        filename = f"consulta_{tab_norm}_{temporada}_{fecha_hoy}.csv"

        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response.write("\ufeff")

        writer = csv.writer(response)
        writer.writerow([header for header, _ in columnas])
        for item in rows:
            writer.writerow([_export_cell_value(item.get(key, "")) for _, key in columnas])
        return response


class ExportarConsultaExcelView(LoginRequiredMixin, JefaturaRequiredMixin, View):
    login_url = reverse_lazy("usuarios:login")

    def get(self, request):
        temporada = request.session.get("temporada_activa") or str(datetime.date.today().year)
        _, filtro_productor, filtro_estado = _filtros_consulta_request(
            request,
            CONSULTA_TAB_DEFAULT,
        )
        fecha_desde, fecha_hasta = _fechas_consulta_request(request)
        force_refresh = _refresh_consulta_request(request)

        lotes, pallets, _ = _consulta_dataset(
            temporada,
            filtro_productor,
            filtro_estado,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            force_refresh=force_refresh,
        )

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter
        except Exception:
            messages.error(
                request,
                "No fue posible generar Excel porque openpyxl no esta instalado en este entorno.",
            )
            return redirect(_url_consulta("lotes", filtro_productor, filtro_estado))

        def _write_sheet(ws, columnas, rows):
            headers = [header for header, _ in columnas]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for item in rows:
                ws.append([_export_cell_value(item.get(key, "")) for _, key in columnas])
            for idx, (header, _) in enumerate(columnas, start=1):
                width = min(max(len(header) + 4, 12), 40)
                ws.column_dimensions[get_column_letter(idx)].width = width

        wb = Workbook()
        ws_lotes = wb.active
        ws_lotes.title = "Lotes"
        _write_sheet(ws_lotes, CONSULTA_COLUMNAS_LOTES, lotes)

        ws_pallets = wb.create_sheet(title="Pallets")
        _write_sheet(ws_pallets, CONSULTA_COLUMNAS_PALLETS, pallets)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        fecha_hoy = datetime.date.today().strftime("%Y%m%d")
        filename = f"control_gestion_{temporada}_{fecha_hoy}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def _handle_result(request, result) -> None:
    if result.ok:
        messages.success(request, result.message)
    else:
        for err in result.errors:
            messages.error(request, err)

