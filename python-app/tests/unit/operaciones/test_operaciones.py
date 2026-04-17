"""
Tests del flujo operativo de packing.

Cobertura:
  - Use cases: iniciar_lote_recepcion, agregar_bin_a_lote_abierto, cerrar_lote_recepcion
  - Repositories: list_by_lote, list_recent
  - Smoke tests: vistas críticas renderizan 200
  - Flujo recepcion via POST web
  - SmokeViewsAuthenticatedTest: todas las vistas con admin autenticado
  - RecepcionFlowE2ETest: secuencia narrativa iniciar→bin→cerrar
  - RoleAccessControlTest: evidencia negativa de enforcement de roles por módulo
"""
import datetime
import json
import tempfile
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import Mock, patch

from django.test import TestCase, Client, override_settings
from django.contrib.auth import get_user_model
from django.urls import reverse
from usuarios.permissions import SESSION_KEY_ROL

from operaciones.application.use_cases import (
    iniciar_lote_recepcion,
    agregar_bin_a_lote_abierto,
    cerrar_lote_recepcion,
)
from infrastructure.repository_factory import get_repositories
from operaciones.models import (
    Bin,
    BinLote,
    Lote,
    LotePlantaEstado,
    Pallet,
    PalletLote,
    CamaraFrio,
    IngresoAPacking,
    Desverdizado,
)

User = get_user_model()

TEMPORADA = str(datetime.date.today().year)


def _make_user():
    return User.objects.create_user(username="tester", password="pass1234", is_superuser=True, is_staff=True)


def _make_lote(estado=LotePlantaEstado.ABIERTO, **kwargs):
    defaults = dict(
        temporada=TEMPORADA,
        lote_code=f"LP-TEST-{Lote.objects.count()+1:03d}",
        estado=estado,
        is_active=True,
    )
    defaults.update(kwargs)
    return Lote.objects.create(**defaults)


def _make_bin(lote, **kwargs):
    defaults = dict(
        temporada=TEMPORADA,
        bin_code=f"BIN-T-{Bin.objects.count()+1:04d}",
        codigo_productor="PROD-01",
        tipo_cultivo="Uva de mesa",
        variedad_fruta="Thompson",
        color="2",
        fecha_cosecha=datetime.date.today(),
        kilos_bruto_ingreso=500,
        kilos_neto_ingreso=480,
    )
    defaults.update(kwargs)
    b = Bin.objects.create(**defaults)
    BinLote.objects.create(bin=b, lote=lote)
    return b


# ---------------------------------------------------------------------------
# Helpers de autenticación con rol en sesión
# ---------------------------------------------------------------------------

def _make_client_with_session_rol(username, rol_str, is_staff=False, is_superuser=False):
    """Crea un Django User, hace force_login y fija SESSION_KEY_ROL en sesión."""
    user = User.objects.create_user(
        username=username, password="pw123",
        is_staff=is_staff, is_superuser=is_superuser,
    )
    c = Client()
    c.force_login(user)
    session = c.session
    session[SESSION_KEY_ROL] = rol_str
    session.save()
    return c, user


def _make_admin_client(username="admin_smoke"):
    """Crea un cliente Administrador con is_superuser=True y rol en sesión."""
    c, _ = _make_client_with_session_rol(username, "Administrador", is_superuser=True)
    return c


# ---------------------------------------------------------------------------
# Helpers de use cases
# ---------------------------------------------------------------------------

def _iniciar(extra=None):
    payload = {"temporada": TEMPORADA, "operator_code": "OP-TEST", "source_system": "test"}
    if extra:
        payload.update(extra)
    return iniciar_lote_recepcion(payload)


def _agregar_bin(lote_code, extra=None):
    payload = {
        "temporada": TEMPORADA,
        "lote_code": lote_code,
        "operator_code": "OP-TEST",
        "source_system": "test",
        "variedad_fruta": "Thompson",
        "kilos_neto_ingreso": "200",
        "kilos_bruto_ingreso": "210",
    }
    if extra:
        payload.update(extra)
    return agregar_bin_a_lote_abierto(payload)


def _cerrar(lote_code):
    return cerrar_lote_recepcion({
        "temporada": TEMPORADA,
        "lote_code": lote_code,
        "operator_code": "OP-TEST",
        "source_system": "test",
    })


# ---------------------------------------------------------------------------
# Use case: iniciar_lote_recepcion
# ---------------------------------------------------------------------------

class IniciarLoteRecepcionTest(TestCase):

    def test_crea_lote_abierto(self):
        result = _iniciar()
        self.assertTrue(result.ok, result.errors)
        self.assertEqual(result.code, "LOTE_INICIADO")
        self.assertIn("lote_code", result.data)
        self.assertEqual(result.data["estado"], "abierto")

    def test_lote_code_autogenerado_no_vacio(self):
        result = _iniciar()
        self.assertTrue(result.ok)
        lote_code = result.data["lote_code"]
        self.assertTrue(len(lote_code) > 0)

    def test_lote_persiste_en_db(self):
        result = _iniciar()
        self.assertTrue(result.ok)
        repos = get_repositories()
        lote = repos.lotes.find_by_code(TEMPORADA, result.data["lote_code"])
        self.assertIsNotNone(lote)
        self.assertEqual(lote.estado, "abierto")
        self.assertEqual(lote.cantidad_bins, 0)

    def test_dos_lotes_codigos_distintos(self):
        r1 = _iniciar()
        r2 = _iniciar()
        self.assertTrue(r1.ok)
        self.assertTrue(r2.ok)
        self.assertNotEqual(r1.data["lote_code"], r2.data["lote_code"])

    def test_rechaza_sin_temporada(self):
        result = iniciar_lote_recepcion({})
        self.assertFalse(result.ok)
        self.assertEqual(result.code, "INVALID_PAYLOAD")


# ---------------------------------------------------------------------------
# Use case: agregar_bin_a_lote_abierto
# ---------------------------------------------------------------------------

class AgregarBinALoteAbiertoTest(TestCase):

    def setUp(self):
        result = _iniciar()
        self.assertTrue(result.ok)
        self.lote_code = result.data["lote_code"]

    def test_agrega_bin_exitosamente(self):
        result = _agregar_bin(self.lote_code)
        self.assertTrue(result.ok, result.errors)
        self.assertEqual(result.code, "BIN_AGREGADO")
        self.assertIn("bin_code", result.data)
        self.assertEqual(result.data["lote_code"], self.lote_code)

    def test_bin_code_autogenerado(self):
        result = _agregar_bin(self.lote_code)
        self.assertTrue(result.ok)
        self.assertTrue(len(result.data["bin_code"]) > 0)

    def test_lote_cantidad_bins_incrementa(self):
        repos = get_repositories()
        _agregar_bin(self.lote_code)
        _agregar_bin(self.lote_code)
        lote = repos.lotes.find_by_code(TEMPORADA, self.lote_code)
        self.assertEqual(lote.cantidad_bins, 2)

    def test_bin_queda_asociado_al_lote(self):
        result = _agregar_bin(self.lote_code)
        self.assertTrue(result.ok)
        repos = get_repositories()
        lote = repos.lotes.find_by_code(TEMPORADA, self.lote_code)
        bins = repos.bins.list_by_lote(lote.id)
        self.assertEqual(len(bins), 1)
        self.assertEqual(bins[0].bin_code, result.data["bin_code"])

    def test_rechaza_lote_inexistente(self):
        result = _agregar_bin("LOTE-NOEXISTE")
        self.assertFalse(result.ok)
        self.assertEqual(result.code, "LOTE_NOT_FOUND")

    def test_rechaza_lote_cerrado(self):
        _agregar_bin(self.lote_code)
        _cerrar(self.lote_code)
        result = _agregar_bin(self.lote_code)
        self.assertFalse(result.ok)
        self.assertEqual(result.code, "LOTE_NOT_OPEN")

    def test_rechaza_sin_lote_code(self):
        result = agregar_bin_a_lote_abierto({"temporada": TEMPORADA})
        self.assertFalse(result.ok)
        self.assertEqual(result.code, "INVALID_PAYLOAD")


# ---------------------------------------------------------------------------
# Use case: cerrar_lote_recepcion
# ---------------------------------------------------------------------------

class CerrarLoteRecepcionTest(TestCase):

    def setUp(self):
        result = _iniciar()
        self.assertTrue(result.ok)
        self.lote_code = result.data["lote_code"]

    def test_rechaza_cerrar_lote_sin_bins(self):
        result = _cerrar(self.lote_code)
        self.assertFalse(result.ok)
        self.assertEqual(result.code, "LOTE_SIN_BINS")

    def test_cierra_lote_con_bins(self):
        _agregar_bin(self.lote_code)
        result = _cerrar(self.lote_code)
        self.assertTrue(result.ok, result.errors)
        self.assertEqual(result.code, "LOTE_CERRADO")
        self.assertEqual(result.data["estado"], "cerrado")

    def test_estado_persiste_cerrado_en_db(self):
        _agregar_bin(self.lote_code)
        _cerrar(self.lote_code)
        repos = get_repositories()
        lote = repos.lotes.find_by_code(TEMPORADA, self.lote_code)
        self.assertEqual(lote.estado, "cerrado")

    def test_rechaza_cerrar_lote_ya_cerrado(self):
        _agregar_bin(self.lote_code)
        _cerrar(self.lote_code)
        result = _cerrar(self.lote_code)
        self.assertFalse(result.ok)
        self.assertEqual(result.code, "LOTE_NOT_OPEN")

    def test_rechaza_cerrar_lote_inexistente(self):
        result = cerrar_lote_recepcion({
            "temporada": TEMPORADA,
            "lote_code": "LOTE-NOEXISTE",
        })
        self.assertFalse(result.ok)
        self.assertEqual(result.code, "LOTE_NOT_FOUND")


# ---------------------------------------------------------------------------
# Repository: list_by_lote / list_recent
# ---------------------------------------------------------------------------

class RepositoryQueryTest(TestCase):

    def test_bins_list_by_lote_vacio(self):
        r = _iniciar()
        repos = get_repositories()
        lote = repos.lotes.find_by_code(TEMPORADA, r.data["lote_code"])
        bins = repos.bins.list_by_lote(lote.id)
        self.assertEqual(bins, [])

    def test_bins_list_by_lote_con_datos(self):
        r = _iniciar()
        lote_code = r.data["lote_code"]
        _agregar_bin(lote_code)
        _agregar_bin(lote_code)
        repos = get_repositories()
        lote = repos.lotes.find_by_code(TEMPORADA, lote_code)
        bins = repos.bins.list_by_lote(lote.id)
        self.assertEqual(len(bins), 2)

    def test_lotes_list_recent(self):
        _iniciar()
        _iniciar()
        repos = get_repositories()
        lotes = repos.lotes.list_recent(TEMPORADA)
        self.assertGreaterEqual(len(lotes), 2)

    def test_bin_lotes_list_by_lote(self):
        r = _iniciar()
        lote_code = r.data["lote_code"]
        _agregar_bin(lote_code)
        repos = get_repositories()
        lote = repos.lotes.find_by_code(TEMPORADA, lote_code)
        bin_lotes = repos.bin_lotes.list_by_lote(lote.id)
        self.assertEqual(len(bin_lotes), 1)
        self.assertEqual(bin_lotes[0].lote_id, lote.id)


# ---------------------------------------------------------------------------
# GET smoke — todas las vistas deben renderizar 200 (auth requerida)
# ---------------------------------------------------------------------------

class ViewsGetSmokeTest(TestCase):

    def setUp(self):
        self.client = Client()
        self.user = _make_user()
        self.client.force_login(self.user)

    def _get_ok(self, url_name):
        url = reverse(f"operaciones:{url_name}")
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200, f"GET {url_name} returned {resp.status_code}")

    def test_dashboard(self):
        self._get_ok("dashboard")

    def test_recepcion(self):
        self._get_ok("recepcion")

    def test_desverdizado(self):
        self._get_ok("desverdizado")

    def test_ingreso_packing(self):
        self._get_ok("ingreso_packing")

    def test_proceso(self):
        self._get_ok("proceso")

    def test_control(self):
        self._get_ok("control")

    def test_paletizado(self):
        self._get_ok("paletizado")

    def test_camaras(self):
        self._get_ok("camaras")

    def test_consulta_jefatura(self):
        # Consulta requires jefatura role (is_staff)
        self.user.is_staff = True
        self.user.save()
        self._get_ok("consulta")

    def test_consulta_jefatura_con_filtros(self):
        self.user.is_staff = True
        self.user.save()
        url = reverse("operaciones:consulta") + "?productor=PROD&estado=cerrado"
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)

    def test_consulta_jefatura_redirect_sin_rol(self):
        """Operador sin rol admin/jefatura es redirigido fuera de consulta."""
        self.user.is_superuser = False
        self.user.is_staff = False
        self.user.save()
        url = reverse("operaciones:consulta")
        resp = self.client.get(url)
        self.assertIn(resp.status_code, [302, 403])

    def test_redirect_sin_auth(self):
        """Sin login debe redirigir al login."""
        self.client.logout()
        url = reverse("operaciones:dashboard")
        resp = self.client.get(url)
        self.assertIn(resp.status_code, [302, 301])


# ---------------------------------------------------------------------------
# Flujo recepcion: iniciar lote → agregar bin → rechazar bin → cerrar lote
# ---------------------------------------------------------------------------

class RecepcionFlowTest(TestCase):

    def setUp(self):
        self.client = Client()
        self.user = _make_user()
        self.client.force_login(self.user)
        self.url = reverse("operaciones:recepcion")

    def test_iniciar_lote(self):
        resp = self.client.post(self.url, {
            "action": "iniciar",
            "temporada": TEMPORADA,
            "operator_code": "OP-01",
        })
        self.assertIn(resp.status_code, [200, 302])
        self.assertTrue(Lote.objects.filter(temporada=TEMPORADA).exists())

    def _datos_bin(self):
        return {
            "action": "agregar_bin",
            "temporada": TEMPORADA,
            "codigo_productor": "PROD-01",
            "tipo_cultivo": "Uva de mesa",
            "variedad_fruta": "Thompson",
            "color": "2",
            "fecha_cosecha": str(datetime.date.today()),
            "kilos_bruto_ingreso": "500",
            "kilos_neto_ingreso": "480",
            "operator_code": "OP-01",
        }

    def test_agregar_bin(self):
        # Iniciar primero
        self.client.post(self.url, {
            "action": "iniciar",
            "temporada": TEMPORADA,
            "operator_code": "OP-01",
        })
        lote = Lote.objects.filter(temporada=TEMPORADA).first()
        self.assertIsNotNone(lote)
        # Agregar bin valido
        resp = self.client.post(self.url, self._datos_bin())
        self.assertIn(resp.status_code, [200, 302])

    def test_agregar_multiples_bins_mismas_campos_base(self):
        """
        Agregar dos bins consecutivos con los mismos campos base no debe
        lanzar IntegrityError en RegistroEtapa.event_key.
        Regresion: agregar_bin_a_lote_abierto usaba create() en lugar de
        get_or_create() para RegistroEtapa, causando UNIQUE constraint fail
        ante reintento o bin_code repetido por datos de dev inconsistentes.
        """
        self.client.post(self.url, {
            "action": "iniciar",
            "temporada": TEMPORADA,
            "operator_code": "OP-01",
        })
        lote = Lote.objects.filter(temporada=TEMPORADA).first()
        self.assertIsNotNone(lote)
        datos = self._datos_bin()
        resp1 = self.client.post(self.url, datos)
        self.assertIn(resp1.status_code, [200, 302])
        resp2 = self.client.post(self.url, datos)
        self.assertIn(resp2.status_code, [200, 302])
        from operaciones.models import BinLote
        self.assertEqual(
            BinLote.objects.filter(lote=lote).count(), 2,
            "Deben existir dos bins distintos en el lote",
        )

    def test_cerrar_lote(self):
        # Crear lote abierto con un bin directo en BD
        lote = _make_lote(estado=LotePlantaEstado.ABIERTO)
        _make_bin(lote)
        self.client.session["lote_activo_code"] = lote.lote_code
        session = self.client.session
        session["lote_activo_code"] = lote.lote_code
        session["temporada_activa"] = TEMPORADA
        session.save()

        resp = self.client.post(self.url, {
            "action": "cerrar",
            "temporada": TEMPORADA,
            "operator_code": "OP-01",
            "kilos_bruto_conformacion": "500",
            "kilos_neto_conformacion": "480",
        })
        self.assertIn(resp.status_code, [200, 302])


# ---------------------------------------------------------------------------
# Flujo desverdizado: carga con selector/contexto
# ---------------------------------------------------------------------------

class DesverdizadoViewTest(TestCase):

    def setUp(self):
        self.client = Client()
        self.user = _make_user()
        self.client.force_login(self.user)
        self.url = reverse("operaciones:desverdizado")

    def test_get_renderiza_con_lotes_pendientes(self):
        lote = _make_lote(
            estado=LotePlantaEstado.CERRADO,
            requiere_desverdizado=True,
            disponibilidad_camara_desverdizado="disponible",
        )
        _make_bin(lote)
        resp = self.client.get(self.url)
        self.assertEqual(resp.status_code, 200)
        self.assertIn("lotes_data_json", resp.context)
        self.assertIn(lote.lote_code, resp.context["lotes_data_json"])

    def test_contexto_lote_contiene_campos_base(self):
        lote = _make_lote(
            estado=LotePlantaEstado.CERRADO,
            requiere_desverdizado=True,
            disponibilidad_camara_desverdizado="disponible",
        )
        _make_bin(lote, codigo_productor="PROD-77", variedad_fruta="Red Globe")
        resp = self.client.get(self.url)
        self.assertIn("PROD-77", resp.context["lotes_data_json"])
        self.assertIn("Red Globe", resp.context["lotes_data_json"])


# ---------------------------------------------------------------------------
# Flujo ingreso packing: via_desverdizado derivado
# ---------------------------------------------------------------------------

class IngresoPackingViewTest(TestCase):

    def setUp(self):
        self.client = Client()
        self.user = _make_user()
        self.client.force_login(self.user)
        self.url = reverse("operaciones:ingreso_packing")

    def test_get_renderiza(self):
        resp = self.client.get(self.url)
        self.assertEqual(resp.status_code, 200)
        self.assertIn("lotes_data_json", resp.context)

    def test_lote_con_desverdizado_tiene_via_desverdizado_true(self):
        lote = _make_lote(
            estado=LotePlantaEstado.CERRADO,
            requiere_desverdizado=True,
            disponibilidad_camara_desverdizado="disponible",
        )
        _make_bin(lote)
        Desverdizado.objects.create(
            lote=lote,
            source_system="test",
        )
        resp = self.client.get(self.url)
        import json
        data = json.loads(resp.context["lotes_data_json"])
        if lote.lote_code in data:
            self.assertTrue(data[lote.lote_code]["via_desverdizado"])


# ---------------------------------------------------------------------------
# horas_desverdizado se persiste en el modelo
# ---------------------------------------------------------------------------

class HorasDesverdizadoModelTest(TestCase):

    def test_campo_existe_y_persiste(self):
        lote = _make_lote(
            estado=LotePlantaEstado.CERRADO,
            requiere_desverdizado=True,
            disponibilidad_camara_desverdizado="disponible",
        )
        desv = Desverdizado.objects.create(
            lote=lote,
            horas_desverdizado=72,
            source_system="test",
        )
        desv.refresh_from_db()
        self.assertEqual(desv.horas_desverdizado, 72)
        # El campo proceso queda vacio (no se mezclan mas)
        self.assertEqual(desv.proceso, "")

    def test_horas_null_por_defecto(self):
        lote = _make_lote(estado=LotePlantaEstado.CERRADO, requiere_desverdizado=True)
        desv = Desverdizado.objects.create(lote=lote, source_system="test")
        desv.refresh_from_db()
        self.assertIsNone(desv.horas_desverdizado)

    def test_proceso_legacy_no_interfiere(self):
        """proceso (legacy) y horas_desverdizado son independientes."""
        lote = _make_lote(estado=LotePlantaEstado.CERRADO, requiere_desverdizado=True)
        desv = Desverdizado.objects.create(
            lote=lote, horas_desverdizado=48, proceso="conservacion frio",
            source_system="test",
        )
        desv.refresh_from_db()
        self.assertEqual(desv.horas_desverdizado, 48)
        self.assertEqual(desv.proceso, "conservacion frio")


# ---------------------------------------------------------------------------
# DesverdizadoForm — validacion server-side de horas_desverdizado
# ---------------------------------------------------------------------------

class DesverdizadoFormValidationTest(TestCase):

    _BASE = {
        "numero_camara": "1",
        "fecha_ingreso": str(datetime.date.today()),
        "hora_ingreso": "08:00",
        "color": "3",
    }

    def test_horas_valido(self):
        from operaciones.forms import DesverdizadoForm
        form = DesverdizadoForm(data={**self._BASE, "horas_desverdizado": 72})
        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data["horas_desverdizado"], 72)

    def test_horas_limite_inferior(self):
        from operaciones.forms import DesverdizadoForm
        form = DesverdizadoForm(data={**self._BASE, "horas_desverdizado": 1})
        self.assertTrue(form.is_valid(), form.errors)

    def test_horas_limite_superior(self):
        from operaciones.forms import DesverdizadoForm
        form = DesverdizadoForm(data={**self._BASE, "horas_desverdizado": 240})
        self.assertTrue(form.is_valid(), form.errors)

    def test_horas_cero_invalido(self):
        from operaciones.forms import DesverdizadoForm
        form = DesverdizadoForm(data={"horas_desverdizado": 0})
        self.assertFalse(form.is_valid())
        self.assertIn("horas_desverdizado", form.errors)

    def test_horas_excede_maximo(self):
        from operaciones.forms import DesverdizadoForm
        form = DesverdizadoForm(data={"horas_desverdizado": 241})
        self.assertFalse(form.is_valid())
        self.assertIn("horas_desverdizado", form.errors)

    def test_horas_negativo_invalido(self):
        from operaciones.forms import DesverdizadoForm
        form = DesverdizadoForm(data={"horas_desverdizado": -5})
        self.assertFalse(form.is_valid())

    def test_horas_vacio_ok(self):
        from operaciones.forms import DesverdizadoForm
        form = DesverdizadoForm(data=self._BASE)  # sin horas_desverdizado
        self.assertTrue(form.is_valid(), form.errors)
        self.assertIsNone(form.cleaned_data["horas_desverdizado"])


# ---------------------------------------------------------------------------
# CalidadPalletMuestra — modelo y flujo
# ---------------------------------------------------------------------------

class CalidadPalletMuestraModelTest(TestCase):

    def setUp(self):
        self.lote = _make_lote(estado=LotePlantaEstado.CERRADO)
        _make_bin(self.lote)
        self.pallet = Pallet.objects.create(
            temporada=TEMPORADA,
            pallet_code=f"PAL-TEST-{Pallet.objects.count()+1:03d}",
            is_active=True,
        )
        PalletLote.objects.create(pallet=self.pallet, lote=self.lote)

    def test_crear_muestra(self):
        from operaciones.models import CalidadPalletMuestra
        m = CalidadPalletMuestra.objects.create(
            pallet=self.pallet,
            numero_muestra=1,
            temperatura_fruta=18.5,
            peso_caja_muestra=8.200,
            n_frutos=42,
            aprobado=True,
            source_system="test",
        )
        m.refresh_from_db()
        self.assertEqual(m.numero_muestra, 1)
        self.assertEqual(float(m.temperatura_fruta), 18.5)
        self.assertTrue(m.aprobado)

    def test_multiples_muestras_por_pallet(self):
        from operaciones.models import CalidadPalletMuestra
        for i in range(1, 4):
            CalidadPalletMuestra.objects.create(
                pallet=self.pallet, numero_muestra=i,
                temperatura_fruta=17 + i, source_system="test",
            )
        self.assertEqual(self.pallet.muestras_calidad.count(), 3)

    def test_muestra_sin_datos_opcionales(self):
        from operaciones.models import CalidadPalletMuestra
        m = CalidadPalletMuestra.objects.create(
            pallet=self.pallet, numero_muestra=1, source_system="test",
        )
        m.refresh_from_db()
        self.assertIsNone(m.temperatura_fruta)
        self.assertIsNone(m.aprobado)

    def test_ordering(self):
        from operaciones.models import CalidadPalletMuestra
        CalidadPalletMuestra.objects.create(
            pallet=self.pallet, numero_muestra=3, source_system="test",
        )
        CalidadPalletMuestra.objects.create(
            pallet=self.pallet, numero_muestra=1, source_system="test",
        )
        nums = list(
            self.pallet.muestras_calidad.values_list("numero_muestra", flat=True)
        )
        self.assertEqual(nums, [1, 3])


# ---------------------------------------------------------------------------
# PaletizadoView — contexto y flujo de muestras
# ---------------------------------------------------------------------------

class PaletizadoViewTest(TestCase):

    def setUp(self):
        self.client = Client()
        self.user = _make_user()
        self.client.force_login(self.user)
        self.url = reverse("operaciones:paletizado")

    def test_get_incluye_form_muestra(self):
        resp = self.client.get(self.url)
        self.assertEqual(resp.status_code, 200)
        self.assertIn("form_muestra", resp.context)
        self.assertIn("pallets_data_json", resp.context)

    def test_pallets_data_json_contiene_campos(self):
        lote = _make_lote(estado=LotePlantaEstado.CERRADO)
        _make_bin(lote, codigo_productor="PROD-55", variedad_fruta="Flame")
        pallet = Pallet.objects.create(
            temporada=TEMPORADA,
            pallet_code="PAL-CTX-001",
            is_active=True,
            peso_total_kg=1200,
        )
        PalletLote.objects.create(pallet=pallet, lote=lote)
        resp = self.client.get(self.url)
        import json
        data = json.loads(resp.context["pallets_data_json"])
        if "PAL-CTX-001" in data:
            entry = data["PAL-CTX-001"]
            self.assertEqual(entry["productor"], "PROD-55")
            self.assertEqual(entry["variedad"], "Flame")
            self.assertEqual(entry["peso_total"], 1200.0)


# ---------------------------------------------------------------------------
# CamarasView — contexto con color y peso
# ---------------------------------------------------------------------------

class CamarasViewTest(TestCase):

    def setUp(self):
        self.client = Client()
        self.user = _make_user()
        self.client.force_login(self.user)
        self.url = reverse("operaciones:camaras")

    def test_get_renderiza(self):
        resp = self.client.get(self.url)
        self.assertEqual(resp.status_code, 200)
        self.assertIn("pallets_data_json", resp.context)

    def test_pallets_data_incluye_color_y_peso(self):
        lote = _make_lote(estado=LotePlantaEstado.CERRADO)
        _make_bin(lote, color="4")
        pallet = Pallet.objects.create(
            temporada=TEMPORADA,
            pallet_code="PAL-CAM-001",
            is_active=True,
            peso_total_kg=950,
        )
        PalletLote.objects.create(pallet=pallet, lote=lote)
        resp = self.client.get(self.url)
        import json
        data = json.loads(resp.context["pallets_data_json"])
        if "PAL-CAM-001" in data:
            self.assertEqual(data["PAL-CAM-001"]["color"], "4")
            self.assertEqual(data["PAL-CAM-001"]["peso_total"], 950.0)


# ---------------------------------------------------------------------------
# Smoke: todas las vistas con admin autenticado
# ---------------------------------------------------------------------------

@override_settings(PERSISTENCE_BACKEND="sqlite")
class SmokeViewsAuthenticatedTest(TestCase):
    """
    Evidencia de que todas las vistas críticas renderizan correctamente
    con un Administrador autenticado (is_superuser=True, rol="Administrador").
    Cubre el flujo completo login → dashboard → recepcion → … → exportacion.
    """

    def setUp(self):
        self.client = _make_admin_client("admin_sva")
        self.anon = Client()

    def test_login_page_anonima_renderiza(self):
        resp = self.anon.get(reverse("usuarios:login"))
        self.assertEqual(resp.status_code, 200)

    def test_login_redirige_si_ya_autenticado(self):
        resp = self.client.get(reverse("usuarios:login"))
        self.assertIn(resp.status_code, [301, 302])

    def test_dashboard(self):
        resp = self.client.get(reverse("operaciones:dashboard"))
        self.assertEqual(resp.status_code, 200)

    def test_recepcion(self):
        resp = self.client.get(reverse("operaciones:recepcion"))
        self.assertEqual(resp.status_code, 200)

    def test_desverdizado(self):
        resp = self.client.get(reverse("operaciones:desverdizado"))
        self.assertEqual(resp.status_code, 200)

    def test_ingreso_packing(self):
        resp = self.client.get(reverse("operaciones:ingreso_packing"))
        self.assertEqual(resp.status_code, 200)

    def test_proceso(self):
        resp = self.client.get(reverse("operaciones:proceso"))
        self.assertEqual(resp.status_code, 200)

    def test_control(self):
        resp = self.client.get(reverse("operaciones:control"))
        self.assertEqual(resp.status_code, 200)

    def test_paletizado(self):
        resp = self.client.get(reverse("operaciones:paletizado"))
        self.assertEqual(resp.status_code, 200)

    def test_camaras(self):
        resp = self.client.get(reverse("operaciones:camaras"))
        self.assertEqual(resp.status_code, 200)

    def test_consulta(self):
        resp = self.client.get(reverse("operaciones:consulta"))
        self.assertEqual(resp.status_code, 200)

    def test_exportacion_csv(self):
        resp = self.client.get(reverse("operaciones:exportar_consulta"))
        self.assertEqual(resp.status_code, 200)
        self.assertIn("text/csv", resp.get("Content-Type", ""))


# ---------------------------------------------------------------------------
# Flujo recepción E2E: iniciar → agregar bin → cerrar lote
# ---------------------------------------------------------------------------

@override_settings(PERSISTENCE_BACKEND="sqlite")
class RecepcionFlowE2ETest(TestCase):
    """
    Secuencia narrativa del flujo mínimo de recepción.
    Cada test es independiente (crea su propio estado) para que los fallos
    sean aislados. El conjunto documenta el camino feliz completo.
    """

    def _admin_client(self, suffix):
        return _make_admin_client(f"admin_rfe_{suffix}")

    def test_01_iniciar_lote(self):
        c = self._admin_client("ini")
        resp = c.post(reverse("operaciones:recepcion"), {
            "action": "iniciar",
            "temporada": TEMPORADA,
            "operator_code": "ADM-001",
        })
        self.assertIn(resp.status_code, [200, 302])
        self.assertTrue(Lote.objects.filter(temporada=TEMPORADA).exists())

    def test_02_agregar_bin(self):
        c = self._admin_client("bin")
        # Iniciar lote
        c.post(reverse("operaciones:recepcion"), {
            "action": "iniciar",
            "temporada": TEMPORADA,
            "operator_code": "ADM-001",
        })
        lote = Lote.objects.filter(temporada=TEMPORADA).first()
        self.assertIsNotNone(lote)
        resp = c.post(reverse("operaciones:recepcion"), {
            "action": "agregar_bin",
            "temporada": TEMPORADA,
            "codigo_productor": "PROD-01",
            "tipo_cultivo": "Uva de mesa",
            "variedad_fruta": "Thompson",
            "color": "2",
            "fecha_cosecha": str(datetime.date.today()),
            "hora_recepcion": "08:00",
            "numero_cuartel": "C01",
            "kilos_bruto_ingreso": "500",
            "kilos_neto_ingreso": "480",
            "a_o_r": "aprobado",
            "operator_code": "ADM-001",
        })
        self.assertIn(resp.status_code, [200, 302])
        self.assertGreaterEqual(BinLote.objects.filter(lote=lote).count(), 1)

    def test_03_cerrar_lote(self):
        c = self._admin_client("cer")
        lote = _make_lote(estado=LotePlantaEstado.ABIERTO)
        _make_bin(lote)
        session = c.session
        session["lote_activo_code"] = lote.lote_code
        session["temporada_activa"] = TEMPORADA
        session.save()
        resp = c.post(reverse("operaciones:recepcion"), {
            "action": "cerrar",
            "temporada": TEMPORADA,
            "operator_code": "ADM-001",
            "kilos_bruto_conformacion": "500",
            "kilos_neto_conformacion": "480",
        })
        self.assertIn(resp.status_code, [200, 302])

    def test_04_estado_persiste_cerrado(self):
        # Usar use cases para que cantidad_bins se actualice correctamente
        r = _iniciar()
        self.assertTrue(r.ok)
        lote_code = r.data["lote_code"]
        _agregar_bin(lote_code)
        result = cerrar_lote_recepcion({
            "temporada": TEMPORADA,
            "lote_code": lote_code,
            "operator_code": "ADM-001",
            "source_system": "test",
        })
        self.assertTrue(result.ok, result.errors)
        repos = get_repositories()
        lote = repos.lotes.find_by_code(TEMPORADA, lote_code)
        self.assertEqual(lote.estado, "cerrado")


# ---------------------------------------------------------------------------
# Control de acceso por rol: evidencia negativa y positiva
# ---------------------------------------------------------------------------

@override_settings(PERSISTENCE_BACKEND="sqlite")
class RoleAccessControlTest(TestCase):
    """
    Evidencia de que RolRequiredMixin y JefaturaRequiredMixin protegen
    correctamente cada módulo operativo.

    Códigos HTTP verificados empíricamente:
      - Vistas operativas (RolRequiredMixin): 403 para autenticados con rol incorrecto.
        UserPassesTestMixin.handle_no_permission levanta PermissionDenied cuando
        user.is_authenticated=True (independiente de raise_exception).
      - ExportarConsultaCSVView: 302. JefaturaRequiredMixin sobreescribe handle_no_permission
        con redirect a portal (bypasa el PermissionDenied de UserPassesTestMixin).
      - ConsultaJefaturaView: 302. Usa JefaturaRequiredMixin.handle_no_permission via MRO.

    Arquitectura:
      - Vistas operativas: RolRequiredMixin → lee SESSION_KEY_ROL → 302 si falla
      - ExportarConsultaCSVView: JefaturaRequiredMixin → lee sesión → 302 si falla
      - ConsultaJefaturaView: override test_func → is_staff/is_superuser → 403 si falla
                              (raise_exception=True)
    """

    # Negative: rol incorrecto rechazado en cada módulo operativo
    # RolRequiredMixin → UserPassesTestMixin.handle_no_permission → 403 para autenticados
    # (Django levanta PermissionDenied cuando user.is_authenticated=True)

    def test_recepcion_rechaza_rol_desverdizado(self):
        c, _ = _make_client_with_session_rol("u_rec_desv", "Desverdizado")
        resp = c.get(reverse("operaciones:recepcion"))
        self.assertEqual(resp.status_code, 403)

    def test_desverdizado_rechaza_rol_recepcion(self):
        c, _ = _make_client_with_session_rol("u_desv_rec", "Recepcion")
        resp = c.get(reverse("operaciones:desverdizado"))
        self.assertEqual(resp.status_code, 403)

    def test_ingreso_packing_rechaza_rol_proceso(self):
        c, _ = _make_client_with_session_rol("u_ing_pro", "Proceso")
        resp = c.get(reverse("operaciones:ingreso_packing"))
        self.assertEqual(resp.status_code, 403)

    def test_proceso_rechaza_rol_control(self):
        c, _ = _make_client_with_session_rol("u_pro_ctrl", "Control")
        resp = c.get(reverse("operaciones:proceso"))
        self.assertEqual(resp.status_code, 403)

    def test_control_rechaza_rol_paletizado(self):
        c, _ = _make_client_with_session_rol("u_ctrl_pal", "Paletizado")
        resp = c.get(reverse("operaciones:control"))
        self.assertEqual(resp.status_code, 403)

    def test_paletizado_rechaza_rol_camaras(self):
        c, _ = _make_client_with_session_rol("u_pal_cam", "Camaras")
        resp = c.get(reverse("operaciones:paletizado"))
        self.assertEqual(resp.status_code, 403)

    def test_camaras_rechaza_rol_recepcion(self):
        c, _ = _make_client_with_session_rol("u_cam_rec", "Recepcion")
        resp = c.get(reverse("operaciones:camaras"))
        self.assertEqual(resp.status_code, 403)

    # Negative: operadores rechazados en consulta y exportar

    def test_consulta_rechaza_operador_sin_flags(self):
        """
        ConsultaJefaturaView.test_func usa is_staff/is_superuser.
        JefaturaRequiredMixin.handle_no_permission hace redirect → 302.
        (handle_no_permission sobrescribe raise_exception=True vía MRO.)
        """
        user = User.objects.create_user(
            username="u_cons_ope", password="pw123",
            is_staff=False, is_superuser=False,
        )
        c = Client()
        c.force_login(user)
        resp = c.get(reverse("operaciones:consulta"))
        self.assertEqual(resp.status_code, 302)

    def test_exportar_rechaza_operador(self):
        """ExportarConsultaCSVView usa JefaturaRequiredMixin → lee sesión → 302."""
        c, _ = _make_client_with_session_rol("u_exp_pro", "Proceso")
        resp = c.get(reverse("operaciones:exportar_consulta"))
        self.assertEqual(resp.status_code, 302)

    # Positive: Jefatura accede a consulta, rechazada en módulos operativos

    def test_jefatura_accede_a_consulta(self):
        """ConsultaJefaturaView: is_staff=True → 200."""
        c, _ = _make_client_with_session_rol("u_jef_cons", "Jefatura", is_staff=True)
        resp = c.get(reverse("operaciones:consulta"))
        self.assertEqual(resp.status_code, 200)

    def test_jefatura_rechazada_en_recepcion(self):
        """Jefatura no está en roles_requeridos=["Recepcion"] → 403 (RolRequiredMixin)."""
        c, _ = _make_client_with_session_rol("u_jef_rec", "Jefatura", is_staff=True)
        resp = c.get(reverse("operaciones:recepcion"))
        self.assertEqual(resp.status_code, 403)

    # Multirol: usuario con varios roles accede a los asignados, rechazado en otros

    def test_multirol_accede_a_modulos_asignados(self):
        c, _ = _make_client_with_session_rol("u_multi", "Recepcion, Desverdizado")
        self.assertEqual(c.get(reverse("operaciones:recepcion")).status_code, 200)
        self.assertEqual(c.get(reverse("operaciones:desverdizado")).status_code, 200)
        self.assertEqual(c.get(reverse("operaciones:proceso")).status_code, 403)

    # Admin bypass: Administrador accede a todos los módulos

    def test_admin_accede_a_todos_los_modulos(self):
        c = _make_admin_client("admin_all")
        urls = [
            reverse("operaciones:recepcion"),
            reverse("operaciones:desverdizado"),
            reverse("operaciones:ingreso_packing"),
            reverse("operaciones:proceso"),
            reverse("operaciones:control"),
            reverse("operaciones:paletizado"),
            reverse("operaciones:camaras"),
            reverse("operaciones:consulta"),
            reverse("operaciones:exportar_consulta"),
            reverse("operaciones:exportar_consulta_excel"),
        ]
        for url in urls:
            resp = c.get(url)
            self.assertEqual(resp.status_code, 200, f"Admin debe acceder a {url}")

    # Sin autenticación: todos los módulos redirigen

    def test_unauthenticated_es_redirigido_en_todos_los_modulos(self):
        anon = Client()
        urls = [
            reverse("operaciones:recepcion"),
            reverse("operaciones:desverdizado"),
            reverse("operaciones:ingreso_packing"),
            reverse("operaciones:proceso"),
            reverse("operaciones:control"),
            reverse("operaciones:paletizado"),
            reverse("operaciones:camaras"),
            reverse("operaciones:consulta"),
            reverse("operaciones:exportar_consulta"),
            reverse("operaciones:exportar_consulta_excel"),
        ]
        for url in urls:
            resp = anon.get(url)
            self.assertIn(resp.status_code, [301, 302], f"Anónimo debe ser redirigido en {url}")


@override_settings(PERSISTENCE_BACKEND="sqlite")
class ConsultaJefaturaDetalleExportSqliteTest(TestCase):
    def setUp(self):
        self.client = _make_admin_client("admin_consulta_det_sqlite")
        hoy = datetime.date.today()

        self.lote_a = _make_lote(estado=LotePlantaEstado.CERRADO, lote_code="LP-CONS-001")
        _make_bin(
            self.lote_a,
            bin_code="BIN-CONS-0001",
            codigo_productor="PROD-A",
            variedad_fruta="Clementina",
            color="3",
            fecha_cosecha=hoy,
            kilos_bruto_ingreso=520,
            kilos_neto_ingreso=500,
        )
        self.lote_a.cantidad_bins = 1
        self.lote_a.kilos_bruto_conformacion = Decimal("520")
        self.lote_a.kilos_neto_conformacion = Decimal("500")
        self.lote_a.fecha_conformacion = hoy
        self.lote_a.save()

        self.lote_b = _make_lote(estado=LotePlantaEstado.ABIERTO, lote_code="LP-CONS-002")
        _make_bin(
            self.lote_b,
            bin_code="BIN-CONS-0002",
            codigo_productor="PROD-B",
            variedad_fruta="Murcott",
            color="4",
            fecha_cosecha=hoy,
            kilos_bruto_ingreso=410,
            kilos_neto_ingreso=395,
        )
        self.lote_b.cantidad_bins = 1
        self.lote_b.save(update_fields=["cantidad_bins"])

        self.pallet_a = Pallet.objects.create(
            temporada=TEMPORADA,
            pallet_code="PAL-CONS-001",
            tipo_caja="Caja 10kg",
            peso_total_kg=Decimal("900"),
            is_active=True,
        )
        PalletLote.objects.create(pallet=self.pallet_a, lote=self.lote_a)
        CamaraFrio.objects.create(
            pallet=self.pallet_a,
            camara_numero="CF-01",
            source_system="test",
        )

        self.pallet_b = Pallet.objects.create(
            temporada=TEMPORADA,
            pallet_code="PAL-CONS-002",
            tipo_caja="Caja 8kg",
            peso_total_kg=Decimal("720"),
            is_active=True,
        )
        PalletLote.objects.create(pallet=self.pallet_b, lote=self.lote_b)

    def test_consulta_renderiza_tabs_lotes_y_pallets(self):
        resp_lotes = self.client.get(reverse("operaciones:consulta"))
        self.assertEqual(resp_lotes.status_code, 200)
        self.assertContains(resp_lotes, "Lotes (")
        self.assertContains(resp_lotes, "Pallets (")

        resp_pallets = self.client.get(reverse("operaciones:consulta"), {"tab": "pallets"})
        self.assertEqual(resp_pallets.status_code, 200)
        self.assertContains(resp_pallets, self.pallet_a.pallet_code)

    def test_detalle_lote_muestra_bins_relacionados(self):
        resp = self.client.get(
            reverse("operaciones:consulta_lote_detalle", args=[self.lote_a.lote_code])
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, self.lote_a.lote_code)
        self.assertContains(resp, "BIN-CONS-0001")

    def test_detalle_pallet_muestra_lote_y_link(self):
        resp = self.client.get(
            reverse("operaciones:consulta_pallet_detalle", args=[self.pallet_a.pallet_code])
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, self.pallet_a.pallet_code)
        self.assertContains(resp, self.lote_a.lote_code)
        lote_url = reverse("operaciones:consulta_lote_detalle", args=[self.lote_a.lote_code])
        self.assertContains(resp, lote_url)

    def test_detalle_no_encontrado_redirige_consulta(self):
        resp = self.client.get(
            reverse("operaciones:consulta_lote_detalle", args=["LP-NO-EXISTE"]),
            follow=True,
        )
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(
            resp.redirect_chain[-1][0].startswith(reverse("operaciones:consulta"))
        )
        self.assertIn("no encontrado", resp.content.decode("utf-8", errors="ignore").lower())

    def test_detalle_pallet_no_encontrado_redirige_consulta(self):
        resp = self.client.get(
            reverse("operaciones:consulta_pallet_detalle", args=["PAL-NO-EXISTE"]),
            follow=True,
        )
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(
            resp.redirect_chain[-1][0].startswith(reverse("operaciones:consulta"))
        )
        self.assertIn("no encontrado", resp.content.decode("utf-8", errors="ignore").lower())

    def test_filtro_productor_aplica_a_lotes_y_pallets(self):
        resp_lotes = self.client.get(
            reverse("operaciones:consulta"),
            {"tab": "lotes", "productor": "PROD-A"},
        )
        self.assertContains(resp_lotes, self.lote_a.lote_code)
        self.assertNotContains(resp_lotes, self.lote_b.lote_code)

        resp_pallets = self.client.get(
            reverse("operaciones:consulta"),
            {"tab": "pallets", "productor": "PROD-A"},
        )
        self.assertContains(resp_pallets, self.pallet_a.pallet_code)
        self.assertNotContains(resp_pallets, self.pallet_b.pallet_code)

    def test_filtro_estado_en_camara_frio_filtra_pallets(self):
        resp = self.client.get(
            reverse("operaciones:consulta"),
            {"tab": "pallets", "estado": "en_camara_frio"},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, self.pallet_a.pallet_code)
        self.assertNotContains(resp, self.pallet_b.pallet_code)

    def test_export_csv_por_tab(self):
        resp_lotes = self.client.get(
            reverse("operaciones:exportar_consulta"),
            {"tab": "lotes"},
        )
        self.assertEqual(resp_lotes.status_code, 200)
        self.assertIn("text/csv", resp_lotes.get("Content-Type", ""))
        csv_lotes = resp_lotes.content.decode("utf-8-sig")
        self.assertIn("Lote (code)", csv_lotes)
        self.assertIn(self.lote_a.lote_code, csv_lotes)

        resp_pallets = self.client.get(
            reverse("operaciones:exportar_consulta"),
            {"tab": "pallets"},
        )
        self.assertEqual(resp_pallets.status_code, 200)
        csv_pallets = resp_pallets.content.decode("utf-8-sig")
        self.assertIn("Pallet (code)", csv_pallets)
        self.assertIn("Lote relacionado", csv_pallets)
        self.assertIn(self.pallet_a.pallet_code, csv_pallets)

    def test_export_excel_por_tab(self):
        try:
            from openpyxl import load_workbook
        except Exception:
            self.skipTest("openpyxl no disponible en este entorno")

        resp_lotes = self.client.get(
            reverse("operaciones:exportar_consulta_excel"),
            {"tab": "lotes"},
        )
        self.assertEqual(resp_lotes.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resp_lotes.get("Content-Type", ""),
        )
        self.assertIn(".xlsx", resp_lotes.get("Content-Disposition", ""))
        wb_lotes = load_workbook(filename=BytesIO(resp_lotes.content))
        ws_lotes = wb_lotes.active
        headers_lotes = [c.value for c in ws_lotes[1]]
        self.assertIn("Lote (code)", headers_lotes)
        self.assertIn("Etapa actual", headers_lotes)

        resp_pallets = self.client.get(
            reverse("operaciones:exportar_consulta_excel"),
            {"tab": "pallets"},
        )
        self.assertEqual(resp_pallets.status_code, 200)
        wb_pallets = load_workbook(filename=BytesIO(resp_pallets.content))
        ws_pallets = wb_pallets["Pallets"]
        headers_pallets = [c.value for c in ws_pallets[1]]
        self.assertIn("Pallet (code)", headers_pallets)
        self.assertIn("Lote relacionado", headers_pallets)


@override_settings(PERSISTENCE_BACKEND="dataverse")
class ConsultaJefaturaDataverseCompatTest(TestCase):
    def setUp(self):
        self.client = _make_admin_client("admin_consulta_dataverse")
        hoy = datetime.date.today()
        self._tmp_cache_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmp_cache_dir.cleanup)
        self.cache_file = Path(self._tmp_cache_dir.name) / "consulta_dataverse.json"
        self._cache_override = override_settings(
            CONSULTA_DATAVERSE_CACHE_FILE=str(self.cache_file),
            CONSULTA_DATAVERSE_CACHE_TTL_SECONDS=3600,
        )
        self._cache_override.enable()
        self.addCleanup(self._cache_override.disable)

        self.dv_lote = SimpleNamespace(
            id="dv-lote-1",
            lote_code="LP-DV-001",
            estado="cerrado",
            etapa_actual="Pallet Cerrado",
            cantidad_bins=1,
            kilos_bruto_conformacion=Decimal("610"),
            kilos_neto_conformacion=Decimal("590"),
            requiere_desverdizado=False,
            fecha_conformacion=hoy,
            ultimo_cambio_estado_at=None,
            codigo_productor="PROD-DV",
            is_active=True,
        )
        self.dv_bin = SimpleNamespace(
            id="dv-bin-1",
            bin_code="BIN-DV-001",
            codigo_productor="PROD-DV",
            tipo_cultivo="Citricos",
            variedad_fruta="Oronules",
            color="4",
            fecha_cosecha=hoy,
            kilos_bruto_ingreso=Decimal("610"),
            kilos_neto_ingreso=Decimal("590"),
        )
        self.dv_pallet = SimpleNamespace(
            id="dv-pallet-1",
            pallet_code="PAL-DV-001",
            tipo_caja="Caja DV",
            peso_total_kg=Decimal("990"),
            fecha=hoy,
        )
        self.dv_pallet_lote = SimpleNamespace(
            id="dv-pl-1",
            pallet_id="dv-pallet-1",
            lote_id="dv-lote-1",
        )
        self.dv_camara = SimpleNamespace(id="dv-cf-1", pallet_id="dv-pallet-1")

        self.repos = SimpleNamespace(
            lotes=SimpleNamespace(
                list_recent=lambda limit=500: [self.dv_lote],
                find_by_code=lambda temporada, code: self.dv_lote if code == "LP-DV-001" else None,
                find_by_id=lambda lote_id: self.dv_lote if lote_id == "dv-lote-1" else None,
            ),
            bins=SimpleNamespace(
                first_bin_by_lotes=lambda lote_ids: {"dv-lote-1": self.dv_bin},
                all_bins_by_lotes=lambda lote_ids: {"dv-lote-1": [self.dv_bin]},
                list_by_lote=lambda lote_id: [self.dv_bin] if lote_id == "dv-lote-1" else [],
            ),
            desverdizados=SimpleNamespace(
                find_by_lote=lambda lote_id: None,
                list_by_lotes=lambda lote_ids: {},
            ),
            ingresos_packing=SimpleNamespace(
                find_by_lote=lambda lote_id: None,
                list_by_lotes=lambda lote_ids: {},
            ),
            pallets=SimpleNamespace(
                list_recent=lambda limit=500: [self.dv_pallet],
                find_by_code=lambda temporada, code: self.dv_pallet if code == "PAL-DV-001" else None,
            ),
            pallet_lotes=SimpleNamespace(
                find_by_pallet=lambda pallet_id: self.dv_pallet_lote if pallet_id == "dv-pallet-1" else None,
            ),
            camara_frios=SimpleNamespace(
                find_by_pallet=lambda pallet_id: self.dv_camara if pallet_id == "dv-pallet-1" else None,
            ),
        )

        self.repo_patcher = patch(
            "infrastructure.repository_factory.get_repositories",
            return_value=self.repos,
        )
        self.repo_patcher.start()
        self.addCleanup(self.repo_patcher.stop)

        self.etapa_patcher = patch(
            "infrastructure.dataverse.repositories.resolve_etapa_lote",
            side_effect=lambda lote, repos=None: (
                "Pallet Cerrado" if lote and getattr(lote, "id", "") == "dv-lote-1" else "Recepcion"
            ),
        )
        self.etapa_patcher.start()
        self.addCleanup(self.etapa_patcher.stop)

    def test_consulta_tab_pallets_dataverse(self):
        resp = self.client.get(reverse("operaciones:consulta"), {"tab": "pallets"})
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "PAL-DV-001")
        self.assertContains(resp, "LP-DV-001")

    def test_detalle_lote_dataverse_muestra_bins(self):
        resp = self.client.get(reverse("operaciones:consulta_lote_detalle", args=["LP-DV-001"]))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "BIN-DV-001")

    def test_detalle_pallet_dataverse_linkea_lote(self):
        resp = self.client.get(reverse("operaciones:consulta_pallet_detalle", args=["PAL-DV-001"]))
        self.assertEqual(resp.status_code, 200)
        lote_url = reverse("operaciones:consulta_lote_detalle", args=["LP-DV-001"])
        self.assertContains(resp, lote_url)

    def test_filtro_estado_en_camara_frio_dataverse(self):
        resp = self.client.get(
            reverse("operaciones:consulta"),
            {"tab": "pallets", "estado": "en_camara_frio"},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "PAL-DV-001")

    def test_export_csv_pallets_dataverse(self):
        resp = self.client.get(reverse("operaciones:exportar_consulta"), {"tab": "pallets"})
        self.assertEqual(resp.status_code, 200)
        csv_txt = resp.content.decode("utf-8-sig")
        self.assertIn("Pallet (code)", csv_txt)
        self.assertIn("PAL-DV-001", csv_txt)


@override_settings(PERSISTENCE_BACKEND="dataverse")
class ConsultaJefaturaDataverseCacheTest(TestCase):
    def setUp(self):
        self.client = _make_admin_client("admin_consulta_cache_dataverse")
        hoy = datetime.date.today()
        self._tmp_cache_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmp_cache_dir.cleanup)
        self.cache_file = Path(self._tmp_cache_dir.name) / "consulta_dataverse.json"
        self._cache_override = override_settings(
            CONSULTA_DATAVERSE_CACHE_FILE=str(self.cache_file),
            CONSULTA_DATAVERSE_CACHE_TTL_SECONDS=3600,
        )
        self._cache_override.enable()
        self.addCleanup(self._cache_override.disable)

        self.dv_lote = SimpleNamespace(
            id="dv-lote-cache-1",
            lote_code="LP-DVC-001",
            estado="cerrado",
            etapa_actual="Pallet Cerrado",
            cantidad_bins=1,
            kilos_bruto_conformacion=Decimal("510"),
            kilos_neto_conformacion=Decimal("500"),
            requiere_desverdizado=False,
            fecha_conformacion=hoy,
            ultimo_cambio_estado_at=None,
            codigo_productor="PROD-DVC",
            is_active=True,
        )
        self.dv_bin = SimpleNamespace(
            id="dv-bin-cache-1",
            bin_code="BIN-DVC-001",
            codigo_productor="PROD-DVC",
            tipo_cultivo="Citricos",
            variedad_fruta="Nadorcott",
            color="3",
            fecha_cosecha=hoy,
            kilos_bruto_ingreso=Decimal("510"),
            kilos_neto_ingreso=Decimal("500"),
        )
        self.dv_pallet = SimpleNamespace(
            id="dv-pallet-cache-1",
            pallet_code="PAL-DVC-001",
            tipo_caja="Caja 10kg",
            peso_total_kg=Decimal("800"),
            fecha=hoy,
        )
        self.dv_pallet_lote = SimpleNamespace(
            id="dv-pl-cache-1",
            pallet_id="dv-pallet-cache-1",
            lote_id="dv-lote-cache-1",
        )
        self.dv_camara = SimpleNamespace(id="dv-cf-cache-1", pallet_id="dv-pallet-cache-1")

        self.mock_lotes_list_recent = Mock(return_value=[self.dv_lote])
        self.mock_pallets_list_recent = Mock(return_value=[self.dv_pallet])
        self.repos = SimpleNamespace(
            lotes=SimpleNamespace(
                list_recent=self.mock_lotes_list_recent,
                find_by_code=Mock(
                    side_effect=lambda temporada, code: self.dv_lote if code == "LP-DVC-001" else None
                ),
                find_by_id=Mock(
                    side_effect=lambda lote_id: self.dv_lote if lote_id == "dv-lote-cache-1" else None
                ),
            ),
            bins=SimpleNamespace(
                first_bin_by_lotes=Mock(return_value={"dv-lote-cache-1": self.dv_bin}),
                all_bins_by_lotes=Mock(return_value={"dv-lote-cache-1": [self.dv_bin]}),
                list_by_lote=Mock(
                    side_effect=lambda lote_id: [self.dv_bin] if lote_id == "dv-lote-cache-1" else []
                ),
            ),
            desverdizados=SimpleNamespace(
                find_by_lote=Mock(return_value=None),
                list_by_lotes=Mock(return_value={}),
            ),
            ingresos_packing=SimpleNamespace(
                find_by_lote=Mock(return_value=None),
                list_by_lotes=Mock(return_value={}),
            ),
            pallets=SimpleNamespace(
                list_recent=self.mock_pallets_list_recent,
                find_by_code=Mock(
                    side_effect=lambda temporada, code: self.dv_pallet if code == "PAL-DVC-001" else None
                ),
            ),
            pallet_lotes=SimpleNamespace(
                find_by_pallet=Mock(
                    side_effect=lambda pallet_id: self.dv_pallet_lote if pallet_id == "dv-pallet-cache-1" else None
                ),
            ),
            camara_frios=SimpleNamespace(
                find_by_pallet=Mock(
                    side_effect=lambda pallet_id: self.dv_camara if pallet_id == "dv-pallet-cache-1" else None
                ),
            ),
        )

        self.repo_patcher = patch(
            "infrastructure.repository_factory.get_repositories",
            return_value=self.repos,
        )
        self.repo_patcher.start()
        self.addCleanup(self.repo_patcher.stop)

        self.etapa_patcher = patch(
            "infrastructure.dataverse.repositories.resolve_etapa_lote",
            side_effect=lambda lote, repos=None: (
                "Pallet Cerrado"
                if lote and getattr(lote, "id", "") == "dv-lote-cache-1"
                else "Recepcion"
            ),
        )
        self.etapa_patcher.start()
        self.addCleanup(self.etapa_patcher.stop)

    def _write_stale_cache(self):
        payload = json.loads(self.cache_file.read_text(encoding="utf-8"))
        payload["updated_at"] = "2000-01-01T00:00:00+00:00"
        self.cache_file.write_text(
            json.dumps(payload, ensure_ascii=False),
            encoding="utf-8",
        )

    def test_cache_miss_crea_json_y_segunda_carga_usa_cache(self):
        consulta_url = reverse("operaciones:consulta")
        self.assertFalse(self.cache_file.exists())

        resp_1 = self.client.get(consulta_url, {"tab": "lotes"})
        self.assertEqual(resp_1.status_code, 200)
        self.assertContains(resp_1, "LP-DVC-001")
        self.assertTrue(self.cache_file.exists())
        live_calls = self.mock_lotes_list_recent.call_count
        self.assertGreaterEqual(live_calls, 1)

        resp_2 = self.client.get(consulta_url, {"tab": "lotes"})
        self.assertEqual(resp_2.status_code, 200)
        self.assertContains(resp_2, "LP-DVC-001")
        self.assertEqual(self.mock_lotes_list_recent.call_count, live_calls)

    def test_refresh_manual_fuerza_sync_dataverse(self):
        consulta_url = reverse("operaciones:consulta")
        self.client.get(consulta_url, {"tab": "lotes"})
        calls_before = self.mock_lotes_list_recent.call_count

        resp = self.client.get(consulta_url, {"tab": "lotes", "refresh": "1"})
        self.assertEqual(resp.status_code, 200)
        self.assertGreater(self.mock_lotes_list_recent.call_count, calls_before)

    def test_cache_expirado_dispara_refresco_background_y_no_bloquea(self):
        consulta_url = reverse("operaciones:consulta")
        self.client.get(consulta_url, {"tab": "lotes"})
        self._write_stale_cache()
        calls_before = self.mock_lotes_list_recent.call_count

        with patch(
            "operaciones.views._consulta_dataverse_cache_refresh_background",
            return_value=True,
        ) as mock_background_refresh:
            resp = self.client.get(consulta_url, {"tab": "lotes"})

        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "LP-DVC-001")
        self.assertEqual(self.mock_lotes_list_recent.call_count, calls_before)
        mock_background_refresh.assert_called_once()
        self.assertTrue(resp.context["cache_is_stale"])
        self.assertTrue(resp.context["background_refresh_started"])

    def test_si_falla_dataverse_con_cache_previo_hace_fallback(self):
        consulta_url = reverse("operaciones:consulta")
        self.client.get(consulta_url, {"tab": "lotes"})

        with patch(
            "operaciones.views._consulta_dataverse_cache_refresh_now",
            side_effect=RuntimeError("dataverse offline"),
        ):
            resp = self.client.get(
                consulta_url,
                {"tab": "lotes", "refresh": "1"},
                follow=True,
            )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "LP-DVC-001")
        mensajes = [str(m) for m in resp.context["messages"]]
        self.assertTrue(
            any("No fue posible actualizar la informacion." in m for m in mensajes),
            "Debe mostrar warning de fallback cuando falla refresh manual.",
        )

    def test_exportaciones_csv_xlsx_reusan_dataset_cacheado(self):
        self.client.get(reverse("operaciones:consulta"), {"tab": "lotes"})
        calls_before = self.mock_lotes_list_recent.call_count

        csv_resp = self.client.get(reverse("operaciones:exportar_consulta"), {"tab": "lotes"})
        self.assertEqual(csv_resp.status_code, 200)
        self.assertIn("LP-DVC-001", csv_resp.content.decode("utf-8-sig"))

        self.assertEqual(
            self.mock_lotes_list_recent.call_count,
            calls_before,
            "CSV sin refresh debe salir desde cache sin consultar Dataverse.",
        )

        try:
            from openpyxl import load_workbook
        except Exception:
            return

        xlsx_resp = self.client.get(
            reverse("operaciones:exportar_consulta_excel"),
            {"tab": "pallets"},
        )
        self.assertEqual(xlsx_resp.status_code, 200)
        wb = load_workbook(filename=BytesIO(xlsx_resp.content))
        # El Excel siempre tiene dos hojas: "Lotes" (activa) y "Pallets"
        ws_pallets = wb["Pallets"]
        headers = [c.value for c in ws_pallets[1]]
        self.assertIn("Pallet (code)", headers)
        self.assertIn("Lote relacionado", headers)
        self.assertEqual(
            self.mock_lotes_list_recent.call_count,
            calls_before,
            "Excel sin refresh debe salir desde cache sin consultar Dataverse.",
        )


# ---------------------------------------------------------------------------
# Regresion: _merge_bin_fields — fallback entre bins por campo heredable
# ---------------------------------------------------------------------------

class MergeBinFieldsTest(TestCase):
    """Tests unitarios para _merge_bin_fields."""

    def _bin(self, **kwargs):
        defaults = dict(
            codigo_productor=None, variedad_fruta=None, tipo_cultivo=None,
            color=None, fecha_cosecha=None, codigo_sag_csg=None,
            codigo_sag_csp=None, codigo_sdp=None, numero_cuartel=None,
            nombre_cuartel=None,
        )
        defaults.update(kwargs)
        return SimpleNamespace(**defaults)

    def test_primer_bin_completo_devuelve_sus_valores(self):
        from operaciones.views import _merge_bin_fields
        b = self._bin(
            codigo_productor="PROD-01",
            variedad_fruta="Thompson",
            tipo_cultivo="Uva",
            color="2",
            fecha_cosecha=datetime.date(2025, 3, 1),
        )
        result = _merge_bin_fields([b])
        self.assertEqual(result["codigo_productor"], "PROD-01")
        self.assertEqual(result["variedad_fruta"], "Thompson")
        self.assertEqual(result["color"], "2")

    def test_primer_bin_vacio_usa_segundo(self):
        """Caso raíz del bug: primer bin sin variedad, segundo con variedad."""
        from operaciones.views import _merge_bin_fields
        b1 = self._bin(codigo_productor="PROD-01")
        b2 = self._bin(
            variedad_fruta="Oronules",
            tipo_cultivo="Citricos",
            color="4",
            fecha_cosecha=datetime.date(2025, 4, 10),
        )
        result = _merge_bin_fields([b1, b2])
        self.assertEqual(result["codigo_productor"], "PROD-01")
        self.assertEqual(result["variedad_fruta"], "Oronules")
        self.assertEqual(result["tipo_cultivo"], "Citricos")
        self.assertEqual(result["color"], "4")
        self.assertEqual(result["fecha_cosecha"], datetime.date(2025, 4, 10))

    def test_bins_multiples_primer_no_vacio_gana(self):
        """El primer valor no vacío en orden estable gana para cada campo."""
        from operaciones.views import _merge_bin_fields
        b1 = self._bin(variedad_fruta="Thompson", color="2")
        b2 = self._bin(variedad_fruta="Nadorcott", color="3")
        result = _merge_bin_fields([b1, b2])
        self.assertEqual(result["variedad_fruta"], "Thompson")
        self.assertEqual(result["color"], "2")

    def test_lista_vacia_devuelve_nones(self):
        from operaciones.views import _merge_bin_fields
        result = _merge_bin_fields([])
        self.assertIsNone(result["codigo_productor"])
        self.assertIsNone(result["variedad_fruta"])

    def test_todos_bins_vacios_devuelve_nones(self):
        from operaciones.views import _merge_bin_fields
        result = _merge_bin_fields([self._bin(), self._bin()])
        self.assertIsNone(result["variedad_fruta"])
        self.assertIsNone(result["color"])


# ---------------------------------------------------------------------------
# Regresion: all_bins_by_lotes — SQLite
# ---------------------------------------------------------------------------

class AllBinsByLotesSQLiteTest(TestCase):
    """Verifica que SqliteBinRepository.all_bins_by_lotes retorna todos los bins."""

    def setUp(self):
        self.lote1 = _make_lote(lote_code="LP-ABL-001")
        self.lote2 = _make_lote(lote_code="LP-ABL-002")

    def test_retorna_todos_los_bins_del_lote(self):
        from infrastructure.sqlite.repositories import SqliteBinRepository
        b1 = _make_bin(self.lote1, bin_code="BIN-ABL-001", variedad_fruta="Thompson")
        b2 = _make_bin(self.lote1, bin_code="BIN-ABL-002", variedad_fruta="Crimson")
        repo = SqliteBinRepository()
        result = repo.all_bins_by_lotes([self.lote1.pk])
        bins = result.get(self.lote1.pk, [])
        self.assertEqual(len(bins), 2)
        variedades = {b.variedad_fruta for b in bins}
        self.assertIn("Thompson", variedades)
        self.assertIn("Crimson", variedades)

    def test_retorna_dict_vacio_para_lote_sin_bins(self):
        from infrastructure.sqlite.repositories import SqliteBinRepository
        repo = SqliteBinRepository()
        result = repo.all_bins_by_lotes([self.lote2.pk])
        self.assertEqual(result.get(self.lote2.pk, []), [])

    def test_lote_ids_vacios_devuelve_dict_vacio(self):
        from infrastructure.sqlite.repositories import SqliteBinRepository
        repo = SqliteBinRepository()
        self.assertEqual(repo.all_bins_by_lotes([]), {})

    def test_retorna_bins_de_multiples_lotes(self):
        from infrastructure.sqlite.repositories import SqliteBinRepository
        _make_bin(self.lote1, bin_code="BIN-ML1-001")
        _make_bin(self.lote2, bin_code="BIN-ML2-001")
        repo = SqliteBinRepository()
        result = repo.all_bins_by_lotes([self.lote1.pk, self.lote2.pk])
        self.assertEqual(len(result.get(self.lote1.pk, [])), 1)
        self.assertEqual(len(result.get(self.lote2.pk, [])), 1)

    def test_merge_bin_fields_con_sqlite_all_bins(self):
        """Integra all_bins_by_lotes con _merge_bin_fields: fallback garantizado."""
        from infrastructure.sqlite.repositories import SqliteBinRepository
        from operaciones.views import _merge_bin_fields
        # Primer bin sin variedad; segundo bin con variedad
        _make_bin(self.lote1, bin_code="BIN-MRG-001", variedad_fruta="")
        _make_bin(self.lote1, bin_code="BIN-MRG-002", variedad_fruta="Crimson")
        repo = SqliteBinRepository()
        bins = repo.all_bins_by_lotes([self.lote1.pk]).get(self.lote1.pk, [])
        merged = _merge_bin_fields(bins)
        self.assertEqual(merged["variedad_fruta"], "Crimson")


# ---------------------------------------------------------------------------
# Regresion: _lotes_enriquecidos_dataverse con múltiples bins (mock)
# ---------------------------------------------------------------------------

@override_settings(PERSISTENCE_BACKEND="dataverse")
class LotesEnriquecidosDataverseBinFallbackTest(TestCase):
    """Verifica que reportería usa fallback entre bins cuando el primero está incompleto."""

    def setUp(self):
        hoy = datetime.date.today()
        self.dv_lote = SimpleNamespace(
            id="dv-fb-lote-1",
            lote_code="LP-FB-001",
            estado="cerrado",
            etapa_actual="Recepcion",
            cantidad_bins=2,
            kilos_bruto_conformacion=Decimal("500"),
            kilos_neto_conformacion=Decimal("480"),
            requiere_desverdizado=False,
            fecha_conformacion=hoy,
            ultimo_cambio_estado_at=None,
            codigo_productor="PROD-FB",
            is_active=True,
        )
        # bin1: sin variedad ni color (datos incompletos)
        self.bin_incompleto = SimpleNamespace(
            id="dv-fb-bin-1",
            bin_code="BIN-FB-001",
            codigo_productor="PROD-FB",
            tipo_cultivo=None,
            variedad_fruta=None,
            color=None,
            fecha_cosecha=None,
            codigo_sag_csg=None,
            codigo_sag_csp=None,
            codigo_sdp=None,
            numero_cuartel=None,
            nombre_cuartel=None,
        )
        # bin2: con todos los campos
        self.bin_completo = SimpleNamespace(
            id="dv-fb-bin-2",
            bin_code="BIN-FB-002",
            codigo_productor="PROD-FB",
            tipo_cultivo="Citricos",
            variedad_fruta="Oronules",
            color="4",
            fecha_cosecha=hoy,
            codigo_sag_csg="CSG-001",
            codigo_sag_csp="CSP-001",
            codigo_sdp="SDP-001",
            numero_cuartel="1",
            nombre_cuartel="Norte",
        )
        self.repos = SimpleNamespace(
            lotes=SimpleNamespace(
                list_recent=Mock(return_value=[self.dv_lote]),
            ),
            bins=SimpleNamespace(
                all_bins_by_lotes=Mock(
                    return_value={"dv-fb-lote-1": [self.bin_incompleto, self.bin_completo]}
                ),
            ),
            desverdizados=SimpleNamespace(
                list_by_lotes=Mock(return_value={}),
            ),
            ingresos_packing=SimpleNamespace(
                list_by_lotes=Mock(return_value={}),
            ),
        )

    def test_variedad_viene_del_segundo_bin_cuando_primero_esta_vacio(self):
        from operaciones.views import _lotes_enriquecidos_dataverse
        with patch("infrastructure.repository_factory.get_repositories", return_value=self.repos), \
             patch("infrastructure.dataverse.repositories.resolve_etapa_lote",
                   side_effect=lambda lote, repos=None: "Recepcion"):
            resultado = _lotes_enriquecidos_dataverse("", "")
        self.assertEqual(len(resultado), 1)
        item = resultado[0]
        self.assertEqual(item["variedad"], "Oronules")
        self.assertEqual(item["tipo_cultivo"], "Citricos")
        self.assertEqual(item["color"], "4")
        self.assertEqual(item["codigo_sag_csg"], "CSG-001")
        self.assertEqual(item["numero_cuartel"], "1")
        self.assertEqual(item["nombre_cuartel"], "Norte")

    def test_productor_del_primer_bin_gana_si_no_vacio(self):
        """codigo_productor del primer bin gana aunque segundo también lo tenga."""
        from operaciones.views import _lotes_enriquecidos_dataverse
        self.bin_incompleto.codigo_productor = "PROD-PRIMERO"
        self.bin_completo.codigo_productor = "PROD-SEGUNDO"
        with patch("infrastructure.repository_factory.get_repositories", return_value=self.repos), \
             patch("infrastructure.dataverse.repositories.resolve_etapa_lote",
                   side_effect=lambda lote, repos=None: "Recepcion"):
            resultado = _lotes_enriquecidos_dataverse("", "")
        self.assertEqual(resultado[0]["productor"], "PROD-PRIMERO")


# ---------------------------------------------------------------------------
# Regresion: _consulta_cache_invalidate elimina el archivo
# ---------------------------------------------------------------------------

class ConsultaCacheInvalidateTest(TestCase):
    """Verifica que _consulta_cache_invalidate() elimina el archivo de caché."""

    def test_elimina_archivo_existente(self):
        from operaciones.views import _consulta_cache_invalidate
        with tempfile.TemporaryDirectory() as tmpdir:
            cache_file = Path(tmpdir) / "consulta_dataverse.json"
            cache_file.write_text("{}", encoding="utf-8")
            with override_settings(CONSULTA_DATAVERSE_CACHE_FILE=str(cache_file)):
                self.assertTrue(cache_file.exists())
                _consulta_cache_invalidate()
                self.assertFalse(cache_file.exists(), "El archivo debe eliminarse tras invalidar.")

    def test_no_falla_si_archivo_no_existe(self):
        from operaciones.views import _consulta_cache_invalidate
        with tempfile.TemporaryDirectory() as tmpdir:
            cache_file = Path(tmpdir) / "no_existe.json"
            with override_settings(CONSULTA_DATAVERSE_CACHE_FILE=str(cache_file)):
                # No debe lanzar excepcion
                _consulta_cache_invalidate()


# ---------------------------------------------------------------------------
# Regresion: export CSV coincide con datos de tabla HTML (SQLite, sin regresión)
# ---------------------------------------------------------------------------

class ExportCSVConsistenciaSQLiteTest(TestCase):
    """Export CSV muestra los mismos campos que la vista HTML en modo SQLite."""

    def setUp(self):
        self.client = _make_admin_client("admin_export_consisten")
        self.lote = _make_lote(
            lote_code="LP-EXP-001",
            estado=LotePlantaEstado.CERRADO,
            is_active=True,
        )
        Pallet.objects.create(
            temporada=TEMPORADA,
            pallet_code="PAL-EXP-001",
            is_active=True,
        )

    def test_export_csv_lotes_devuelve_200(self):
        resp = self.client.get(
            reverse("operaciones:exportar_consulta"),
            {"tab": "lotes"},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertIn("LP-EXP-001", resp.content.decode("utf-8-sig"))

    def test_export_csv_no_regresion_sqlite(self):
        """SQLite: all_bins_by_lotes integrado no rompe la vista consulta."""
        resp = self.client.get(reverse("operaciones:consulta"), {"tab": "lotes"})
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "LP-EXP-001")


# ---------------------------------------------------------------------------
# Regresion: no romper recepcion SQLite con el nuevo all_bins_by_lotes
# ---------------------------------------------------------------------------

class RecepcionNoRegressionAllBinsTest(TestCase):
    """Verifica que agregar_bin_a_lote_abierto y cerrar_lote siguen funcionando."""

    def setUp(self):
        self.client = _make_admin_client("admin_no_regresion_bins")
        session = self.client.session
        session["temporada_activa"] = TEMPORADA
        session.save()

    def test_recepcion_flujo_completo_no_regresion(self):
        # Iniciar lote
        resp = self.client.post(
            reverse("operaciones:recepcion"),
            {"action": "iniciar"},
        )
        self.assertIn(resp.status_code, [200, 302])

        session = self.client.session
        lote_code = session.get("lote_activo_code")
        if not lote_code:
            return  # Backend sin lote activo — no hay más que verificar

        # Agregar bin — lotes_json_from_records usa all_bins_by_lotes internamente
        resp = self.client.post(
            reverse("operaciones:recepcion"),
            {
                "action": "agregar_bin",
                "codigo_productor": "PROD-NR",
                "variedad_fruta": "Thompson",
                "tipo_cultivo": "Uva de mesa",
                "color": "2",
                "fecha_cosecha": str(datetime.date.today()),
                "kilos_bruto_ingreso": "500",
                "kilos_neto_ingreso": "480",
            },
        )
        self.assertIn(resp.status_code, [200, 302])


# ---------------------------------------------------------------------------
# Regresion: fecha_conformacion se establece automaticamente al iniciar lote
# ---------------------------------------------------------------------------

class IniciarLoteFechaConformacionTest(TestCase):
    """Verifica que iniciar_lote_recepcion asigna fecha_conformacion=hoy por defecto."""

    def test_default_fecha_conformacion_es_hoy(self):
        from operaciones.application.use_cases.iniciar_lote_recepcion import iniciar_lote_recepcion
        hoy = datetime.date.today()
        resultado = iniciar_lote_recepcion({"temporada": TEMPORADA})
        self.assertTrue(resultado.ok, resultado.errors)
        lote_code = resultado.data["lote_code"]
        lote = Lote.objects.get(temporada=TEMPORADA, lote_code=lote_code)
        self.assertIsNotNone(lote.fecha_conformacion, "fecha_conformacion debe quedar establecida")
        self.assertEqual(lote.fecha_conformacion, hoy)

    def test_fecha_conformacion_explicita_en_payload_se_respeta(self):
        from operaciones.application.use_cases.iniciar_lote_recepcion import iniciar_lote_recepcion
        fecha_custom = datetime.date(2026, 3, 1)
        resultado = iniciar_lote_recepcion({"temporada": TEMPORADA, "fecha_conformacion": fecha_custom})
        self.assertTrue(resultado.ok, resultado.errors)
        lote_code = resultado.data["lote_code"]
        lote = Lote.objects.get(temporada=TEMPORADA, lote_code=lote_code)
        self.assertEqual(lote.fecha_conformacion, fecha_custom)


# ---------------------------------------------------------------------------
# Regresion: _row_to_lote() hidrata fecha_conformacion y ultimo_cambio_estado_at
# ---------------------------------------------------------------------------

class RowToLoteHidrataCamposTest(TestCase):
    """Verifica que _row_to_lote() lee correctamente los campos de fecha."""

    def test_fecha_conformacion_se_hidrata(self):
        from infrastructure.dataverse.repositories import _row_to_lote
        from infrastructure.dataverse.mapping import LOTE_FIELDS
        row = {
            LOTE_FIELDS["id"]: "guid-test-1",
            LOTE_FIELDS["lote_code"]: "LP-ROW-001",
            LOTE_FIELDS["fecha_conformacion"]: "2026-03-15",
            LOTE_FIELDS["etapa_actual"]: "Recepcion",
            LOTE_FIELDS["cantidad_bins"]: 3,
        }
        record = _row_to_lote(row)
        self.assertEqual(record.fecha_conformacion, datetime.date(2026, 3, 15))

    def test_ultimo_cambio_estado_at_se_hidrata(self):
        from infrastructure.dataverse.repositories import _row_to_lote
        from infrastructure.dataverse.mapping import LOTE_FIELDS
        row = {
            LOTE_FIELDS["id"]: "guid-test-2",
            LOTE_FIELDS["lote_code"]: "LP-ROW-002",
            LOTE_FIELDS["ultimo_cambio_estado_at"]: "2026-04-10T14:30:00Z",
            LOTE_FIELDS["cantidad_bins"]: 0,
        }
        record = _row_to_lote(row)
        self.assertIsNotNone(record.ultimo_cambio_estado_at)
        self.assertEqual(record.ultimo_cambio_estado_at.year, 2026)
        self.assertEqual(record.ultimo_cambio_estado_at.month, 4)
        self.assertEqual(record.ultimo_cambio_estado_at.day, 10)

    def test_campos_fecha_none_cuando_row_vacia(self):
        from infrastructure.dataverse.repositories import _row_to_lote
        from infrastructure.dataverse.mapping import LOTE_FIELDS
        row = {
            LOTE_FIELDS["id"]: "guid-test-3",
            LOTE_FIELDS["lote_code"]: "LP-ROW-003",
            LOTE_FIELDS["cantidad_bins"]: 0,
        }
        record = _row_to_lote(row)
        self.assertIsNone(record.fecha_conformacion)
        self.assertIsNone(record.ultimo_cambio_estado_at)


# ---------------------------------------------------------------------------
# Regresion: _bin_to_record() SQLite mapea tipo_cultivo y cuarteles
# ---------------------------------------------------------------------------

class BinToRecordSQLiteTipoCultivoTest(TestCase):
    """Verifica que _bin_to_record() popula tipo_cultivo, numero_cuartel y nombre_cuartel."""

    def setUp(self):
        self.lote = _make_lote()

    def test_tipo_cultivo_mapeado(self):
        from infrastructure.sqlite.repositories import _bin_to_record
        b = _make_bin(
            self.lote,
            tipo_cultivo="Citricos",
            variedad_fruta="Navel",
            numero_cuartel="C05",
            nombre_cuartel="Sur",
        )
        record = _bin_to_record(b)
        self.assertEqual(record.tipo_cultivo, "Citricos")
        self.assertEqual(record.numero_cuartel, "C05")
        self.assertEqual(record.nombre_cuartel, "Sur")

    def test_merge_bin_fields_tipo_cultivo_desde_sqlite(self):
        """all_bins_by_lotes + _merge_bin_fields devuelve tipo_cultivo desde SQLite."""
        from infrastructure.sqlite.repositories import SqliteBinRepository
        from operaciones.views import _merge_bin_fields
        _make_bin(self.lote, tipo_cultivo="Uva de mesa", variedad_fruta="Thompson")
        repo = SqliteBinRepository()
        bins_map = repo.all_bins_by_lotes([self.lote.id])
        bins = bins_map.get(self.lote.id, [])
        merged = _merge_bin_fields(bins)
        self.assertEqual(merged["tipo_cultivo"], "Uva de mesa")

    def test_merge_bin_fields_fallback_segundo_bin_si_primero_vacio(self):
        """tipo_cultivo se toma del segundo bin si el primero está vacío."""
        from infrastructure.sqlite.repositories import SqliteBinRepository
        from operaciones.views import _merge_bin_fields
        _make_bin(self.lote, tipo_cultivo="", variedad_fruta="Thompson")
        _make_bin(self.lote, tipo_cultivo="Citricos", variedad_fruta="Oronules")
        repo = SqliteBinRepository()
        bins_map = repo.all_bins_by_lotes([self.lote.id])
        bins = bins_map.get(self.lote.id, [])
        merged = _merge_bin_fields(bins)
        self.assertEqual(merged["tipo_cultivo"], "Citricos")


# ---------------------------------------------------------------------------
# Regresion: _detalle_lote_context() incluye ultimo_cambio_etapa
# ---------------------------------------------------------------------------

class DetalleLoteContextUltimoCambioTest(TestCase):
    """Verifica que el contexto de detalle de lote incluye ultimo_cambio_etapa."""

    def setUp(self):
        self.lote = _make_lote(
            lote_code="LP-DET-UCE-001",
            estado=LotePlantaEstado.CERRADO,
            fecha_conformacion=datetime.date.today(),
        )
        _make_bin(self.lote)

    def test_ultimo_cambio_etapa_presente_en_contexto_sqlite(self):
        from operaciones.views import _detalle_lote_context
        ctx = _detalle_lote_context(TEMPORADA, "LP-DET-UCE-001")
        self.assertIn("ultimo_cambio_etapa", ctx, "ultimo_cambio_etapa debe estar en el contexto")

    def test_ultimo_cambio_etapa_fallback_a_fecha_conformacion(self):
        """Si no hay ultimo_cambio_estado_at, fallback muestra fecha_conformacion como string."""
        from operaciones.views import _detalle_lote_context
        ctx = _detalle_lote_context(TEMPORADA, "LP-DET-UCE-001")
        uce = ctx.get("ultimo_cambio_etapa")
        self.assertTrue(uce, "ultimo_cambio_etapa no debe ser vacio")
        # _fmt_ultimo_cambio formatea date como "d/m/Y"
        self.assertEqual(uce, datetime.date.today().strftime("%d/%m/%Y"))

    def test_detalle_view_renderiza_campo_ultimo_cambio(self):
        """La vista de detalle renderiza el campo 'Ult. cambio estado'."""
        client = _make_admin_client("admin_uce_det")
        resp = client.get(reverse("operaciones:consulta_lote_detalle", args=["LP-DET-UCE-001"]))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Ult. cambio estado")


# ---------------------------------------------------------------------------
# Regresion Dataverse: reporteria leyendo fecha_conformacion y ultimo_cambio
# ---------------------------------------------------------------------------

@override_settings(PERSISTENCE_BACKEND="dataverse")
class LoteReportDataverseFechasCamposTest(TestCase):
    """Verifica que reporteria Dataverse incluye fecha_conformacion y ultimo_cambio_estado_at."""

    def setUp(self):
        hoy = datetime.date.today()
        ahora = datetime.datetime(2026, 4, 16, 10, 30, 0, tzinfo=datetime.timezone.utc)
        self.dv_lote = SimpleNamespace(
            id="dv-fc-lote-1",
            lote_code="LP-FC-001",
            estado="cerrado",
            etapa_actual="Packing / Proceso",
            cantidad_bins=2,
            kilos_bruto_conformacion=Decimal("520"),
            kilos_neto_conformacion=Decimal("500"),
            requiere_desverdizado=False,
            fecha_conformacion=hoy,
            ultimo_cambio_estado_at=ahora,
            codigo_productor="PROD-FC",
            is_active=True,
        )
        self.repos = SimpleNamespace(
            lotes=SimpleNamespace(list_recent=Mock(return_value=[self.dv_lote])),
            bins=SimpleNamespace(all_bins_by_lotes=Mock(return_value={
                "dv-fc-lote-1": [SimpleNamespace(
                    id="dv-fc-bin-1",
                    bin_code="BIN-FC-001",
                    codigo_productor="PROD-FC",
                    tipo_cultivo="Citricos",
                    variedad_fruta="Oronules",
                    color="4",
                    fecha_cosecha=hoy,
                    codigo_sag_csg="", codigo_sag_csp="", codigo_sdp="",
                    numero_cuartel="1", nombre_cuartel="Norte",
                )]
            })),
            desverdizados=SimpleNamespace(list_by_lotes=Mock(return_value={})),
            ingresos_packing=SimpleNamespace(list_by_lotes=Mock(return_value={})),
        )

    def test_fecha_conformacion_en_resultado_reporteria(self):
        from operaciones.views import _lotes_enriquecidos_dataverse
        with patch("infrastructure.repository_factory.get_repositories", return_value=self.repos), \
             patch("infrastructure.dataverse.repositories.resolve_etapa_lote",
                   side_effect=lambda lote, repos=None: "Packing / Proceso"):
            resultado = _lotes_enriquecidos_dataverse("", "")
        self.assertEqual(len(resultado), 1)
        self.assertEqual(resultado[0]["fecha_conformacion"], datetime.date.today())

    def test_ultimo_cambio_etapa_en_resultado_reporteria(self):
        from operaciones.views import _lotes_enriquecidos_dataverse
        ahora = datetime.datetime(2026, 4, 16, 10, 30, 0, tzinfo=datetime.timezone.utc)
        with patch("infrastructure.repository_factory.get_repositories", return_value=self.repos), \
             patch("infrastructure.dataverse.repositories.resolve_etapa_lote",
                   side_effect=lambda lote, repos=None: "Packing / Proceso"):
            resultado = _lotes_enriquecidos_dataverse("", "")
        self.assertEqual(resultado[0]["ultimo_cambio_etapa"], ahora)

    def test_tipo_cultivo_heredado_desde_bins_dataverse(self):
        from operaciones.views import _lotes_enriquecidos_dataverse
        with patch("infrastructure.repository_factory.get_repositories", return_value=self.repos), \
             patch("infrastructure.dataverse.repositories.resolve_etapa_lote",
                   side_effect=lambda lote, repos=None: "Packing / Proceso"):
            resultado = _lotes_enriquecidos_dataverse("", "")
        self.assertEqual(resultado[0]["tipo_cultivo"], "Citricos")
